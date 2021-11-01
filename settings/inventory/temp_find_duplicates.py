"""
Find duplicates and produce excel output from inventory_combined_raw.xlsx
"""
from openpyxl import load_workbook


def get_combined_data(worksheet_combined: load_workbook) -> list:
    """
    Combined data for later manipulation.
    :return:
    """
    combined_data: list = []

    count: int = 2
    while count < 61_961:
        serial_number: str = worksheet_combined[f"A{count}"].value
        # print(f"Get Row {count} - {serial_number}")

        if serial_number is not None:
            part_number: str = worksheet_combined[f"B{count}"].value
            section: str = worksheet_combined[f"C{count}"].value
            notes: str = worksheet_combined[f"D{count}"].value

            row_data: tuple = (serial_number, part_number, section, notes)
            combined_data.append(row_data)
            count += 1

    return combined_data


def find_duplicates(combined_data: list) -> list:
    """
    Find duplicates after getting combined data.
    :param combined_data:
    :return:
    """
    duplicates: list = []
    all_box: list = []
    for index, row_data in enumerate(combined_data, start=1):
        serial_number: str = row_data[0]
        # print(f"Find Row - {index} - {serial_number}")

        if serial_number in all_box:
            duplicates.append(row_data)
        elif serial_number not in all_box:
            all_box.append(serial_number)

    return duplicates


def output_duplicates(duplicates: list, worksheet_duplicates: load_workbook) -> None:
    """
    Output into excel.
    :param worksheet_duplicates: load_workbook
    :param duplicates:
    :return:
    """
    for index, entry in enumerate(duplicates, start=2):
        serial_number: str = entry[0]
        part_number: str = entry[1]
        section: str = entry[2]
        notes: str = entry[3]

        worksheet_duplicates[f"A{index}"].value = serial_number
        worksheet_duplicates[f"B{index}"].value = part_number
        worksheet_duplicates[f"C{index}"].value = section
        worksheet_duplicates[f"D{index}"].value = notes


def get_recurring_duplicates(duplicates: list) -> dict:
    """

    :param duplicates:
    :return:
    """
    recurring: dict = {}
    for entry in duplicates:
        serial_number: str = entry[0]
        part_number: str = entry[1]
        section: str = entry[2]
        notes: str = entry[3]
        print(serial_number)

        if serial_number not in recurring:
            recurring[serial_number]: dict = {}
            recurring[serial_number]["count"]: int = 1
            recurring[serial_number]["part_numbers"]: list = [part_number]
            recurring[serial_number]["sections"]: list = [section]
            recurring[serial_number]["notes"]: list = [notes]

        elif serial_number in recurring:
            recurring[serial_number]["count"] += 1
            recurring[serial_number]["part_numbers"].append(part_number)
            recurring[serial_number]["sections"].append(section)
            recurring[serial_number]["notes"].append(notes)

    return recurring


def output_recurring(worksheet_recurring, recurring) -> None:
    """

    :param worksheet_recurring:
    :param recurring:
    :return:
    """
    for index, serial_number in enumerate(recurring, start=2):
        count: int = recurring[serial_number]["count"]
        part_numbers: list = recurring[serial_number]["part_numbers"]
        sections: list = recurring[serial_number]["sections"]
        notes: list = recurring[serial_number]["notes"]

        worksheet_recurring[f"A{index}"].value = serial_number
        worksheet_recurring[f"B{index}"].value = count
        worksheet_recurring[f"C{index}"].value = ", ".join(part_numbers)
        worksheet_recurring[f"D{index}"].value = ", ".join(sections)
        worksheet_recurring[f"E{index}"].value = ", ".join(notes)


def get_difference_serials(worksheet_difference):
    original: list = []
    # non_duplicates
    pass


def main_method() -> None:
    """

    :return:
    """
    file_name: str = "inventory_combined_raw.xlsx"

    workbook = load_workbook(file_name)
    worksheet_combined: load_workbook = workbook["Combined"]
    worksheet_duplicates: load_workbook = workbook["Duplicates"]
    worksheet_recurring: load_workbook = workbook["Recurring - Duplicates"]
    # worksheet_difference: load_workbook = workbook["Difference"]

    combined_data: list = get_combined_data(worksheet_combined)
    duplicates: list = find_duplicates(combined_data)
    output_duplicates(duplicates, worksheet_duplicates)
    recurring: dict = get_recurring_duplicates(duplicates)
    output_recurring(worksheet_recurring, recurring)

    workbook.save("inventory_combined_duplicates.xlsx")


main_method()



