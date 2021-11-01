"""
Temporary solution to store quarantine serial numbers inside database.

8/31/2021
"""
from openpyxl import load_workbook
from pipe_cleaner.src.log_database import access_database_document
from datetime import datetime


def get_quarantine_data() -> dict:
    """
    Get quarantine data including part, serial, and supplier.
    :return: serial -> part / supplier
    """
    worksheet = get_quarantine_worksheet()

    quarantine: dict = {"data": {},
                        "stats": {"before": 0,
                                  "after": 0}}

    for row_number in range(2, worksheet.max_row):
        if worksheet[f"A{row_number}"].value is None:
            break

        quarantine_data: dict = quarantine["data"]

        serial_number = str(worksheet[f"B{row_number}"].value).replace(" ", "")
        part_number = str(worksheet[f"A{row_number}"].value).replace(" ", "")
        supplier = str(worksheet[f"C{row_number}"].value).replace(" ", "")

        clean_serial: str = clean_serial_number(serial_number, part_number)
        clean_part: str = clean_part_number(part_number)
        clean_supplier: str = clean_supplier_name(supplier)

        if clean_serial not in quarantine_data:
            current_serial = quarantine_data[clean_serial] = {}

            current_serial["part"]: str = clean_part
            current_serial["supplier"]: str = clean_supplier

            quarantine["stats"]["after"] += 1

        quarantine["stats"]["before"] += 1

    return quarantine["data"]


def get_quarantine_worksheet():
    workbook = load_workbook("quarantine_serials.xlsx")
    worksheet = workbook["Quarantine"]
    return worksheet


def get_inventory_transactions() -> dict:
    """
    This is not from the database but the transactions provided from Rich's data.
    :return: serial -> part / supplier
    """
    workbook = load_workbook("kirkland_inventory.xlsx")
    worksheet = workbook["transactions"]

    transaction_data: dict = {}

    for row_number in range(2, worksheet.max_row):
        if worksheet[f"A{row_number}"].value is None:
            break

        part_number = str(worksheet[f"A{row_number}"].value).replace(" ", "")
        serial_number = str(worksheet[f"B{row_number}"].value).replace(" ", "")
        action_needed = str(worksheet[f"C{row_number}"].value).replace(" ", ""). \
            replace("NeedsTransacted", "").upper()

        clean_serial: str = clean_serial_number(serial_number, part_number)

        if len(serial_number) != len(clean_serial):
            print(f"serial_number: {serial_number}")
            print(f"part_number: {part_number}")

        transaction_data[serial_number]: dict = {}
        transaction_data[serial_number]["part_number"]: str = part_number
        transaction_data[serial_number]["action_needed"]: str = action_needed

    return transaction_data


def clean_data(field: str) -> str:
    """
    Clean data so that string output is consistent.
    :param field:
    :return:
    """
    return field.title().replace("Hgst", "HGST").replace("Skhynix", "SK Hynix"). \
        replace("Incage", "Cage").replace("Inserver", "Rack").replace("Westerndigitial", "Western Digital"). \
        replace("HGST/WesternDigital", "HGST / Western Digital"). \
        replace("Aspen/Westerndigital", "Aspen / Western Digital")


def clean_serial_number(serial_number: str, part_number) -> str:
    """
    Clean serial number based on delimiter and part number.
    :param serial_number:
    :param part_number:
    :return:
    """
    if serial_number == part_number and len(serial_number) == len(part_number):
        return serial_number

    return serial_number. \
        replace(f"_{part_number}", ""). \
        replace(f" {part_number}", ""). \
        replace(f"-{part_number}", ""). \
        replace(f".{part_number}", ""). \
        replace(f"+{part_number}", ""). \
        replace(f"{part_number}_", ""). \
        replace(f"{part_number} ", ""). \
        replace(f"{part_number}-", ""). \
        replace(f"{part_number}.", ""). \
        replace(f"{part_number}+", ""). \
        replace(f"{part_number}", "").upper().replace(" ", "")


def clean_supplier_name(supplier_name: str) -> str:
    """

    :param supplier_name:
    :return:
    """
    return supplier_name. \
        replace("WesternDigital", "Western Digital"). \
        replace("SKHynix", "SK Hynix"). \
        replace("Aspen/WesternDigital", "Aspen / Western Digital"). \
        replace("Toshiba/Kioxia", "Toshiba / Kioxia"). \
        replace("SEAGATE", "Seagate"). \
        replace("TOSHIBA", "Toshiba"). \
        replace("Aspen/Western Digital", "Aspen / Western Digital").\
        replace("HGST/Western Digital", "HGST / Western Digital")


def clean_location_name(location_name: str) -> str:
    """

    :param location_name:
    :return:
    """
    return location_name. \
        replace("InCage", "Cage"). \
        replace("InServer", "Rack"). \
        replace("In Cage", "Cage"). \
        replace("In Server", "Rack")


def clean_part_number(part_number: str) -> str:
    """

    :param part_number:
    :return:
    """
    return part_number.upper().replace(" ", "")


def update_inventory_data(fields: dict) -> dict:
    """
    Update inventory data with new inventory logs.
    :param fields: data structure containing
    :return: update inventory data
    """
    serial_data: dict = {}
    current_original = serial_data["original"] = {}
    current_clean = serial_data["clean"] = {}
    current_logs = serial_data["logs"] = {}

    clean_serial: str = fields["clean"]["serial"]
    clean_part: str = fields["clean"]["part"]
    clean_supplier: str = fields["clean"]["supplier"]
    clean_location: str = fields["clean"]["location"]

    current_original["serial"]: str = fields["original"]["serial"]
    current_original["part"]: str = fields["original"]["part"]
    current_original["supplier"]: str = fields["original"]["supplier"]
    current_original["location"]: str = fields["original"]["location"]

    current_clean["serial"]: str = clean_serial
    current_clean["part"]: str = clean_part
    current_clean["supplier"]: str = clean_supplier
    current_clean["location"]: str = clean_location

    current_logs["serials"]: list = [clean_serial]
    current_logs["parts"]: list = [clean_part]
    current_logs["suppliers"]: list = [clean_supplier]
    current_logs["locations"]: list = [clean_location]

    return serial_data


def update_combined_data(fields: dict) -> dict:
    """
    Update inventory data with new inventory logs.
    :param fields: data structure containing
    :return: update inventory data
    """
    serial_data: dict = {}
    current_original = serial_data["original"] = {}
    current_clean = serial_data["clean"] = {}
    current_logs = serial_data["logs"] = {}

    clean_serial: str = fields["clean"]["serial"]
    clean_part: str = fields["clean"]["part"]
    clean_supplier: str = fields["clean"]["supplier"]
    clean_location: str = fields["clean"]["location"]
    clean_section: str = fields["clean"]["section"]
    clean_note: str = fields["clean"]["note"]

    current_original["serial"]: str = fields["original"]["serial"]
    current_original["part"]: str = fields["original"]["part"]
    current_original["supplier"]: str = fields["original"]["supplier"]
    current_original["location"]: str = fields["original"]["location"]
    current_original["section"]: str = fields["original"]["section"]
    current_original["note"]: str = fields["original"]["note"]

    current_clean["serial"]: str = clean_serial
    current_clean["part"]: str = clean_part
    current_clean["supplier"]: str = clean_supplier
    current_clean["section"]: str = clean_section
    current_clean["notes"]: str = fields["original"]["note"]

    current_logs["serials"]: list = [clean_serial]
    current_logs["parts"]: list = [clean_part]
    current_logs["suppliers"]: list = [clean_supplier]
    current_logs["sections"]: list = [clean_section]
    current_logs["notes"]: list = [clean_note]

    if clean_note == "IN":
        current_clean["location"]: str = "Cage"
        current_logs["locations"]: list = ["Cage"]

    elif clean_note == "OUT":
        current_clean["location"]: str = "Out"
        current_logs["locations"]: list = ["Out"]

    else:
        current_clean["location"]: str = clean_location
        current_logs["locations"]: list = [clean_location]

    return serial_data


def update_inventory_commodities(current_serial: dict, fields: dict) -> dict:
    """

    :param current_serial:
    :param fields:
    :return:
    """
    clean_serial: str = fields["clean"]["serial"]
    clean_part: str = fields["clean"]["part"]
    clean_supplier: str = fields["clean"]["supplier"]
    clean_location: str = fields["clean"]["location"]
    clean_section: str = fields["clean"]["section"]
    clean_note: str = fields["clean"]["note"]

    current_serial["logs"]["serials"].append(clean_serial)
    current_serial["logs"]["parts"].append(clean_part)
    current_serial["logs"]["suppliers"].append(clean_supplier)
    current_serial["logs"]["locations"].append(clean_location)
    current_serial["logs"]["sections"].append(clean_section)
    current_serial["logs"]["notes"].append(clean_note)

    if clean_note == "IN":
        current_serial["logs"]["locations"].append("Cage")

    elif clean_note == "OUT":
        current_serial["logs"]["locations"].append("Out")

    else:
        current_serial["logs"]["locations"].append(clean_location)

    return current_serial


# def get_kirkland_inventory() -> dict:
#     """
#     This is not from the database but the Kirkland inventory provided from Rich's data.
#     :return: serial -> part / supplier
#     """
#     worksheet: load_workbook = get_inventory_worksheet()
#
#     inventory: dict = setup_inventory_dict()
#
#     for row_number in range(2, worksheet.max_row):
#         if worksheet[f"A{row_number}"].value is None:
#             break
#
#         inventory_data: dict = inventory["data"]
#         inventory_stats: dict = inventory["stats"]
#         fields: dict = get_inventory_fields(row_number, worksheet)
#
#         original_serial: str = fields["original"]["serial"]
#         clean_serial: str = fields["clean"]["serial"]
#
#         if clean_serial not in inventory_data:
#             inventory_data[clean_serial]: dict = {}
#             inventory_data[clean_serial]: dict = update_inventory_data(fields)
#
#             inventory_stats["after"] += 1
#
#         elif clean_serial in inventory_data:
#             current_serial: dict = inventory_data[clean_serial]
#             inventory_data[clean_serial]: dict = update_inventory_commodities(current_serial, fields)
#
#             inventory_stats["duplicates"] += 1
#
#         if len(original_serial) != len(clean_serial):
#             inventory_stats["doubles"] += 1
#
#         unique_part_numbers = list(set(inventory_data[clean_serial]["logs"]["parts"]))
#         if len(unique_part_numbers) >= 2:
#             print(f"Serial Number: {clean_serial}")
#             print(f"Parts:         {', '.join(inventory_data[clean_serial]['logs']['parts'])}")
#             print(f"Suppliers:     {', '.join(inventory_data[clean_serial]['logs']['suppliers'])}")
#             print(f"Locations:     {', '.join(inventory_data[clean_serial]['logs']['locations'])}\n")
#             inventory_stats["different_part_numbers"] += 1
#
#         inventory_stats["before"] += 1
#
#     import json
#     foo = json.dumps(inventory["stats"], sort_keys=True, indent=4)
#     print(foo)
#     input()
#
#     return inventory["data"]


def get_inventory_worksheet():
    workbook = load_workbook("kirkland_inventory.xlsx")
    return workbook["inventory"]


def setup_inventory_dict() -> dict:
    """
    Structure inventory data for future analysis and comparison.
    :return:
    """
    return {"data": {},
            "stats": {"before": 0,
                      "after": 0,
                      "duplicates": 0,
                      "part_in_serial": 0,
                      "different_part_numbers": 0}}


def get_inventory_fields(row_number: int, worksheet: load_workbook) -> dict:
    """
    Get Rich's data given the fields.
    :param row_number:
    :param worksheet:
    :return:
    """
    serial = str(worksheet[f"A{row_number}"].value)
    part = str(worksheet[f"B{row_number}"].value)
    supplier = str(worksheet[f"C{row_number}"].value)
    location = str(worksheet[f"D{row_number}"].value)

    clean_serial: str = clean_serial_number(serial, part)
    clean_part: str = clean_part_number(part)
    clean_supplier: str = clean_supplier_name(supplier)
    clean_location: str = clean_location_name(location)

    return {"original": {"serial": serial,
                         "part": part,
                         "supplier": supplier,
                         "location": location},
            "clean": {"serial": clean_serial,
                      "part": clean_part,
                      "supplier": clean_supplier,
                      "location": clean_location}}


def get_clean_note(note: str) -> str:
    """

    :param note:
    :return:
    """
    return note.replace(" ", "").replace("NeedsTransacted", "").upper().strip()


def setup_combined_data(row_number: int, worksheet: load_workbook) -> dict:
    """
    Get Rich's data given the fields.
    :param row_number:
    :param worksheet:
    :return:
    """
    serial = str(worksheet[f"A{row_number}"].value)
    part = str(worksheet[f"B{row_number}"].value)
    location = str(worksheet[f"C{row_number}"].value)
    supplier = str(worksheet[f"D{row_number}"].value)
    section = str(worksheet[f"E{row_number}"].value)
    note = str(worksheet[f"F{row_number}"].value)

    clean_serial: str = clean_serial_number(serial, part)
    clean_part: str = clean_part_number(part)
    clean_supplier: str = clean_supplier_name(supplier)
    clean_location: str = clean_location_name(location)
    clean_section: str = section
    clean_note: str = get_clean_note(note)

    return {"original": {"serial": serial,
                         "part": part,
                         "supplier": supplier,
                         "location": location,
                         "section": section,
                         "note": note},

            "clean": {"serial": clean_serial,
                      "part": clean_part,
                      "supplier": clean_supplier,
                      "location": clean_location,
                      "section": clean_section,
                      "note": clean_note}}


def get_month_name(raw_month: str) -> str:
    """

    :param raw_month:
    :return:
    """
    if raw_month == "01":
        return "January"

    elif raw_month == "02":
        return "February"

    elif raw_month == "03":
        return "March"

    elif raw_month == "04":
        return "April"

    elif raw_month == "05":
        return "May"

    elif raw_month == "06":
        return "June"

    elif raw_month == "07":
        return "July"

    elif raw_month == "08":
        return "August"

    elif raw_month == "09":
        return "September"

    elif raw_month == "10":
        return "October"

    elif raw_month == "11":
        return "November"

    elif raw_month == "12":
        return "December"

    else:
        return "None"


def get_current_date() -> str:
    raw_date: str = datetime.now().strftime('%Y-%m-%d')
    raw_year: str = raw_date[0:4]
    raw_month: str = raw_date[5:7]
    raw_day: str = raw_date[8:10]

    return f"{raw_month}/{raw_day}/{raw_year}"


def get_serial_numbers_database():
    """
    Get serial numbers for all records of Kirkland.
    :return:
    """
    return access_database_document("serial_numbers", 'version_02')


def setup_serial_number() -> dict:
    """
    Structure of serial number per transaction.
    :return: serial_number_id
    """
    return {"part_numbers": [],
            "from_locations": [],
            "to_locations": [],
            "scanners": [],
            "requesters": [],
            "times": [],
            "dates": [],
            "suppliers": [],
            "notes": [],
            "tool_versions": [],
            "in_house": [],
            "_id": "None",
            "specific_locations": {"to_pipes": [],
                                   "from_pipes": [],
                                   "to_machines": [],
                                   "from_machines": [],
                                   "to_outside": [],
                                   "from_outside": [],
                                   "to_cage": [],
                                   "from_cage": []}}


def store_transactions_in_database(transactions_data: dict) -> None:
    """
    Store transactions data from Rich inside database.
    :return:
    """
    serials_database = get_serial_numbers_database()

    current_date: str = get_current_date()
    current_time: str = datetime.today().strftime("%I:%M %p")

    for index, entry_serial_number in enumerate(transactions_data, start=1):
        print(f"Transactions ({index}): {entry_serial_number}")

        serial_data: dict = transactions_data[entry_serial_number]

        entry_part_number: str = serial_data["part_number"]
        entry_action_needed: str = serial_data["action_needed"]

        serial_number_data = serials_database.find_one({"_id": entry_serial_number})

        if serial_number_data is None:

            if entry_action_needed == "OUT":
                setup_serial: dict = setup_serial_number()

                setup_serial["_id"]: str = entry_serial_number
                setup_serial["in_house"].append("False")
                setup_serial["part_numbers"].append(entry_part_number)
                setup_serial["from_locations"].append("Unknown")
                setup_serial["to_locations"].append("Unknown")
                setup_serial["scanners"].append("Unknown")
                setup_serial["requesters"].append("Unknown")
                setup_serial["dates"].append(current_date)
                setup_serial["times"].append(current_time)
                setup_serial["suppliers"].append("Unknown")
                setup_serial["tool_versions"].append("2.7.3")
                setup_serial["notes"].append("From Rich")

                serials_database.insert_one(setup_serial)

            elif entry_action_needed == "IN":
                setup_serial: dict = setup_serial_number()

                setup_serial["_id"]: str = entry_serial_number
                setup_serial["in_house"].append("True")
                setup_serial["part_numbers"].append(entry_part_number)
                setup_serial["from_locations"].append("Unknown")
                setup_serial["to_locations"].append("Unknown")
                setup_serial["scanners"].append("Unknown")
                setup_serial["requesters"].append("Unknown")
                setup_serial["dates"].append(current_date)
                setup_serial["times"].append(current_time)
                setup_serial["suppliers"].append("Unknown")
                setup_serial["tool_versions"].append("2.7.3")
                setup_serial["notes"].append("From Rich")

                serials_database.insert_one(setup_serial)

        else:
            if entry_action_needed == "OUT":
                serial_number_data["in_house"].append("False")
                serial_number_data["part_numbers"].append(entry_part_number)
                serial_number_data["from_locations"].append("Unknown")
                serial_number_data["to_locations"].append("Unknown")
                serial_number_data["scanners"].append("Unknown")
                serial_number_data["requesters"].append("Unknown")
                serial_number_data["dates"].append(current_date)
                serial_number_data["times"].append(current_time)
                serial_number_data["suppliers"].append("Unknown")
                serial_number_data["tool_versions"].append("2.7.3")
                serial_number_data["notes"].append("From Rich")

                serials_database.update_one({"_id": entry_serial_number},
                                            {"$set": serial_number_data},
                                            upsert=True)

            elif entry_action_needed == "IN":
                serial_number_data["in_house"].append("True")
                serial_number_data["part_numbers"].append(entry_part_number)
                serial_number_data["from_locations"].append("Unknown")
                serial_number_data["to_locations"].append("Unknown")
                serial_number_data["scanners"].append("Unknown")
                serial_number_data["requesters"].append("Unknown")
                serial_number_data["dates"].append(current_date)
                serial_number_data["times"].append(current_time)
                serial_number_data["suppliers"].append("Unknown")
                serial_number_data["tool_versions"].append("2.7.3")
                serial_number_data["notes"].append("From Rich")

                serials_database.update_one({"_id": entry_serial_number},
                                            {"$set": serial_number_data},
                                            upsert=True)


def store_quarantine_in_database(quarantine_data: dict) -> None:
    """
    Store quarantine data inside database.
    :return:
    """
    serials_database = get_serial_numbers_database()

    current_date: str = get_current_date()
    current_time: str = datetime.today().strftime("%I:%M %p")

    for index, entry_serial_number in enumerate(quarantine_data, start=1):
        print(f"Quarantine ({index}): {entry_serial_number}")
        serial_data: dict = quarantine_data[entry_serial_number]

        entry_part_number: str = serial_data["part"]
        entry_supplier: str = serial_data["supplier"]

        serial_number_data = serials_database.find_one({"_id": entry_serial_number})

        if serial_number_data is None:
            setup_serial: dict = setup_serial_number()

            setup_serial["_id"]: str = entry_serial_number
            setup_serial["in_house"].append("True")
            setup_serial["part_numbers"].append(entry_part_number)
            setup_serial["from_locations"].append("Unknown")
            setup_serial["to_locations"].append("Quarantine")
            setup_serial["scanners"].append("Unknown")
            setup_serial["requesters"].append("Unknown")
            setup_serial["dates"].append(current_date)
            setup_serial["times"].append(current_time)
            setup_serial["suppliers"].append(entry_supplier)
            setup_serial["tool_versions"].append("2.7.3")
            setup_serial["notes"].append("From Inventory Team")

            serials_database.insert_one(setup_serial)

        else:
            serial_number_data["in_house"].append("True")
            serial_number_data["part_numbers"].append(entry_part_number)
            serial_number_data["from_locations"].append("Unknown")
            serial_number_data["to_locations"].append("Quarantine")
            serial_number_data["scanners"].append("Unknown")
            serial_number_data["requesters"].append("Unknown")
            serial_number_data["dates"].append(current_date)
            serial_number_data["times"].append(current_time)
            serial_number_data["suppliers"].append(entry_supplier)
            serial_number_data["tool_versions"].append("2.7.3")
            serial_number_data["notes"].append("From Inventory Team")

            serials_database.update_one({"_id": entry_serial_number},
                                        {"$set": serial_number_data},
                                        upsert=True)


def store_inventory_in_database(inventory_data: dict) -> None:
    """
    Store quarantine data inside database.
    :return:
    """
    serials_database = get_serial_numbers_database()

    current_date: str = get_current_date()
    current_time: str = datetime.today().strftime("%I:%M %p")

    for index, entry_serial_number in enumerate(inventory_data, start=1):
        print(f"Inventory ({index}): {entry_serial_number}")
        serial_data: dict = inventory_data[entry_serial_number]

        entry_part_number: str = serial_data["clean"]["part"]
        entry_supplier: str = serial_data["clean"]["supplier"]
        entry_location: str = serial_data["clean"]["location"]

        serial_number_data = serials_database.find_one({"_id": entry_serial_number})

        if serial_number_data is None:
            setup_serial: dict = setup_serial_number()

            setup_serial["_id"]: str = entry_serial_number
            setup_serial["in_house"].append("True")
            setup_serial["part_numbers"].append(entry_part_number)
            setup_serial["from_locations"].append("Unknown")
            setup_serial["to_locations"].append(entry_location)
            setup_serial["scanners"].append("Unknown")
            setup_serial["requesters"].append("Unknown")
            setup_serial["dates"].append(current_date)
            setup_serial["times"].append(current_time)
            setup_serial["suppliers"].append(entry_supplier)
            setup_serial["tool_versions"].append("2.7.3")
            setup_serial["notes"].append("From Rich")

            serials_database.insert_one(setup_serial)

        else:
            serial_number_data["in_house"].append("True")
            serial_number_data["part_numbers"].append(entry_part_number)
            serial_number_data["from_locations"].append("Unknown")
            serial_number_data["to_locations"].append(entry_location)
            serial_number_data["scanners"].append("Unknown")
            serial_number_data["requesters"].append("Unknown")
            serial_number_data["dates"].append(current_date)
            serial_number_data["times"].append(current_time)
            serial_number_data["suppliers"].append(entry_supplier)
            serial_number_data["tool_versions"].append("2.7.3")
            serial_number_data["notes"].append("From Rich")

            serials_database.update_one({"_id": entry_serial_number},
                                        {"$set": serial_number_data},
                                        upsert=True)


def store_combined_data_in_database(combined_data: dict) -> None:
    """
    Store quarantine data inside database.
    :return:
    """
    serials_database = get_serial_numbers_database()

    current_date: str = get_current_date()
    current_time: str = datetime.today().strftime("%I:%M %p")

    for index, entry_serial_number in enumerate(combined_data, start=1):
        print(f"Inventory - {index}: {entry_serial_number}")
        serial_data: dict = combined_data[entry_serial_number]

        locations: list = serial_data["logs"]["locations"]
        part_numbers: list = serial_data["logs"]["parts"]
        suppliers: list = serial_data["logs"]["suppliers"]
        section: list = serial_data["original"]["section"]

        serial_number_data = serials_database.find_one({"_id": entry_serial_number})

        if serial_number_data is None:
            setup_serial: dict = setup_serial_number()

            setup_serial["_id"]: str = entry_serial_number

            setup_serial["part_numbers"].append(part_numbers[-1])
            setup_serial["from_locations"].append("Unknown")
            setup_serial["scanners"].append("Unknown")
            setup_serial["requesters"].append("Unknown")
            setup_serial["dates"].append(current_date)
            setup_serial["times"].append(current_time)
            setup_serial["suppliers"].append(suppliers[-1])
            setup_serial["tool_versions"].append("2.7.3")

            setup_serial: dict = update_database_locations(locations, setup_serial)
            setup_serial: dict = update_database_notes(section, setup_serial)

            serials_database.insert_one(setup_serial)

        else:
            print("\tEXCEPTION <----------------------------------")


def update_database_locations(locations: list, setup_serial) -> dict:
    """

    :param locations:
    :param setup_serial:
    :return:
    """
    if "In" in locations:
        setup_serial["in_house"].append("True")
        setup_serial["to_locations"].append(locations[-1])

    elif "Out" in locations:
        setup_serial["in_house"].append("False")
        setup_serial["to_locations"].append("Out")

    else:
        setup_serial["in_house"].append("True")
        setup_serial["to_locations"].append(locations[-1])

    return setup_serial


def update_database_notes(section: list, setup_serial: dict) -> dict:
    """
    Update database notes.
    :param section:
    :param setup_serial:
    :return:
    """
    if section == "Quarantine":
        setup_serial["notes"].append("From Inventory Team")

    elif section == "Inventory":
        setup_serial["notes"].append("From Rich")

    return setup_serial


def get_combined_data() -> dict:
    """
    Combined data includes inventory, transaction, and quarantine.
    :return:
    """
    worksheet = get_combined_worksheet()

    inventory: dict = setup_inventory_dict()

    for row_number in range(2, worksheet.max_row):
        if worksheet[f"A{row_number}"].value is None:
            break

        fields: dict = setup_combined_data(row_number, worksheet)
        inventory_data: dict = inventory["data"]
        inventory_stats: dict = inventory["stats"]

        clean_serial: str = fields["clean"]["serial"]
        original_serial: str = fields["original"]["serial"]

        if clean_serial not in inventory_data:
            inventory_data[clean_serial]: dict = {}
            inventory_data[clean_serial]: dict = update_combined_data(fields)

            inventory_stats["after"] += 1

        elif clean_serial in inventory_data:
            current_serial: dict = inventory_data[clean_serial]
            inventory_data[clean_serial]: dict = update_inventory_commodities(current_serial, fields)

            # print(f"Serial Number: {inventory_data[clean_serial]['original']['serial']}")
            # print(f"Clean SN:      {', '.join(inventory_data[clean_serial]['logs']['serials'])}")
            # print(f"Parts:         {', '.join(inventory_data[clean_serial]['logs']['parts'])}")
            # print(f"Suppliers:     {', '.join(inventory_data[clean_serial]['logs']['suppliers'])}")
            # print(f"Locations:     {', '.join(inventory_data[clean_serial]['logs']['locations'])}\n")

            inventory_stats["duplicates"] += 1


        if len(original_serial) != len(clean_serial):

            # print(f"Serial Number: {inventory_data[clean_serial]['original']['serial']}")
            # print(f"Clean SN:      {', '.join(inventory_data[clean_serial]['logs']['serials'])}")
            # print(f"Parts:         {', '.join(inventory_data[clean_serial]['logs']['parts'])}")
            # print(f"Suppliers:     {', '.join(inventory_data[clean_serial]['logs']['suppliers'])}")
            # print(f"Locations:     {', '.join(inventory_data[clean_serial]['logs']['locations'])}\n")

            inventory_stats["part_in_serial"] += 1

        unique_part_numbers = list(set(inventory_data[clean_serial]["logs"]["parts"]))
        if len(unique_part_numbers) >= 2:

            # print(f"Serial Number: {clean_serial}")
            # print(f"Parts:         {', '.join(inventory_data[clean_serial]['logs']['parts'])}")
            # print(f"Suppliers:     {', '.join(inventory_data[clean_serial]['logs']['suppliers'])}")
            # print(f"Locations:     {', '.join(inventory_data[clean_serial]['logs']['locations'])}\n")

            inventory_stats["different_part_numbers"] += 1

        inventory_stats["before"] += 1

    import json
    foo = json.dumps(inventory["stats"], sort_keys=True, indent=4)
    print(foo)
    input()

    return inventory["data"]


def get_combined_worksheet():
    workbook = load_workbook("inventory_combined_raw.xlsx")
    return workbook["Combined"]


def main_method() -> None:
    """
    Starting point.
    :return:
    """
    combined_data: dict = get_combined_data()
    # inventory_data: dict = get_kirkland_inventory()
    # transactions_data: dict = get_inventory_transactions()
    # quarantine_data: dict = get_quarantine_data()

    # store_combined_data_in_database(combined_data)
    # store_inventory_in_database(inventory_data)
    # store_transactions_in_database(transactions_data)
    # store_quarantine_in_database(quarantine_data)


main_method()
