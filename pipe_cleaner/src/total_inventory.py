"""
Get total from Kirkland - 5/6/2021
"""
import os
import sys
from time import strftime

import xlsxwriter
from colorama import Fore, Style
from openpyxl import load_workbook
from csv import reader

from pipe_cleaner.src.data_console_server import main_method as get_console_server_data
from pipe_cleaner.src.excel_properties import Structure


def get_clean_user_name(default_user_name):
    if '.' not in default_user_name:
        indexed_clean_name: list = []
        for index, character in enumerate(default_user_name, start=0):
            index_character: str = default_user_name[index]
            if index_character.isupper() and not index == 0:
                indexed_clean_name.append(' ')
                indexed_clean_name.append(index_character)
            else:
                indexed_clean_name.append(index_character)

        return ''.join(indexed_clean_name).replace('-Ext', '')
    else:
        return default_user_name.replace('.', ' ').title().replace('-Ext', '')


def get_last_entry_row(excel_sheet) -> int:
    """
    
    :param excel_sheet:
    :return:
    """
    index: int = 2

    for row in range(5_000):
        date_value = str(excel_sheet[f'D{index}'].value)
        name_value = str(excel_sheet[f'E{index}'].value)

        if date_value == 'None' and name_value == 'None':
            return index

        elif date_value == '' and name_value == '':
            return index

        elif not date_value and not name_value:
            return index

        index += 1


def get_drive_last_row(excel_sheet) -> int:
    """

    :param excel_sheet:
    :return:
    """
    index: int = 10
    for row in range(10_000):
        date_value = str(excel_sheet[f'D{index}'].value)
        name_value = str(excel_sheet[f'E{index}'].value)

        if date_value == 'None' and name_value == 'None':
            return index

        elif date_value == '' and name_value == '':
            return index

        elif not date_value and not name_value:
            return index

        index += 1


def get_current_transactions_data(worksheet, last_row) -> list:
    """
    Grab data from Z: Drive on transactions data.
    :param worksheet:
    :param last_row:
    :return:
    """
    transactions_data: list = []

    index: int = 10
    for row in range(10_000):
        current_row: list = [worksheet[f'B{index}'].value,
                             worksheet[f'C{index}'].value,
                             worksheet[f'D{index}'].value,
                             worksheet[f'E{index}'].value,
                             worksheet[f'F{index}'].value,
                             worksheet[f'G{index}'].value,
                             worksheet[f'H{index}'].value,
                             worksheet[f'I{index}'].value,
                             worksheet[f'J{index}'].value,
                             worksheet[f'K{index}'].value,
                             worksheet[f'L{index}'].value,
                             worksheet[f'M{index}'].value,
                             worksheet[f'N{index}'].value,
                             worksheet[f'O{index}'].value]

        clean_data: dict = {'approved_date': current_row[0],
                            'approved': current_row[1],
                            'date': current_row[2],
                            'name': current_row[3],
                            'from': current_row[4],
                            'to': current_row[5],
                            'type': current_row[6],
                            'part_number': current_row[7],
                            'quantity': current_row[8],
                            'supplier': current_row[9],
                            'pipe_number': current_row[10],
                            'task_number': current_row[11],
                            'po_number': current_row[12],
                            'notes': current_row[13]}

        transactions_data.append(clean_data)

        if index == last_row:
            transactions_data.pop(-1)
            return transactions_data

        else:
            index += 1


def get_excel_date(raw_data: str) -> str:
    clean_data = raw_data.replace('datetime.datetime(', '').replace(', 0, 0)', '').replace(' 00:00:00', '')
    year = clean_data[0:4]
    month = clean_data[5:7]
    day = clean_data[8:10]

    return f'{month}/{day}/{year}'


def clean_part_numbers(raw_data: str) -> str:
    clean_data = str(raw_data).split(' ')[0].replace(r'\u00a0', '').strip().upper()
    return clean_data


def clean_numbers(raw_data: str) -> int:
    clean_data = str(raw_data).strip()

    if not clean_data or clean_data.upper() == 'NONE' or clean_data == '':
        return 0
    else:
        try:
            return int(clean_data)
        except ValueError:
            return 0


def get_old_inventory_data(worksheet, last_row: int) -> list:
    """
    Grab data from local directory.
    """
    transactions_data: list = []

    index: int = 2
    for row in range(10_000):
        current_row: list = [worksheet[f'A{index}'].value,
                             worksheet[f'B{index}'].value,
                             worksheet[f'C{index}'].value,
                             worksheet[f'D{index}'].value]

        current_data: dict = {'type': str(current_row[0]).strip(),
                              'supplier': str(current_row[1]).strip(),
                              'part_number': clean_part_numbers(current_row[2]),
                              'quantity': clean_numbers(current_row[3])}

        transactions_data.append(current_data)

        if index == last_row:
            return transactions_data

        else:
            index += 1


def get_manual_transactions_data(worksheet, last_row: int) -> list:
    """
    Grab data from Z: Drive on transactions data.

    :param worksheet:
    :param last_row:
    :return:
    """
    transactions_data: list = []

    index: int = 2
    for row in range(10_000):
        current_row: list = [get_excel_date(str(worksheet[f'A{index}'].value)),
                             worksheet[f'B{index}'].value,
                             worksheet[f'C{index}'].value,
                             worksheet[f'D{index}'].value,
                             worksheet[f'E{index}'].value,
                             worksheet[f'F{index}'].value,
                             clean_part_numbers(str(worksheet[f'G{index}'].value)),
                             worksheet[f'H{index}'].value,
                             worksheet[f'I{index}'].value]

        current_data: dict = {'date': current_row[0],
                              'person': current_row[1],
                              'from': str(current_row[2]).split('_')[0].lower(),
                              'to': str(current_row[2]).split('_')[-1].lower(),
                              'supplier': current_row[5],
                              'part_number': current_row[6],
                              'quantity': current_row[7],
                              'notes': current_row[8]}

        transactions_data.append(current_data)

        if index == last_row:
            transactions_data.pop(-1)
            return transactions_data

        else:
            index += 1


def clean_current_transactions(current_transactions: list) -> dict:
    """

    :param current_transactions:
    :return:
    """
    clean_data: dict = {}

    for row_data in current_transactions:
        approved = str(row_data['approved']).upper()

        if approved == 'DA' or approved == 'CS' or approved == 'JS' or approved == 'MH' or approved == 'RM' \
                or approved == 'JT':
            part_number: str = row_data['part_number']
            from_location: int = row_data['from']
            to_location: int = row_data['to']
            quantity = int(clean_numbers(row_data['quantity']))

            if from_location == 'Taking Pictures' or to_location == 'Taking Pictures':
                pass

            elif not part_number and not quantity:
                pass

            elif part_number in clean_data:
                if from_location == 'Cage':
                    clean_data[part_number]['quantity'] -= quantity
                else:
                    clean_data[part_number]['quantity'] += quantity

            else:
                clean_data[part_number]: dict = {}
                clean_data[part_number]['supplier']: str = row_data['supplier']
                clean_data[part_number]['type']: str = row_data['type']
                if from_location == 'Cage':
                    clean_data[part_number]['quantity']: int = 0 - int(quantity)
                else:
                    clean_data[part_number]['quantity']: int = quantity

    return clean_data


def get_transactions_from_drive(file_path: str) -> dict:
    """
    Given the path way to shared drive, look for current inventory transactions through agreed automation design.

    :param file_path: either shared main path or alt main path as expressed in the Z: Drive path
    :return: contains clean data structure for clean comparison later.
    """
    try:
        workbook = load_workbook(file_path)
        worksheet = workbook['All Inventory Transaction']

        last_row: int = get_drive_last_row(worksheet)
        current_transactions = get_current_transactions_data(worksheet, last_row)
        return clean_current_transactions(current_transactions)
    except OSError:
        print(f'\n\tNeed access to Z: Drive (172.30.1.100 network) !!!')
        print(f'\tPlease login into Z: Drive and run Pipe Cleaner again...')
        input()
        sys.exit()


def get_current_inventory_transactions(file_paths: dict) -> dict:
    """
    Get current inventory transactions.

    WARNING: Might have an alternative path due to Z: Drive having 172 IP call vs just Z: Drive call.  Both the same,
    however certain machines might be privy to one or the other network.  For this reason, there are two file
    pathways present to account for both situation.

    :return: clean data for later comparison.
    """
    try:
        return get_transactions_from_drive(file_paths['shared_alt'])

    except FileNotFoundError:
        return get_transactions_from_drive(file_paths['shared_main'])


def clean_old_transactions(old_transactions_data: list) -> dict:
    """
    Convert old transactions to same data structure.

    :param old_transactions_data:
    :return:
    """
    clean_data: dict = {}

    for row_data in old_transactions_data:
        part_number: str = row_data['part_number']
        quantity = int(row_data['quantity'])

        if part_number in clean_data:
            clean_data[part_number]['quantity'] += quantity

        else:
            clean_data[part_number]: dict = {}
            clean_data[part_number]['quantity']: int = quantity
            clean_data[part_number]['supplier']: str = row_data['supplier']

    return clean_data


def get_manual_transactions(file_paths: dict) -> dict:
    """
    Old transactions that are manual counted via inventory management.

    :return: Contains clean structure to allow apples to apples comparison later.
    """
    worksheet_1 = load_workbook(file_paths['manual_transaction'])['Sheet1']

    last_row: int = get_last_entry_row(worksheet_1)

    manual_transactions_data: list = get_manual_transactions_data(worksheet_1, last_row)
    return clean_old_transactions(manual_transactions_data)


def clean_old_inventory(raw_old_inventory: list) -> dict:
    """
    Clean old inventory to combine similar data

    :param raw_old_inventory:
    :return:
    """
    clean_data: dict = {}

    for row_data in raw_old_inventory:
        part_number: str = row_data['part_number']

        if part_number == 'NONE' or not part_number or part_number == '':
            pass

        else:
            try:
                if part_number in clean_data:
                    clean_data[part_number]['quantity'] += int(row_data['quantity'])

                else:
                    clean_data[part_number]: dict = {}
                    clean_data[part_number]['type'] = row_data['type']
                    clean_data[part_number]['supplier'] = row_data['supplier']
                    clean_data[part_number]['quantity'] = int(row_data['quantity'])

            except KeyError:
                clean_data[part_number]: dict = {}
                clean_data[part_number]['type'] = row_data['type']
                clean_data[part_number]['supplier'] = row_data['supplier']
                clean_data[part_number]['quantity'] = int(row_data['quantity'])

    return clean_data


def get_last_manual_inventory_count(file_paths: dict) -> dict:
    """
    Get manual count of inventory.  This is where the inventory manager accounts for manual count of inventory.
    """
    worksheet_2 = load_workbook(file_paths['manual_transaction'])['Sheet2']

    last_row: int = get_last_entry_row(worksheet_2)

    raw_old_inventory = get_old_inventory_data(worksheet_2, last_row)
    return clean_old_inventory(raw_old_inventory)


def consolidate_data(old_transactions: dict, current_transactions: dict) -> dict:
    """
    Put together old transactions and current transactions (Shared Drive) for easier comparison later.

    :param old_transactions: static data from given excel file
    :param current_transactions: snapshot data taken from shared drive
    :return:
    """
    clean_data: dict = {}

    combined_data: list = merge_dictionaries(old_transactions, current_transactions)

    for part in combined_data:
        try:
            part_number: str = part['part_number']
            supplier: str = part['supplier']
            part_type: str = part['type']
            quantity: int = int(part['quantity'])

            if part_number in clean_data:
                clean_data[part_number]['quantity'] += quantity

            else:
                clean_data[part_number]: dict = {}
                current_data: dict = clean_data[part_number]

                current_data['supplier']: str = supplier
                current_data['type']: str = part_type
                current_data['quantity']: int = quantity
        except ValueError:
            pass

    return clean_data


def merge_dictionaries(dictionary_1: dict, dictionary_2: dict) -> list:
    """
    Combined data for easier iterations.
    """
    all_data: list = []

    all_data: list = add_data_to_total(all_data, dictionary_1)
    all_data: list = add_data_to_total(all_data, dictionary_2)

    return all_data


def add_data_to_total(all_data: list, commodities: dict) -> list:
    for part_number in commodities:
        current_part: dict = {}
        current_data: dict = commodities[part_number]

        current_part['part_number']: str = part_number
        current_part['quantity']: int = current_data.get('quantity', 'None')
        current_part['type']: int = current_data.get('type', 'None')
        current_part['supplier']: int = current_data.get('supplier', 'None')

        all_data.append(current_part)

    return all_data


def flatten_console_server_data(console_server_data: dict) -> dict:
    """
    Convert console server data into easier structure to iterate.
    :param console_server_data:
    :return:
    """
    clean_data: dict = {}

    commodities: dict = console_server_data['commodities']

    clean_data: dict = clean_commodity_data(clean_data, commodities['dimms'], 'dimm')
    clean_data: dict = clean_commodity_data(clean_data, commodities['disks'], 'disk')
    clean_data: dict = clean_commodity_data(clean_data, commodities['nvmes'], 'nvme')

    return clean_data


def clean_commodity_data(clean_data: dict, commodity: dict, part_type: str) -> dict:
    """
    Iterate through DIMMs, NVMEs, and disks to flatten
    :param clean_data:
    :param commodity:
    :param part_type:
    :return:
    """
    for part_number in commodity:
        quantity: int = commodity[part_number]

        if part_number in clean_data:
            clean_data[part_number]['quantity'] += quantity

        else:
            clean_data[part_number]: dict = {}
            clean_data[part_number]['quantity'] = quantity
            clean_data[part_number]['type'] = part_type

    return clean_data


def convert_index_to_letter(index: int) -> str:
    """

    :param index: Current index due to how many columns we care about in the excel output sheet
    """
    lower_character = chr(ord('a') + index)
    return str(lower_character).upper()


def get_letter_for_column_position(initial: int, left_padding: int) -> str:
    """
    For positioning the column title based on starting point of the left padding.
    :return: letter of excel column
    """
    return convert_index_to_letter(initial + left_padding)


def get_column_title_position(header_height: int, index: int, left_padding: int) -> str:
    """
    Get position of the column title based on excel position from the letter and number ex. A1, B4, C3
    """
    letter: str = get_letter_for_column_position(index, left_padding)
    return f'{letter}{header_height}'


def add_white_cell(position: str, current_setup) -> None:
    """
    Account for empty cells that don't have column title.
    Meant for giving space between different groups of data.
    """
    worksheet: xlsxwriter = current_setup.get('worksheet')
    structure: xlsxwriter = current_setup.get('structure')

    worksheet.write(position, '', structure.white)


def add_column_titles(current_setup: dict) -> None:
    """
    Set up Column Names in the Excel table for categorizing into vertical data later
    """
    header_height: int = current_setup.get('header_height')
    left_padding: int = current_setup.get('left_padding')
    column_names: tuple = current_setup.get('column_names')

    for index, column_title in enumerate(column_names, start=0):
        position: str = get_column_title_position(header_height, index, left_padding)

        if not column_title:
            add_white_cell(position, current_setup)

        else:
            add_column_title(position, column_title, current_setup)


def add_vse_logo_top_right(current_setup: dict) -> None:
    """
    Creates VSE Logo on the top left corner
    :param current_setup:
    """
    worksheet: xlsxwriter = current_setup.get('worksheet')

    worksheet.insert_image('A1', 'pipe_cleaner/img/vse_logo.png')


def add_column_title(position: str, column_title: str, current_setup: dict) -> None:
    """
    Add column title to the current excel sheet
    """
    worksheet: xlsxwriter = current_setup.get('worksheet')
    structure: xlsxwriter = current_setup.get('structure')

    worksheet.write(position, column_title, structure.teal_middle_16)


def set_excel_design(current_setup: dict) -> None:
    """
    Set up excel output design/parameters.
    """
    set_rows_and_columns_sizes(current_setup)

    add_column_titles(current_setup)
    add_freeze_panes(current_setup)
    add_vse_logo_top_right(current_setup)



def add_freeze_panes(current_setup: dict) -> None:
    """
    Allows information to the left to stay
    """
    header_height: int = current_setup.get('header_height')
    freeze_pane_position: int = current_setup.get('freeze_pane_position')
    worksheet: xlsxwriter = current_setup.get('worksheet')

    worksheet.freeze_panes(header_height, freeze_pane_position)


def add_header_user_name(current_setup: dict):
    """
    Add clean user name to the top left corner.
    """
    worksheet: xlsxwriter = current_setup.get('worksheet')
    structure: xlsxwriter = current_setup.get('structure')

    worksheet.write('C5', 'Total Inventory - Kirkland', structure.black_35)


def set_header_rows_height(rows_height: tuple, worksheet: xlsxwriter, structure: xlsxwriter) -> None:
    """
    Establishes current worksheet row vertical heights for the header.
    """
    for index, row_size in enumerate(rows_height, start=0):
        worksheet.set_row(index, row_size, structure.white)


def set_rows_and_columns_sizes(current_setup) -> None:
    """
    Beginning of the Excel Structure
    """
    rows_height: tuple = current_setup.get('rows_height')
    columns_width: tuple = current_setup.get('columns_width')

    worksheet: xlsxwriter = current_setup.get('worksheet')
    structure: xlsxwriter = current_setup.get('structure')

    set_header_rows_height(rows_height, worksheet, structure)
    set_excel_column_width(columns_width, worksheet, structure)


def set_excel_column_width(columns_width: tuple, worksheet: xlsxwriter, structure: xlsxwriter) -> None:
    """
    Establishes current worksheet column widths.
    """
    for index in range(0, len(columns_width)):
        current_letter: str = convert_index_to_letter(index)

        worksheet.set_column(f'{current_letter}:{current_letter}',
                             columns_width[index],
                             structure.white)


def clean_pipe_cleaner_version(pipe_cleaner_version) -> str:
    """
    Version for documentation
    :param pipe_cleaner_version:
    :return: cleaner version
    """
    return f"v{pipe_cleaner_version.split(' ')[0]}"


def add_header_date_and_version(current_setup: dict) -> None:
    """
    Adds the current date/time and Pipe Cleaner version to the header area in the top left corner
    """
    current_time: str = strftime('%I:%M %p')
    current_date: str = strftime('%m/%d/%Y')
    pipe_cleaner_version: str = current_setup.get('version')

    worksheet: xlsxwriter = current_setup.get('worksheet')
    structure: xlsxwriter = current_setup.get('structure')

    pipe_cleaner_version: str = clean_pipe_cleaner_version(pipe_cleaner_version)

    worksheet.write('C6', f' {current_date} - {current_time} - {pipe_cleaner_version}',
                    structure.italic_blue_font)


def add_header_data(current_setup: dict) -> None:
    """
    Add header data on ex. username, date, version, etc.
    """
    add_header_user_name(current_setup)
    add_header_date_and_version(current_setup)


def set_sheet_structure(current_setup) -> None:
    """
    Create dashboard structure
    """
    set_excel_design(current_setup)
    add_header_data(current_setup)


def get_file_paths() -> dict:
    """
    Get file paths for extracting data from.
    """
    return {'manual_transaction': 'settings/transaction_logs.xlsx',
            'shared_main': r'Z:\Kirkland_Lab\PipeCleaner\transaction_logs\_inventory_transactions.xlsx',
            'baseline_4_22': 'settings/baseline_cage_4_22.csv',
            'teams': 'settings/transactions.csv',
            'shared_alt': r'//172.30.1.100/pxe/Kirkland_Lab/PipeCleaner/transaction_logs/_inventory_transactions.xlsx'}


def create_inventory_transaction(pipe_cleaner_version: str, default_user_name: str, site_location: str) -> dict:
    """
    Current excel sheet design to setup the excel tab for data to fill in later.
    """
    sheet_title: str = 'Inventory Template'

    check_opened_pipe_cleaner()

    workbook = xlsxwriter.Workbook('inventory_transaction.xlsx')

    excel_setup: dict = {'sheet_title': sheet_title,
                         'worksheet': workbook.add_worksheet(sheet_title),
                         'structure': Structure(workbook),
                         'workbook': workbook,
                         'version': pipe_cleaner_version,
                         'site_location': site_location,
                         'default_user_name': default_user_name,
                         'clean_name': get_clean_user_name(default_user_name),
                         'header_height': 9,
                         'body_position': 15,
                         'left_padding': 2,
                         'freeze_pane_position': 6,
                         'host_group_hyperlink': 'http://172.30.1.100/console/host_groups.php',
                         'rows_height': (15.0, 15.0, 15.0, 15.0, 36.0, 15.75, 15.75, 15.75, 30.0, 60.0, 60.0, 60.0,
                                         60.0, 60.0, 60.0, 60.0, 60.0, 60.0, 60.0, 60.0, 60.0, 60.0),
                         'columns_width': (2.00, 2.00, 17.0, 23.0, 22.0, 22.0, 15.0, 30.0, 12.0, 18.0, 17.0, 17.0,
                                           17.0, 40.0, 21.0, 21.0, 21.0, 21.0, 25.0, 25.0, 25.0, 25.0),
                         'column_names': ('Date',
                                          'Name',
                                          'From',
                                          'To',
                                          'Type',
                                          'Part Number',
                                          'QTY',
                                          'Supplier',
                                          'Pipe #',
                                          'Task #',
                                          'PO #',
                                          'Notes')}

    return excel_setup


def remove_excel_green_corners(current_setup) -> None:
    """
    Excel sometimes have green corners within a cell. Removes to clear up look of excel output.
    :param current_setup: Current worksheet
    """
    worksheet = current_setup.get('worksheet')

    worksheet.ignore_errors({'number_stored_as_text': 'A1:XFD1048576'})


def get_color_left(index: int, structure):

    count: int = index % 2
    if count == 0:
        return structure.blue_left_12
    else:
        return structure.alt_blue_left_12


def get_color_middle(index: int, structure):

    count: int = index % 2
    if count == 0:
        return structure.blue_middle_12
    else:
        return structure.alt_blue_middle_12


def get_quantity_value(inventory_source: dict, part_number: str) -> str:

    try:
        return inventory_source[part_number]['quantity']

    except KeyError:
        return '0'


def get_type_value(inventory_source: dict, part_number: str) -> str:

    try:
        return inventory_source[part_number]['type'].upper()

    except KeyError:
        return 'NONE'


def get_final_type(inventory_source_1: dict, inventory_source_2: dict, part_number: str) -> str:

    value_1 = get_type_value(inventory_source_1, part_number)
    value_2 = get_type_value(inventory_source_2, part_number)

    if value_1 == 'NONE' and value_2 == 'NONE':
        return 'None'
    elif value_1 == 'NONE' and value_2.isascii():
        return value_2
    elif value_1.isascii() and value_2 == 'NONE':
        return value_1
    elif value_1.isascii() and value_2.isascii():
        total: list = [value_1, value_2]
        return ", ".join(total)


def get_final_supplier(inventory_source_1: dict, inventory_source_2: dict, part_number: str) -> str:

    value_1 = get_supplier_value(inventory_source_1, part_number)
    value_2 = get_supplier_value(inventory_source_2, part_number)

    if value_1 == 'None' and value_2 == 'None':
        return 'None'
    elif value_1 == 'None' and value_2.isascii():
        return value_2
    elif value_1.isascii() and value_2 == 'None':
        return value_1
    elif value_1.isascii() and value_2.isascii():
        total: list = [value_1, value_2]
        return ", ".join(total)


def get_supplier_value(inventory_source: dict, part_number: str) -> str:

    try:
        return inventory_source[part_number]['supplier']

    except KeyError:
        return 'None'


def create_excel_output(total_data: dict, inventory_data: dict, console_server_data: dict, current_setup) -> None:
    """
    Create an excel output.
    """
    worksheet = current_setup['worksheet']
    structure = current_setup['structure']

    for index, part_number in enumerate(sorted(total_data), start=10):
        color_left = get_color_left(index, structure)
        color_middle = get_color_middle(index, structure)
        worksheet.set_row(index - 1, 19.5)

        total_quantity: int = total_data[part_number]['quantity']
        cage_quantity: str = get_quantity_value(inventory_data, part_number)
        rack_quantity: str = get_quantity_value(console_server_data, part_number)

        final_type: str = get_final_type(inventory_data, console_server_data, part_number)
        final_supplier: str = get_final_supplier(inventory_data, console_server_data, part_number)

        worksheet.write(f'C{index}', part_number, color_left)
        worksheet.write(f'D{index}', total_quantity, color_middle)
        worksheet.write(f'E{index}', cage_quantity, color_middle)
        worksheet.write(f'F{index}', rack_quantity, color_middle)

        worksheet.write(f'G{index}', final_type, color_middle)
        worksheet.write(f'H{index}', final_supplier, color_middle)

    current_setup['workbook'].close()


def check_opened_pipe_cleaner():
    try:
        os.remove('pipes/main_dashboard.xlsx')
    except FileNotFoundError:
        pass
    except PermissionError:
        print(f'\tExcel file {Fore.RED}main_dashboard.xlsx{Style.RESET_ALL} already up.\n'
              f'\tPlease close excel file and restart Pipe Cleaner.')
        print(f'\n')
        print(f'\tPress {Fore.BLUE}enter{Style.RESET_ALL} to close program.\n')
        input()
        sys.exit()


def create_setup(basic_data) -> dict:
    """
    Current excel sheet design to setup the excel tab for data to fill in later.
    """
    sheet_title: str = 'Inventory Template'
    site: str = basic_data['site']
    username: str = basic_data['username']
    version: str = basic_data['version']

    check_opened_pipe_cleaner()

    workbook = xlsxwriter.Workbook('total_kirkland_inventory.xlsx')

    excel_setup: dict = {'sheet_title': sheet_title,
                         'worksheet': workbook.add_worksheet(sheet_title),
                         'structure': Structure(workbook),
                         'workbook': workbook,
                         'version': version,
                         'site_location': site,
                         'default_user_name': username,
                         'clean_name': get_clean_user_name(username),
                         'header_height': 9,
                         'body_position': 15,
                         'left_padding': 2,
                         'freeze_pane_position': 6,
                         'host_group_hyperlink': 'http://172.30.1.100/console/host_groups.php',
                         'rows_height': (15.0, 15.0, 15.0, 15.0, 36.0, 15.75, 15.75, 15.75, 24.0),
                         'columns_width': (2.00, 2.00, 36.0, 14.0, 14.0, 14.0, 24.0, 26.0, 20.0, 20.0, 20.0, 20.0,
                                           20.0, 20.0, 20.0, 20.0, 20.0, 20.0),
                         'column_names': ('Part Number',
                                          'Total',
                                          'Cage',
                                          'Racks',
                                          'Type',
                                          'Supplier')}

    return excel_setup


def auto_filter(current_setup: dict) -> None:
    """
    Provide filters in excel in the header columns within the excel sheet.
    """
    worksheet = current_setup['worksheet']
    worksheet.autofilter('C9:H9')


def clean_part_number(part_number: str) -> str:
    """
    Standardized part number for consistent comparison later
    """
    return part_number.upper().strip()


def get_baseline_cage(file_paths) -> dict:
    """
    As of April 22nd, baseline count for cage inventory has be done.  This fetches data from that event.
    """
    baseline: dict = {}

    with open(file_paths['baseline_4_22']) as file:
        csv_data = reader(file, delimiter=",")

        for index, row in enumerate(csv_data, start=0):
            if index == 0:
                continue

            part_number: str = clean_part_number(row[0])
            count = int(row[1])

            if part_number in baseline:
                baseline[part_number]['quantity'] += count

            elif part_number not in baseline:
                baseline[part_number]: dict = {}
                baseline[part_number]['quantity'] = count

    return baseline


def get_teams_transactions(file_paths) -> dict:
    """
    As of April 22nd, baseline count for cage inventory has be done.  This fetches data after that event.
    """
    transactions: dict = {}

    with open(file_paths['teams']) as file:
        csv_data = reader(file, delimiter=",")

        for index, row in enumerate(csv_data, start=0):

            if index == 0:
                continue

            destination = str(row[1]).upper()
            part_number: str = clean_part_number(row[2])
            count = int(str(row[3]).strip())

            if part_number in transactions:

                if destination == 'CAGE':
                    transactions[part_number]['quantity'] += count

                elif 'QUAR' in destination:  # QUAR for Quarantine. Accounts for misspelling
                    transactions[part_number]['quantity'] += count

                else:
                    transactions[part_number]['quantity'] -= count

            elif part_number not in transactions:
                transactions[part_number]: dict = {}

                if destination == 'CAGE' or 'QUAR' in destination:
                    transactions[part_number]['quantity'] = count

                else:
                    transactions[part_number]['quantity'] = 0 - count

    return transactions


def main_method(basic_data: dict) -> None:
    """
    Consolidate available data to produce excel output of Kirkland inventory total.
    """
    file_paths: dict = get_file_paths()

    drive_transactions: dict = get_cage_data(file_paths)
    console_server_data: dict = get_console_server_data()
    # azure_devops_data: dict = get_all_ticket_data(console_server_data)
    # part_numbers: dict = get_console_server_data()['part_numbers']

    console_server_data: dict = flatten_console_server_data(get_console_server_data()['inventory'])

    manual_transactions: dict = get_manual_transactions(file_paths)

    manual_inventory_data: dict = get_last_manual_inventory_count(file_paths)

    transactions_data: dict = consolidate_data(manual_transactions, drive_transactions)
    inventory_data: dict = consolidate_data(transactions_data, manual_inventory_data)
    total_data: dict = consolidate_data(inventory_data, console_server_data)

    try:
        current_setup: dict = create_setup(basic_data)

        set_sheet_structure(current_setup)
        remove_excel_green_corners(current_setup)
        auto_filter(current_setup)
        create_excel_output(total_data, inventory_data, console_server_data, current_setup)

    except PermissionError:
        print(f'\n\tPlease close excel output total_kirkland_inventory.xlsx..."')
        print(f'\tPress enter to exit')
        input()


def get_cage_data(file_paths: dict) -> dict:
    """
    Gather data from last manual count, latest Teams transactions, and Shared Drive Transactions.
    """
    cage_data: dict = {}

    baseline_cage: dict = get_baseline_cage(file_paths)
    teams_transactions: dict = get_teams_transactions(file_paths)
    drive_transactions: dict = get_current_inventory_transactions(file_paths)

    cage_data: dict = store_cage_data(baseline_cage, cage_data, 'baseline')
    cage_data: dict = store_cage_data(teams_transactions, cage_data, 'teams')
    cage_data: dict = store_cage_data(drive_transactions, cage_data, 'drive')

    return cage_data


def store_cage_data(inventory_source: dict, cage_data: dict, key_name: str) -> dict:
    """
    Gather from different inventory sources to consolidate data into one container.
    """
    for part_number in inventory_source:
        quantity: int = inventory_source[part_number]['quantity']

        if part_number in cage_data:
            cage_data[part_number]['total_count'] += quantity
            locations: dict = cage_data[part_number]['locations']

            if key_name in locations:
                locations[key_name]['count'] += quantity

            else:
                locations[key_name]: dict = {}
                locations[key_name]['count'] = quantity

        else:
            cage_data[part_number]: dict = {}
            cage_data[part_number]['locations']: dict = {}
            cage_data[part_number]['locations'][key_name]: dict = {}

            cage_data[part_number]['total_count']: int = quantity
            cage_data[part_number]['locations'][key_name]['count'] = quantity

    return cage_data
