"""
Cage Count
"""
from openpyxl import load_workbook
from os import system

from pipe_cleaner.src.log_database import access_database_document


# def get_descriptions() -> dict:
#     parts_sheet = load_workbook(f'settings/part_numbers.xlsx')['descriptions']
#
#     part_numbers_library: dict = {}
#     for row_number in range(2, 177):
#         part_number = str(parts_sheet[f'D{row_number}'].value).upper().strip()
#
#         if part_number not in part_numbers_library:
#             friendly_name: str = parts_sheet[f'A{row_number}'].value
#             item_type: str = parts_sheet[f'B{row_number}'].value
#             model_number: str = parts_sheet[f'C{row_number}'].value
#             item_supplier: str = parts_sheet[f'E{row_number}'].value
#             description: str = parts_sheet[f'F{row_number}'].value
#             rank: str = parts_sheet[f'G{row_number}'].value
#
#             part_numbers_library[part_number]: dict = {}
#             part_numbers_library[part_number]['friendly_name'] = friendly_name
#             part_numbers_library[part_number]['item_type'] = item_type
#             part_numbers_library[part_number]['model_number'] = model_number
#             part_numbers_library[part_number]['item_supplier'] = item_supplier
#             part_numbers_library[part_number]['description'] = description
#             part_numbers_library[part_number]['rank'] = rank
#
#     return part_numbers_library


def main_method() -> None:
    parts_sheet = load_workbook(f'settings/part_numbers.xlsx')['parts']
    transactions_sheet = load_workbook(f'settings/part_numbers.xlsx')['transactions']

    part_numbers: dict = {}
    for row_number in range(2, 170):
        part_number = str(parts_sheet[f'A{row_number}'].value).upper().strip()

        if part_number not in part_numbers:
            part_numbers[part_number] = int(parts_sheet[f'B{row_number}'].value)
        else:
            part_numbers[part_number] += int(parts_sheet[f'B{row_number}'].value)

    for row_number in range(2, 301):
        transact_part_number = str(transactions_sheet[f'C{row_number}'].value).upper().strip()
        flow: str = transactions_sheet[f'A{row_number}'].value
        quantity = int(transactions_sheet[f'D{row_number}'].value)

        if transact_part_number not in part_numbers:
            part_numbers[transact_part_number] = int(transactions_sheet[f'D{row_number}'].value)

        else:

            if flow == 'Pipe_To_Cage':
                part_numbers[transact_part_number] += quantity

            elif flow == 'Cage_To_Pipe':
                part_numbers[transact_part_number] -= quantity

            elif flow == 'Shipment_From_Cage':
                part_numbers[transact_part_number] -= quantity

            elif flow == 'Receipt_To_Cage':
                part_numbers[transact_part_number] += quantity

    workbook = load_workbook('settings/all_part_numbers_template.xlsx')
    worksheet = workbook['cage']

    document = access_database_document('transactions', '021')
    all_transactions = document.find({})

    for transaction in all_transactions:
        part_number: str = transaction['part_number']
        scanned_count: int = len(transaction['scanned'])

        if part_number not in part_numbers:
            part_numbers[part_number] = scanned_count

        else:
            part_numbers[part_number] += scanned_count

    descriptions: dict = get_descriptions()

    for index, part_number in enumerate(part_numbers, start=2):
        worksheet[f'A{index}'] = part_number
        worksheet[f'B{index}'] = part_numbers[part_number]
        worksheet[f'C{index}'] = descriptions.get(part_number, {}).get('friend_name', '')
        worksheet[f'D{index}'] = descriptions.get(part_number, {}).get('item_type', '')
        worksheet[f'E{index}'] = descriptions.get(part_number, {}).get('model_number', '')
        worksheet[f'F{index}'] = descriptions.get(part_number, {}).get('item_supplier', '')
        worksheet[f'G{index}'] = descriptions.get(part_number, {}).get('description', '')
        worksheet[f'H{index}'] = descriptions.get(part_number, {}).get('rank', '')

    workbook.save(fr'library/all_cage_numbers.xlsx')
    system(fr'start EXCEL.EXE library/all_cage_numbers.xlsx')


def get_descriptions() -> dict:
    parts_sheet = load_workbook(f'settings/part_numbers.xlsx')['descriptions']

    part_numbers_library: dict = {}
    for row_number in range(2, 177):
        part_number = str(parts_sheet[f'D{row_number}'].value).upper().strip()

        if part_number not in part_numbers_library:
            friendly_name: str = parts_sheet[f'A{row_number}'].value
            item_type: str = parts_sheet[f'B{row_number}'].value
            model_number: str = parts_sheet[f'C{row_number}'].value
            item_supplier: str = parts_sheet[f'E{row_number}'].value
            description: str = parts_sheet[f'F{row_number}'].value
            rank: str = parts_sheet[f'G{row_number}'].value

            part_numbers_library[part_number]: dict = {}
            part_numbers_library[part_number]['friendly_name'] = friendly_name
            part_numbers_library[part_number]['item_type'] = item_type
            part_numbers_library[part_number]['model_number'] = model_number
            part_numbers_library[part_number]['item_supplier'] = item_supplier
            part_numbers_library[part_number]['description'] = description
            part_numbers_library[part_number]['rank'] = rank

    return part_numbers_library
