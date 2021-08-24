"""
Export monthly transactions on inventory from latest to earliest.
"""
from os import system

from openpyxl import load_workbook
from openpyxl.styles import Alignment

from pipe_cleaner.src.log_database import access_database_document


def reverse_transactions(all_transactions: list) -> list:
    """

    """
    order: list = []

    for entry in all_transactions:
        order.append(entry)

    new_order: list = order[::-1]

    return new_order


def add_excel_data(document, worksheet: load_workbook) -> None:
    """
    Add serial number data to excel.
    """
    all_transactions: list = document.find({})
    transactions: list = reverse_transactions(all_transactions)

    for index, transaction_log in enumerate(transactions, start=2):
        # import json
        # foo = json.dumps(transaction_log, sort_keys=True, indent=4)
        # print(foo)
        # input()

        date: str = transaction_log["time"]["date_logged"]
        approved_by: str = transaction_log["source"]["approved_by"]
        form_number: str = transaction_log["source"]["form_number"]
        previous: str = transaction_log["location"]["previous"]
        current: str = transaction_log["location"]["current"]
        pipe: str = transaction_log["location"]["pipe"]
        part_number: str = transaction_log["part_number"]
        quantity = str(len(transaction_log["scanned"]))
        task: str = transaction_log["source"]["task"]
        scanned: list = transaction_log["scanned"]

        worksheet[f'A{index}']: str = date
        worksheet[f'B{index}']: str = approved_by
        worksheet[f'C{index}']: str = form_number
        worksheet[f'D{index}']: str = previous
        worksheet[f'E{index}']: str = current
        worksheet[f'F{index}']: str = pipe
        worksheet[f'G{index}']: str = part_number
        worksheet[f'H{index}']: str = quantity
        worksheet[f'I{index}']: str = task
        worksheet[f'J{index}']: str = ", ".join(str(v) for v in scanned)

        wrap_text(index, worksheet)


def wrap_text(index: int, worksheet):
    """
    Make sure text wraps on certain fields.
    """
    worksheet[f'A{index}'].alignment = Alignment(wrap_text=True, horizontal='center')
    worksheet[f'B{index}'].alignment = Alignment(wrap_text=True, horizontal='center')
    worksheet[f'C{index}'].alignment = Alignment(wrap_text=True, horizontal='center')
    worksheet[f'D{index}'].alignment = Alignment(wrap_text=True, horizontal='center')
    worksheet[f'E{index}'].alignment = Alignment(wrap_text=True, horizontal='center')
    worksheet[f'F{index}'].alignment = Alignment(wrap_text=True, horizontal='center')
    worksheet[f'G{index}'].alignment = Alignment(wrap_text=True, horizontal='center')
    worksheet[f'H{index}'].alignment = Alignment(wrap_text=True, horizontal='center')
    worksheet[f'I{index}'].alignment = Alignment(wrap_text=True)


def main_method() -> None:
    """

    """
    print(f'\n\tGetting transactions data from database...')

    document = access_database_document('transactions', '021')

    workbook = load_workbook(fr'settings/transaction_logs_template.xlsx')
    worksheet = workbook['Sheet1']

    print(f'\n\tCreating excel output...')
    add_excel_data(document, worksheet)

    workbook.save(fr'pipes/transaction_logs.xlsx')
    system(fr'start EXCEL.EXE pipes/transaction_logs.xlsx')
