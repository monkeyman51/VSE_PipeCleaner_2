"""
Fetch part numbers library data from database to output data.
"""
from os import system

from openpyxl import load_workbook
from openpyxl.styles import Alignment

from pipe_cleaner.src.log_database import access_database_document


def add_excel_data(all_part_numbers: list, worksheet: load_workbook) -> None:
    """
    Add serial number data to excel.
    """
    serial_numbers_count: int = 0
    for index, part_number_data in enumerate(all_part_numbers, start=3):
        serial_numbers_count += 1

        worksheet[f'A{index}'] = part_number_data['_id']
        current_source: dict = part_number_data['sources'][-1]

        worksheet[f'B{index}'] = current_source['source']
        worksheet[f'C{index}'] = current_source['date_logged']
        worksheet[f'D{index}'] = current_source['time_logged']

        worksheet[f'E{index}'] = current_source['friendly_name']
        worksheet[f'F{index}'] = current_source['model_number']
        worksheet[f'G{index}'] = current_source['supplier']
        worksheet[f'H{index}'] = current_source['manufacturer']

        worksheet[f'I{index}'] = current_source['type']
        worksheet[f'J{index}'] = current_source['size']
        worksheet[f'K{index}'] = current_source['speed']
        worksheet[f'L{index}'] = current_source['ddr']
        worksheet[f'M{index}'] = current_source['voltage']
        worksheet[f'N{index}'] = current_source['description']
        worksheet[f'O{index}'] = current_source['comment']

        for number in range(1, 18):
            current_letter = str(chr(ord('@')+number))
            worksheet[f'{current_letter}{index}'].alignment = Alignment(horizontal='center')

    worksheet['A1'] = f'Total P/N - {str(serial_numbers_count)}'
    worksheet['A1'].alignment = Alignment(horizontal='center')


def main_method() -> None:
    """

    """
    print(f'\n\tGetting part_numbers data from database...')
    all_part_numbers: list = access_database_document('part_numbers', 'all').find({})
    workbook = load_workbook(fr'settings/pn_library_template.xlsx')
    worksheet = workbook['part_numbers']

    print(f'\n\tCreating excel output...')
    add_excel_data(all_part_numbers, worksheet)
    workbook.save(fr'pipes/serial_numbers.xlsx')
    system(fr'start EXCEL.EXE pipes/serial_numbers.xlsx')
