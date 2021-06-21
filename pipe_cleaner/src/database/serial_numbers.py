"""
Fetch serial numbers from cloud database and output excel file with serial numbers data.
"""
from os import system

from openpyxl import load_workbook
from openpyxl.styles import Alignment

from pipe_cleaner.src.log_database import access_database_document


def add_excel_data(all_serial_numbers: list, worksheet: load_workbook) -> None:
    """
    Add serial number data to excel.
    """
    serial_numbers_count: int = 0
    for index, serial_number_data in enumerate(all_serial_numbers, start=3):
        serial_numbers_count += 1

        worksheet[f'A{index}'] = serial_number_data['_id']
        worksheet[f'B{index}'] = serial_number_data['part_number']
        worksheet[f'C{index}'] = len(serial_number_data['transactions'])

        current_transaction: dict = serial_number_data['transactions'][-1]

        worksheet[f'D{index}'] = current_transaction['time']['date_logged']
        worksheet[f'E{index}'] = current_transaction['time']['time_logged']

        worksheet[f'F{index}'] = current_transaction['location']['current']
        worksheet[f'G{index}'] = current_transaction['location']['previous']
        worksheet[f'H{index}'] = current_transaction['location']['rack']
        worksheet[f'I{index}'] = current_transaction['location']['machine']
        worksheet[f'J{index}'] = current_transaction['location']['pipe']
        worksheet[f'K{index}'] = current_transaction['location']['site']

        worksheet[f'L{index}'] = current_transaction['source']['approved_by']
        worksheet[f'M{index}'] = current_transaction['source']['verified_by']
        worksheet[f'N{index}'] = current_transaction['source']['trr']
        worksheet[f'O{index}'] = current_transaction['source']['version']
        worksheet[f'P{index}'] = current_transaction['source']['comment']
        worksheet[f'Q{index}'] = current_transaction['source']['task']

        for number in range(1, 18):
            current_letter = str(chr(ord('@')+number))
            worksheet[f'{current_letter}{index}'].alignment = Alignment(horizontal='center')

    worksheet['A1'] = f'Total S/N - {str(serial_numbers_count)}'
    worksheet['A1'].alignment = Alignment(horizontal='center')


def main_method() -> None:
    """

    """
    print(f'\n\tGetting serial numbers data from database...')
    all_serial_numbers: list = access_database_document('serial_numbers', 'all').find({})
    workbook = load_workbook(fr'settings/serial_numbers_template.xlsx')
    worksheet = workbook['serial_numbers']

    print(f'\n\tCreating excel output...')
    add_excel_data(all_serial_numbers, worksheet)
    workbook.save(fr'pipes/serial_numbers.xlsx')
    system(fr'start EXCEL.EXE pipes/serial_numbers.xlsx')
