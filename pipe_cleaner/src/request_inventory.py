"""
Responsible for handling request input for inventory transactions.  Includes handling part number, logistics, task name,
and other important information needed to be sent to the inventory team to handle material movement.

6/3/2021

"""
import sys
from os import system
from time import strftime, sleep

import win32com.client as client
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from psutil import process_iter as task_manager, NoSuchProcess
from pipe_cleaner.src.log_database import access_database_document


def add_basic_documents_info(workbook, basic_info: dict, sheet_name: str) -> dict:
    """
    Handles basic documenting information like data, time, user, and site.
    """
    sheet_name: str = sheet_name.title()

    print(f'\t\t- Creating {sheet_name} excel')

    return add_excel_data(sheet_name, basic_info, workbook['Request'])


def get_template_dictionary(start_location: str, end_location: str) -> dict:
    """
    Get template for fill later.  Everything is optional until it is assigned as main.
    """
    normal_default: str = 'Copy From Task'

    return {'task_name': {'cell': '$C$13',
                          'value': 'optional',
                          'default': normal_default},

            'part_number': {'cell': '$C$17',
                            'value': 'optional',
                            'default': normal_default},

            'start_location': {'cell': '$C$20',
                               'value': start_location.title(),
                               'default': 'None'},

            'end_location': {'cell': '$C$23',
                             'value': end_location.title(),
                             'default': 'None'},

            'notes': {'cell': '$C$26',
                      'value': 'optional',
                      'default': 'None'},

            'pipe_number': {'cell': '$C$29',
                            'value': 'optional',
                            'default': normal_default},

            'machine_name': {'cell': '$C$32',
                             'value': 'optional',
                             'default': normal_default},

            'trr_number': {'cell': '$C$35',
                           'value': 'optional',
                           'default': normal_default},

            'purchase_order': {'cell': '$C$38',
                               'value': 'optional',
                               'default': 'Copy From Package'},

            'main_fields': {'cell': '$F$10',
                            'value': 0,
                            'default': 'None'}}


def get_template_pipe_cage(start_location: str, end_location: str) -> dict:
    """
    Get request template based on start and end locations.
    """
    template: dict = get_template_dictionary(start_location, end_location)

    template['task_name']['value']: str = 'main'
    template['part_number']['value']: str = 'main'
    template['pipe_number']['value']: str = 'main'
    template['machine_name']['value']: str = 'main'
    template['trr_number']['value']: str = 'main'

    return get_total_mains(template)


def get_total_mains(template: dict) -> dict:
    """
    Get total of main fields needed to be filled.
    """
    main_fields: int = 0
    for field_name in template:
        state: str = template[field_name]['value']

        if state == 'main':
            main_fields += 1

    template['main_fields']['value'] = main_fields

    return template


def get_template_pipe_to_shipment(start_location: str, end_location: str) -> dict:
    """
    Get request template based on start and end locations.
    """
    template: dict = get_template_dictionary(start_location, end_location)

    template['part_number']['value']: str = 'main'
    template['pipe_number']['value']: str = 'main'
    template['machine_name']['value']: str = 'main'
    template['purchase_order']['value']: str = 'main'

    return get_total_mains(template)


def get_template_image(start_location: str, end_location: str) -> dict:
    """
    Get request template based on start and end locations.
    """
    template: dict = get_template_dictionary(start_location, end_location)

    template['task_name']['value']: str = 'main'
    template['part_number']['value']: str = 'main'

    return get_total_mains(template)


def get_locations_template(basic_info: dict) -> dict:
    """
    Get the excel template based off of start location and end location.
    """
    start: str = basic_info['start'].upper()
    end: str = basic_info['end'].upper()

    if start == 'PIPE' and end == 'CAGE':
        return get_template_pipe_cage(start, end)

    elif start == 'PIPE' and end == 'SHIPMENT':
        return get_template_pipe_to_shipment(start, end)

    elif start == 'CAGE' and end == 'PIPE':
        return get_template_pipe_cage(start, end)

    elif start == 'CAGE' and end == 'SHIPMENT':
        return get_template_pipe_cage(start, end)

    elif start == 'SHIPMENT' and end == 'CAGE':
        return get_template_pipe_to_shipment(start, end)

    elif start == 'SHIPMENT' and end == 'PIPE':
        return get_template_pipe_to_shipment(start, end)

    elif basic_info["letter"] == 'I':
        return get_template_image(start, 'Picture Area')

    else:
        return get_template_pipe_cage(start, end)


def add_excel_colors(template: dict, worksheet) -> None:
    """
    Add colors based on template given.
    """
    for field_name in template:
        value = str(template[field_name]['value'])
        cell: str = template[field_name]['cell']
        main_cell: str = cell.replace('$C', '$B')

        if cell == '$C$20' or cell == '$C$23':
            worksheet[main_cell].fill = PatternFill(start_color="A5A5A5", fill_type="solid")
            worksheet[cell].fill = PatternFill(start_color="DBDBDB", fill_type="solid")

        elif value == 'optional':
            worksheet[main_cell].fill = PatternFill(start_color="A5A5A5", fill_type="solid")
            worksheet[cell].fill = PatternFill(start_color="DBDBDB", fill_type="solid")


def add_field_data(template: dict, worksheet) -> None:
    """
    Add field based on template given.
    """
    valid_fields: list = []

    for field_name in template:
        value = str(template[field_name]['value'])
        cell: str = template[field_name]['cell']

        if cell == '$F$10':
            continue

        if cell == '$C$20' or cell == '$C$23':
            worksheet[cell].value = value

        elif value != 'optional':
            default = template[field_name]['default']

            worksheet[cell].value = str(default)
            worksheet[cell].font = Font(bold=True)
            validation: str = f'COUNTIFS({cell},"<>{default}",{cell},"<>")'
            valid_fields.append(validation)

    worksheet['D10'].value = f'={"+".join(valid_fields)}'


def add_excel_data(sheet_name: str, basic_info: dict, worksheet) -> dict:
    """
    Adds info to amount of inventory movement and location.
    """
    template: dict = get_locations_template(basic_info)

    main_fields = str(template['main_fields']['value'])
    add_excel_colors(template, worksheet)
    add_field_data(template, worksheet)

    worksheet['D7']: str = f'=IF($D$10={main_fields},"Save and Exit. Check Pipe Cleaner.","")'
    worksheet['F10']: str = main_fields

    add_default_values(basic_info, worksheet, sheet_name)
    update_database_request_form(basic_info)

    return template


def add_default_values(basic_info: dict, worksheet, sheet_name: str) -> None:
    """
    Add default values to the excel form.
    """
    start_location: str = basic_info['start']
    end_location: str = basic_info['end']

    worksheet['D5']: str = f'{sheet_name.upper()} FORM - {start_location.upper()} TO {end_location.upper()}'
    worksheet['K2']: str = strftime('%m/%d/%Y')
    worksheet['K3']: str = strftime('%I:%M %p')
    worksheet['K4']: str = basic_info['name']
    worksheet['K5']: str = basic_info['location']
    worksheet['K6']: str = basic_info['request_number']
    worksheet['K7']: str = basic_info['version']
    worksheet['K10'] = int(basic_info['quantity'])


def update_database_request_form(basic_info: dict) -> None:
    """
    Add default values to the excel form.
    """
    request_number: str = basic_info['request_number']

    request_category: str = get_request_form_category(basic_info)
    request_document = access_database_document('request_forms', request_category)

    new_entry: dict = {'_id': request_number,
                       'basic_info': {'start': basic_info['start'],
                                      'end': basic_info['end'],
                                      'user': basic_info['name'],
                                      'location': basic_info['location'],
                                      'version': basic_info['version'],
                                      'quantity': int(basic_info['quantity']),
                                      'date': strftime('%m/%d/%Y'),
                                      'time': strftime('%I:%M %p'),
                                      'seconds': strftime('%S')},
                       'excel_data': {}}
    request_document.insert_one(new_entry)


def clean_empty_fields(field_data: str):
    """
    Assure database update fields avoid default data or optional data.  Want to be consistent and have none.
    """
    clean_field_name = str(field_data).upper()

    if clean_field_name == 'OPTIONAL':
        return 'None'

    elif clean_field_name == 'COPY FROM TASK':
        return 'None'

    else:
        return field_data


def update_request_excel_data(request_fields: dict, basic_info: dict) -> None:
    """
    Add default values to the excel form.
    """
    request_number: str = basic_info['request_number']

    request_category: str = get_request_form_category(basic_info)
    request_document = access_database_document('request_forms', request_category)

    task_name: str = get_clean_field_data(request_fields, 'task_name')
    part_number: str = get_clean_field_data(request_fields, 'part_number')
    pipe_number: str = get_clean_field_data(request_fields, 'pipe_number')
    machine_name: str = get_clean_field_data(request_fields, 'machine_name')
    trr_number: str = get_clean_field_data(request_fields, 'trr_number')
    purchase_order: str = get_clean_field_data(request_fields, 'purchase_order')
    notes: str = get_clean_field_data(request_fields, 'notes')

    excel_data: dict = {'task_name': task_name,
                        'part_number': part_number,
                        'pipe_number': pipe_number,
                        'machine_name': machine_name,
                        'trr_number': trr_number,
                        'purchase_order': purchase_order,
                        'notes': notes,
                        'date': strftime('%m/%d/%Y'),
                        'time': strftime('%I:%M %p'),
                        'seconds': strftime('%S')}

    request_document.update_one({"_id": request_number},
                                {"$set": {"excel_data": excel_data}},
                                upsert=False)


def get_clean_field_data(request_fields: dict, field_name: str) -> str:
    """
    Assure that clean data enters database and puts consistent None respond if excel field is empty.
    """
    return clean_empty_fields(request_fields[field_name])


def handle_new_file(basic_info: dict) -> load_workbook:
    """
    Delete file. Start new excel file.
    """
    template_source: str = basic_info['request_template']

    try:
        print(f'\n\tRequest Form...')
        return load_workbook(template_source)

    except FileNotFoundError:
        return load_workbook(template_source)

    except PermissionError:
        print(f'\n\tWARNING!!!')
        print(f'\tWARNING!!!')
        print(f'\tWARNING!!!')
        print(f'\n\tClose down request_form to request inventory movement.')
        input(f'\tPress enter to exit.')
        sys.exit()


def clean_default_username(default_user_name: str) -> str:
    """
    Remove unnecessary characters from name to be clean and presentable in the excel document.
    """
    if not default_user_name or default_user_name.isdigit():
        return 'None'

    else:
        return default_user_name.upper().replace('.', ' ').replace('-EXT', '').title().strip()


def get_request_file_name(user_data: dict) -> str:
    """
    Return file name based on the given start_location and end_location.
    """
    start_location: str = user_data['start']
    end_location: str = user_data['end']

    start: str = start_location.lower().strip().replace(' ', '_')
    end: str = end_location.lower().strip().replace(' ', '_')

    return f'{start}_to_{end}.xlsx'


def get_backup_name() -> str:
    """

    """
    date: str = strftime('%m/%d/%Y').replace('/', '')
    time: str = strftime('%I:%M %p').replace(':', '').replace(' ', '')

    return f'request_{date}_{time}'


def get_month_name() -> str:
    """

    """
    month: str = strftime('%m/%d/%Y')[0:2]

    if month == '01':
        return 'january'

    elif month == '02':
        return 'february'

    elif month == '03':
        return 'march'

    elif month == '04':
        return 'april'

    elif month == '05':
        return 'may'

    elif month == '06':
        return 'june'

    elif month == '07':
        return 'july'

    elif month == '08':
        return 'august'

    elif month == '09':
        return 'september'

    elif month == '10':
        return 'october'

    elif month == '11':
        return 'november'

    elif month == '12':
        return 'december'


def merge_basic_info(user_data: dict) -> dict:
    """
    Request inventory amount and gather basic documents info.
    """
    current_month: str = get_month_name()

    user_data['request_file_name'] = f'logs/{get_backup_name()}'
    user_data['name'] = clean_default_username(user_data['name'])
    user_data['request_file'] = 'request_template.xlsx'
    user_data['month'] = current_month
    user_data['request_number'] = get_request_number(user_data)
    user_data['request_template'] = 'settings/request_template.xlsx'
    user_data['location'] = str(user_data['location']).replace(' Lab Site', '')

    return user_data


def setup_excel_output(basic_info: dict) -> dict:
    """
    For launching excel sheet.
    """
    workbook: load_workbook = handle_new_file(basic_info)

    template: dict = add_basic_documents_info(workbook, basic_info, 'request')

    try:
        workbook.save(f'request_form.xlsx')
        return template

    except PermissionError:
        print_line_divider('Problem: request_form.xlsx already open.')
        print(f'\tSolution: Close request_form.xlsx and run Pipe Cleaner again.')
        input(f'\n\tPress enter to close:')
        sys.exit()


def print_line_divider(new_section_title: str) -> None:
    """
    Print terminal line divider.
    """
    print(f'\n\n\t{"-" * 60}\n\n')
    print(f'\t{new_section_title}')


def open_excel_file_for_inventory_forms(file_name: str) -> None:
    """
    Automatically opens excel file.
    """
    system(f'start EXCEL.EXE {file_name}')


def close_message(file_name: str) -> None:
    """
    Close message when excel closed.
    """
    print(f'\t\t- {file_name} closed')


def is_excel_file_running(file_name: str) -> bool:
    """
    Check if excel file is running through the task manager.
    """
    for application in task_manager():
        if 'EXCEL.EXE' in application.name().upper():
            try:
                for excel_file in application.as_dict()['cmdline']:
                    if file_name in excel_file:
                        return True

            except ProcessLookupError:
                close_message(file_name)
                return False
            except NoSuchProcess:
                close_message(file_name)
                return False
    else:
        close_message(file_name)
        return False


def is_excel_file_closed(file_name: str, form_type: str) -> bool:
    """
    Pipe Cleaner continues to check excel file runtime until it terminates by the user.
    """
    print(f'\t\t- {file_name} opened')
    print(f'\t\t- Fill out {form_type} form...\n')

    while True:
        if not is_excel_file_running(file_name):
            return True


def is_field_correct(field_input: str) -> bool:
    """
    Checks if valid input for a given field or is empty.
    """
    try:
        field_input: str = field_input.upper()

        if not field_input:
            return False

        elif field_input == '':
            return False

        elif field_input == 'COPY TASK TITLE':
            return False

        elif field_input == 'REQUIRED':
            return False

        elif field_input == 'COPY FROM TASK':
            return False

        elif field_input == 'RACK / STORAGE / OFFSITE':
            return False

        elif field_input == 'OPTIONAL':
            return False

        elif field_input == 'COPY FROM PACKAGE':
            return False

        elif field_input == 'INVENTORY SUPERVISOR':
            return False

        elif field_input == 'SCAN MATERIAL P/N':
            return False

        elif field_input == 'SCAN RACK / STORAGE / OFFSITE':
            return False

        else:
            return True

    except AttributeError:
        return False


def clean_field(task_name: str) -> str:
    """
    Cleans data.
    """
    return task_name.strip()


def respond_request_form_wrong(request_fields: dict) -> None:
    """
    Message printed if request form is filled incorrectly.
    """
    bad_message: str = 'Missing or Invalid'

    start_location: str = request_fields['start_location'].upper()
    end_location: str = request_fields['end_location'].upper()

    if 'SHIPMENT' in start_location or 'SHIPMENT' in end_location:
        print_request_field_missing(bad_message, request_fields, 'part_number')
        print_request_field_missing(bad_message, request_fields, 'pipe_number')
        print_request_field_missing(bad_message, request_fields, 'machine_name')
        print_request_field_missing(bad_message, request_fields, 'purchase_order')

    else:
        print_request_field_missing(bad_message, request_fields, 'task_name')
        print_request_field_missing(bad_message, request_fields, 'machine_name')
        print_request_field_missing(bad_message, request_fields, 'part_number')
        print_request_field_missing(bad_message, request_fields, 'pipe_number')
        print_request_field_missing(bad_message, request_fields, 'trr_number')


def print_request_field_missing(bad_message: str, request_fields: dict, field_name: str) -> None:
    """
    Print fields out if necessary fields are empty.
    """
    field_title: str = field_name.replace('_', ' ').title()

    if not is_field_correct(request_fields[field_name]):
        print(f'\t\t\t- {field_title}: {bad_message}\n')


def respond_update_form_wrong(update_fields: dict) -> None:
    """
    Message printed if update form is filled incorrectly.
    """
    bad_message: str = 'Invalid or Missing'

    if not is_field_correct(update_fields['approved_by']):
        print(f'\t\t\t- Approved By: {bad_message}')

    if not is_field_correct(update_fields['task_name']):
        print(f'\t\t\t- Task Name: {bad_message}')

    if not is_field_correct(update_fields['part_number']):
        print(f'\t\t\t- Part Number: {bad_message}')

    if not is_field_correct(update_fields['start_location']):
        print(f'\t\t\t- Start Location: {bad_message}')

    if not is_field_correct(update_fields['end_location']):
        print(f'\t\t\t- End Location: {bad_message}')

    if not is_all_serial_numbers_scanned(update_fields):
        print(f'\t\t\t- Serial Numbers Scanned: Not Enough Scanned')

    print(f'')


def get_request_field_inputs(file_name: str) -> dict:
    """
    Get request fields information.
    """
    worksheet = load_workbook(file_name)['Request']

    return {'task_name': worksheet['C13'].value,
            'part_number': worksheet['C17'].value,
            'start_location': worksheet['C20'].value,
            'end_location': worksheet['C23'].value,
            'notes': worksheet['C26'].value,
            'pipe_number': worksheet['C29'].value,
            'machine_name': worksheet['C32'].value,
            'trr_number': worksheet['C35'].value,
            'purchase_order': worksheet['C38'].value,
            'quantity': worksheet['K10'].value,
            'form_number': worksheet['K6'].value}


def get_update_fields(request_form: dict, update_worksheet) -> dict:
    """
    Get update fields information.
    """
    quantity: str = request_form['basic_info']['quantity']
    machine_name: str = request_form['excel_data']['machine_name']
    notes: str = request_form['excel_data']['notes']
    pipe_number: str = request_form['excel_data']['pipe_number']
    purchase_order: str = request_form['excel_data']['purchase_order']
    task_name: str = request_form['excel_data']['task_name']
    trr_number: str = request_form['excel_data']['trr_number']

    return {'approved_by': update_worksheet['C13'].value,
            'part_number': update_worksheet['C16'].value,
            'start_location': update_worksheet['C19'].value,
            'end_location': update_worksheet['C22'].value,
            'task_name': task_name,
            'notes': clean_optional_field(notes),
            'pipe_number': clean_optional_field(pipe_number),
            'machine_name': clean_optional_field(machine_name),
            'trr_number': clean_optional_field(trr_number),
            'purchase_order': clean_optional_field(purchase_order),
            'current_quantity': get_serial_numbers(update_worksheet),
            'total_quantity': quantity,
            'date': update_worksheet['K2'].value,
            'time': update_worksheet['K3'].value,
            'user': update_worksheet['K4'].value,
            'site': update_worksheet['K5'].value,
            'form': update_worksheet['K6'].value,
            'version': update_worksheet['K7'].value}


def clean_optional_field(field_input: str) -> str:
    """
    Get rid of optional or empty of optional fields.
    """
    upper_field_input = str(field_input).upper()

    if 'OPTIONAL' in upper_field_input:
        return 'None'
    elif upper_field_input == '':
        return 'None'
    else:
        return field_input


def get_serial_numbers(worksheet: load_workbook) -> list:
    """
    Get serial numbers from Update Form sheet.
    """
    print(f'\t\t- Collecting Serial Numbers')
    serial_numbers: list = []
    for number in range(12, 1012):
        current_value: str = worksheet[f'H{number}'].value

        if not current_value or current_value == '' or current_value == 'Scan Here':
            pass
        else:
            serial_numbers.append(current_value)

    return serial_numbers


def print_notification_receipt(request_fields: dict, basic_info: dict) -> None:
    """
    Print in terminal the receipt of the notification that will be sent to the inventory team from the request form.
    """
    print_line_divider('Request Form:')
    print(f'\t\t- Quantity: {request_fields["quantity"]}\n')

    for index, field_name in enumerate(request_fields, start=0):

        if index == 4:
            print(f'')
        if 'QUANTITY' in field_name.upper():
            continue

        field_input: str = request_fields[field_name]

        if is_field_correct(field_input):
            field_name: str = field_name.replace('_', ' ').title()
            print(f'\t\t- {field_name}: {field_input}')

        else:
            field_name: str = field_name.replace('_', ' ').title()
            print(f'\t\t- {field_name}: None')

    notification_response: str = get_inventory_notification_response()
    process_notification_input(notification_response, request_fields, basic_info)


def get_inventory_notification_response() -> str:
    """
    For email notification on inventory.
    """
    print_line_divider('Notify Inventory Team:')
    print(f'\t\tY  ->  Yes')
    print(f'\t\tN  ->  No')
    return input(f'\n\tResponse: ')


def add_email_body(request_fields: dict) -> str:
    """
    Add the inventory body of the email.
    """
    start_location: str = request_fields['start_location']
    end_location: str = request_fields['end_location']

    if 'SHIPMENT' in start_location or 'SHIPMENT' in end_location:
        get_shipment_email_message(request_fields)

    elif 'CAGE' in start_location or 'PICTURE AREA' in end_location:
        get_image_request_email(request_fields)

    else:
        return get_request_email_message(request_fields)


def get_request_email_message(request_fields: dict) -> str:
    """
    Get non-shipment email.
    """
    body_message: str = 'Inventory Team,'
    body_message += f'\nNew request form has been made for inventory transaction.  Use Pipe Cleaner to update ' \
                    f'(command "U") to enter form number to update inventory movement.'
    body_message += f'\nForm Number: {request_fields["form_number"]}'
    body_message += f'\n\nInventory Request:'
    body_message += f'\n\t- Quantity: {request_fields["quantity"]}\n'

    for index, field in enumerate(request_fields, start=0):

        if field.upper() == 'QUANTITY' or field.upper() == 'FORM NUMBER':
            continue

        field_input: str = request_fields[field]
        field_name: str = field.replace('_', ' ').title()

        if index == 4:
            body_message += f'\n'

        if is_field_correct(field_input):
            body_message += f'\n\t- {field_name}: {field_input}'
        else:
            body_message += f'\n\t- {field_name}: None'

    return body_message


def get_shipment_email_message(request_fields: dict) -> str:
    """
    Get non-shipment email.
    """
    body_message: str = 'Inventory Team,'
    body_message += f'\nNew request form has been made for inventory shipment.  No need to confirm.  ' \
                    f'\nUse Pipe Cleaner to update ' \
                    f'(command "U") to enter form number to update inventory movement.'
    body_message += f'\nForm Number: {request_fields["form_number"]}'
    body_message += f'\n\nInventory Request:'
    body_message += f'\n\t- Quantity: {request_fields["quantity"]}\n'

    for index, field in enumerate(request_fields, start=0):

        if field.upper() == 'QUANTITY' or field.upper() == 'FORM NUMBER':
            continue

        field_input: str = request_fields[field]
        field_name: str = field.replace('_', ' ').title()

        if index == 4:
            body_message += f'\n'

        if is_field_correct(field_input):
            body_message += f'\n\t- {field_name}: {field_input}'
        else:
            body_message += f'\n\t- {field_name}: None'

    return body_message


def get_image_request_email(request_fields: dict) -> str:
    """
    Get non-shipment email.
    """
    body_message: str = 'Inventory Team,'
    body_message += f'\nNew request form has been made for taking images for VSS.  Please respond to user via email' \
                    f'to confirm.'
    body_message += f'\nForm Number: {request_fields["form_number"]}'
    body_message += f'\n\nInventory Request:'
    body_message += f'\n\t- Quantity: {request_fields["quantity"]}\n'

    for index, field in enumerate(request_fields, start=0):

        if field.upper() == 'QUANTITY' or field.upper() == 'FORM NUMBER':
            continue

        field_input: str = request_fields[field]
        field_name: str = field.replace('_', ' ').title()

        if index == 4:
            body_message += f'\n'

        if is_field_correct(field_input):
            body_message += f'\n\t- {field_name}: {field_input}'
        else:
            body_message += f'\n\t- {field_name}: None'

    return body_message


def email_inventory_request(request_fields: dict) -> None:
    """
    Sends email to people responsible dealing with inventory.
    """
    print(f'\t\t- Writing email to Inventory_Kirkland@veritasdcservices.com')
    location: str = 'Inventory_Kirkland@veritasdcservices.com'
    # location: str = 'joe.ton@VeritasDCservices.com'

    create_email_notification(location, request_fields)


def end_request_form():
    print(f'\t\t- Email sent and excel data sent to database.')
    print(f'\n\tWait for a response from inventory team via email to confirm from Inventory Team.')
    input(f'\tPress enter to exit:')
    sys.exit()


def create_email_notification(person_location: str, request_fields: dict) -> None:
    outlook = client.Dispatch('Outlook.Application')
    message = outlook.CreateItem(0)
    message.To = person_location

    start_location: str = request_fields.get("start_location", "None")
    end_location: str = request_fields.get("end_location", "None")

    if 'SHIPMENT' in start_location.upper() or 'SHIPMENT' in end_location.upper():
        form_number: str = request_fields['form_number']
        message.Subject = f'{start_location} to {end_location} - {form_number}'

    else:
        message.Subject = request_fields['task_name']

    message.Body = add_email_body(request_fields)
    message.Send()


def process_notification_input(user_response: str, request_fields: dict, basic_info: dict) -> None:
    """
    Handle the response to notification on whether to send to inventory team.
    """
    user_response: str = user_response.upper()

    if user_response == 'Y':
        email_inventory_request(request_fields)
        update_request_excel_data(request_fields, basic_info)
        end_request_form()

    elif 'YES' in user_response:
        email_inventory_request(request_fields)
        update_request_excel_data(request_fields, basic_info)
        end_request_form()

    elif user_response == 'N':
        print(f'\n\tNotification Stopped.')
        input(f'\tPress enter to exit Pipe Cleaner.')
        sys.exit()

    elif 'NO' in user_response:
        print(f'\n\tNotification Stopped.')
        input(f'\tPress enter to exit Pipe Cleaner.')
        sys.exit()


def validate_request_form(file_name: str, template: dict, basic_info: dict) -> None:
    """
    Check if inventory request form done correctly.
    """
    print(f'\t\t- Checking request form...')

    request_fields_inputs: dict = get_request_field_inputs(file_name)
    fields_filled: bool = is_essential_fields_filled(request_fields_inputs, template)

    # Sometimes excel application does not save properly.  Gives excel time to register new version.
    sleep(2)

    if fields_filled:
        print_notification_receipt(request_fields_inputs, basic_info)

    else:
        print(f'\n-------------------------------------------------------------------')
        print(f'\n\t\t*** INVALID REQUEST FORM ***\n')
        respond_request_form_wrong(request_fields_inputs)

        print(f'-------------------------------------------------------------------')
        print(f'\t\t* Need to fix inventory request form...')
        input(f'\t\t- Press Enter to open request form: ')

        open_excel_file_for_inventory_forms(file_name)
        start_request_stage(file_name, template, basic_info)


def get_update_form_data(fields_data: dict) -> dict:
    """
    Get data from update form to log in the cloud database.
    """
    serial_number_logs: dict = {}

    serial_numbers_entered: list = fields_data['current_quantity']

    for serial_number in serial_numbers_entered:
        serial_number_logs[serial_number]: dict = {}
        current: dict = serial_number_logs[serial_number]

        current['approved_by'] = fields_data['approved_by']
        current['date'] = fields_data['date']
        current['end_location'] = fields_data['end_location']
        current['form'] = fields_data['form']
        current['machine_name'] = fields_data['machine_name']
        current['notes'] = fields_data['notes']
        current['part_number'] = fields_data['part_number']
        current['pipe_number'] = fields_data['pipe_number']
        current['pipe_number'] = fields_data['pipe_number']
        current['purchase_order'] = fields_data['purchase_order']
        current['site'] = fields_data['site']
        current['start_location'] = fields_data['start_location']
        current['task_name'] = fields_data['task_name']
        current['time'] = fields_data['time']
        current['total_quantity'] = fields_data['total_quantity']
        current['trr_number'] = fields_data['trr_number']
        current['user'] = fields_data['user']
        current['version'] = fields_data['version']

    return serial_number_logs


def check_update_form(file_name: str, request_form: dict, form_number: str) -> None:
    """
    Check if inventory request form done correctly.
    """
    print(f'\t\t- Checking update form')

    fields_data: dict = get_update_fields(request_form, load_workbook(f'update_form.xlsx')['Update'])

    if is_update_fields_filled(fields_data):
        print_line_divider('Log Serial Numbers and Inventory Transaction:')
        print(f'\t\tY  ->  Yes')
        print(f'\t\tN  ->  No')
        response: str = input(f'\n\tResponse: ').upper()

        if 'Y' == response:
            update_serial_numbers(fields_data)
            update_transactions(fields_data, form_number)
            print(f'\n\tSuccess: Updates to Serial Numbers and Transactions database...')
            input(f'\tPress enter to exit: ')
            sys.exit()

        elif 'N' == response:
            print(f'\tNot updated to database.')
            input(f'\tPress enter to exit: ')
            sys.exit()

    else:
        print(f'\n-------------------------------------------------------------------')
        print(f'\n\t\t*** INVALID UPDATE FORM ***')
        respond_update_form_wrong(fields_data)
        print(f'-------------------------------------------------------------------')
        print(f'\n\t\t* Need to fix inventory update form...')
        input(f'\t\t- Press enter to open excel:')
        open_excel_file_for_inventory_forms(file_name)
        check_update_form(file_name, request_form, form_number)


def update_serial_numbers(fields_data: dict) -> None:
    """
    Update serial numbers based off of update form's fields.
    """
    update_data: dict = get_update_form_data(fields_data)
    document = access_database_document('serial_numbers', 'all')

    for serial_number in update_data:

        db_serial_numbers: dict = document.find_one({'_id': serial_number})

        if not db_serial_numbers:
            entry: dict = get_new_serial_number_log(fields_data, serial_number)

            document.insert_one(entry)

        else:
            transaction: dict = get_transaction_for_serial_number_log(fields_data)
            document.update_one({"_id": serial_number},
                                {"$push": {"transactions": transaction}},
                                upsert=False)


def get_current_transactions_count(document) -> int:
    all_transactions: list = document.find({})

    count: int = 0
    for item in all_transactions:
        count += 1
    return count


def update_transactions(fields_data: dict, form_number: str) -> None:
    """
    Update serial numbers based off of update form's fields.
    """
    document = access_database_document('transactions', '021')
    transactions_count: int = get_current_transactions_count(document)
    new_count = str(transactions_count + 1)

    transaction_log: dict = get_transaction_log(fields_data, new_count, form_number)

    document.insert_one(transaction_log)


def get_transaction_log(fields_data: dict, _id: str, form_number: str) -> dict:
    """
    Data being submitted to database as new inventory transaction log.
    """
    transaction_log: dict = {"_id": str(_id),
                             "part_number": fields_data['part_number'],
                             "scanned": fields_data['current_quantity'],
                             "time": {},
                             "location": {},
                             "source": {}}

    transaction_log['time']['date_entry'] = fields_data['date']
    transaction_log['time']['time_entry'] = fields_data['time']
    transaction_log['time']['date_logged'] = strftime('%m/%d/%Y')
    transaction_log['time']['time_logged'] = strftime('%I:%M %p')

    transaction_log['location']['site'] = fields_data['site']
    transaction_log['location']['current'] = fields_data['end_location']
    transaction_log['location']['previous'] = fields_data['start_location']
    transaction_log['location']['rack'] = 'None'
    transaction_log['location']['machine'] = fields_data['machine_name']
    transaction_log['location']['pipe'] = fields_data['pipe_number']

    transaction_log['source']['approved_by'] = fields_data['approved_by']
    transaction_log['source']['verified_by'] = fields_data['user']
    transaction_log['source']['version'] = fields_data['version']
    transaction_log['source']['task'] = fields_data['task_name']
    transaction_log['source']['trr'] = fields_data['trr_number']
    transaction_log['source']['comment'] = fields_data['notes']
    transaction_log['source']['form_number'] = form_number

    return transaction_log


def get_new_serial_number_log(fields_data: dict, _id: str) -> dict:
    """
    Data being submitted to database as new inventory transaction log.
    """
    transaction_log: dict = {"_id": str(_id),
                             "part_number": fields_data['part_number'],
                             "transactions": []}

    transaction: dict = get_transaction_for_serial_number_log(fields_data)
    transaction_log["transactions"].append(transaction)

    return transaction_log


def get_transaction_for_serial_number_log(fields_data: dict) -> dict:
    """
    Transaction for serial numbers.
    """
    return {"time": {"time_entry": fields_data['time'],
                     "date_entry": fields_data['date'],
                     "time_logged": strftime('%m/%d/%Y'),
                     "date_logged": strftime('%I:%M %p')},

            "location": {"site": fields_data['site'],
                         "current": fields_data[
                             'end_location'],
                         "previous": fields_data[
                             'start_location'],
                         "rack": "None",
                         "machine": fields_data[
                             'machine_name'],
                         "pipe": fields_data[
                             'pipe_number']},

            "source": {"approved_by": fields_data['approved_by'],
                       "verified_by": fields_data['user'],
                       "version": fields_data['version'],
                       "task": fields_data['task_name'],
                       "trr": fields_data['trr_number'],
                       "comment": fields_data['notes']}}


def get_year_month():
    """
    Get year month order for database key call.
    """
    month: str = strftime('%m/%d/%Y')[0:2]
    year: str = strftime('%m/%d/%Y')[6:10]
    return f'{year}_{month}'


def is_essential_fields_filled(request_fields_input: dict, template: dict) -> bool:
    """
    Check main fields
    """
    main_fields: int = template['main_fields']['value']
    # start_location: str = request_fields_input['start_location'].upper()
    # end_location: str = request_fields_input['end_location'].upper()

    count: int = 0
    for field_name in template:

        try:
            value: str = template[field_name]['value']
            field_input: str = request_fields_input[field_name]

            if 'main' in value and field_input != 'Copy From Task':
                count += 1

        except KeyError:
            pass

    if main_fields == count:
        return True

    else:
        return False


def is_update_fields_filled(request_fields: dict) -> bool:
    """
    Checks to make sure the four essential fields are filled in for the request form.
    """
    is_quantity_match: bool = is_all_serial_numbers_scanned(request_fields)

    if is_field_correct(request_fields['approved_by']) and \
            is_field_correct(request_fields['part_number']) and \
            is_field_correct(request_fields['start_location']) and \
            is_field_correct(request_fields['end_location']) and \
            is_field_correct(request_fields['task_name']) and \
            is_quantity_match:
        return True

    else:
        return False


def is_all_serial_numbers_scanned(request_fields: dict) -> bool:
    """
    Are all the serial numbers scanned properly. Do they match the request with the total scanned?
    """
    total_quantity = int(request_fields['total_quantity'])
    current_quantity = int(len(request_fields['current_quantity']))

    if total_quantity == current_quantity:
        return True
    elif total_quantity != current_quantity:
        return False


def setup_update_form(request_form: dict) -> None:
    """
    Setup up the fields to have consistent look.
    """
    import json
    foo = json.dumps(request_form, sort_keys=True, indent=4)
    print(foo)
    input()

    form_number: str = request_form['_id']

    date: str = request_form['basic_info']['date']
    end: str = request_form['basic_info']['end']
    location: str = request_form['basic_info']['location']
    quantity: str = request_form['basic_info']['quantity']
    seconds: str = request_form['basic_info']['seconds']
    start: str = request_form['basic_info']['start']
    time: str = request_form['basic_info']['time']
    user: str = request_form['basic_info']['user']
    version: str = request_form['basic_info']['version']

    excel_date: str = request_form['excel_data']['date']  # TODO
    machine_name: str = request_form['excel_data']['machine_name']
    notes: str = request_form['excel_data']['notes']
    part_number: str = request_form['excel_data']['part_number']
    pipe_number: str = request_form['excel_data']['pipe_number']
    purchase_order: str = request_form['excel_data']['purchase_order']
    seconds: str = request_form['excel_data']['seconds']
    task_name: str = request_form['excel_data']['task_name']
    excel_time: str = request_form['excel_data']['time']
    trr_number: str = request_form['excel_data']['trr_number']

    update_workbook: load_workbook = load_workbook(f'settings/update_template.xlsx')
    update_worksheet = update_workbook['Update']

    update_worksheet['C13'] = 'Inventory Supervisor'
    update_worksheet['C16'] = 'Scan Material P/N'
    update_worksheet['C19'] = 'Scan Rack / Storage / Offsite'
    update_worksheet['C22'] = 'Scan Rack / Storage / Offsite'

    update_worksheet['D5'] = f'{start.upper()} TO {end.upper()} - UPDATE FORM'

    update_worksheet['C25'] = task_name
    update_worksheet['C28'] = notes
    update_worksheet['C29'] = pipe_number
    update_worksheet['C30'] = machine_name
    update_worksheet['C31'] = trr_number
    update_worksheet['C32'] = purchase_order

    update_worksheet['K2'] = excel_date
    update_worksheet['K3'] = excel_time
    update_worksheet['K4'] = user
    update_worksheet['K5'] = location
    update_worksheet['K6'] = form_number
    update_worksheet['K7'] = version
    update_worksheet['M10'] = quantity

    try:
        update_workbook.save(f'update_form.xlsx')

    except PermissionError:
        print(f'\tPermission Denied: update_template.xlsx is already open')
        print(f'\tPlease close file and restart Pipe Cleaner')
        input(f'\tPress enter to close:')
        sys.exit()


def start_request_stage(file_name: str, template: dict, basic_info: dict) -> None:
    """
    Start the request stage before notifying inventory team
    """
    file_closed: bool = is_excel_file_closed(file_name, 'request')

    if file_closed:
        validate_request_form(file_name, template, basic_info)


def start_update_form(form_number: str, basic_data: dict) -> None:
    """
    After sending email to inventory team. Have update form ready.
    """
    file_name: str = 'update_form.xlsx'
    site = str(basic_data['site']).replace(' Lab Site', '').replace('Kirkland', '0')

    current_year: str = strftime('%m/%d/%Y')[8:10]
    document = access_database_document('request_forms', f'{site}{current_year}')
    request_form: dict = document.find_one({'_id': form_number})

    setup_update_form(request_form)
    open_excel_file_for_inventory_forms(file_name)

    if is_excel_file_closed(file_name, 'update'):
        check_update_form(file_name, request_form, form_number)


def get_request_number(user_data: dict) -> str:
    """
    Get next request number
    """
    request_category: str = get_request_form_category(user_data)
    document = access_database_document('request_forms', request_category)
    all_requests = document.find({})

    return get_current_request_number(all_requests, request_category)


def get_request_form_category(user_data: dict) -> str:
    """
    Get the request number first three digits. ie. site-year
    """
    current_year: str = strftime('%m/%d/%Y')[8:10]
    current_location: str = user_data['location']

    if 'KIRKLAND' in current_location.upper():
        return f'0{current_year}'


def get_current_request_number(all_requests, request_category: str) -> str:
    """
    Get total request forms stored in database.
    """
    total_requests = int(all_requests.count()) + 1
    request_string = str(total_requests)
    request_length = len(str(request_string))
    request_first_part = request_category[0:3]

    if request_length == 1:
        return f'{request_first_part}-000-00{request_string}'

    elif request_length == 2:
        request_part: str = request_string[0:2]

        return f'{request_first_part}-000-0{request_part}'

    elif request_length == 3:
        request_part: str = request_string[0:3]

        return f'{request_first_part}-000-{request_part}'

    elif request_length == 4:
        request_part_1: str = request_string[0:1]
        request_part_2: str = request_string[1:4]

        return f'{request_first_part}-00{request_part_1}-{request_part_2}'

    elif len(request_string) == 5:
        request_part_1: str = request_string[0:2]
        request_part_2: str = request_string[2:5]

        return f'{request_first_part}-0{request_part_1}-{request_part_2}'

    elif len(request_string) == 6:
        request_part_1: str = request_string[0:3]
        request_part_2: str = request_string[3:6]

        return f'{request_first_part}-{request_part_1}-{request_part_2}'


def start_main_method_for_request_form(user_data: dict) -> None:
    """
    Starting point for handling inventory requests sent to inventory team.
    """
    basic_info: dict = merge_basic_info(user_data)
    template: dict = setup_excel_output(basic_info)

    open_excel_file_for_inventory_forms('request_form.xlsx')
    start_request_stage('request_form.xlsx', template, basic_info)
