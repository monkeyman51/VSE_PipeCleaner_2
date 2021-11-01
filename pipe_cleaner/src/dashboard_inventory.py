from time import strftime

import xlsxwriter

import openpyxl


def process_pipe_name(pipe_name: str):
    """
    Shorten pipe name to fit into excel output
    :param pipe_name:
    :return:
    """
    clean_data: str = pipe_name. \
        replace('[', ''). \
        replace(']', ''). \
        replace("'", '')

    last_part: str = clean_data.split(' ')[-1]

    return clean_data.replace('Pipe-', '').replace(last_part, '')


def check_missing(data: str) -> str:
    """

    :param data:
    :return:
    """
    if data == 'None' or data == '' or data is None:
        return 'None'
    else:
        return data


def get_current_color(index, structure):
    result: int = index % 2
    if result == 0:
        return structure.blue_middle_22
    elif result == 1:
        return structure.alt_blue_middle_22


def get_current_color_11(index, structure):
    result: int = index % 2
    if result == 0:
        return structure.blue_middle
    elif result == 1:
        return structure.alt_blue_middle


def get_total_tickets(pipe_name, processed_console_server) -> int:
    total_tickets: int = 0
    pipe_total_trr: int = len(processed_console_server.get(pipe_name, {}).get('group_unique_tickets'))

    return increment_total_tickets(pipe_total_trr, total_tickets)


def increment_total_tickets(pipe_total_trr, total_tickets):
    if pipe_total_trr == 0:
        total_tickets += 1
    else:
        total_tickets += pipe_total_trr
    return total_tickets


def get_base_position(column, current_position) -> str:
    return f'{column}{current_position}'


def get_merge_position(base_position, letter, max_number):
    max_position: str = get_max_position(letter, max_number)
    return f'{base_position}:{max_position}'


def get_max_position(letter, max_number) -> str:
    return f'{letter}{max_number}'


def add_empty_merge_cell(merge_position, structure, worksheet):
    worksheet.merge_range(merge_position, '', structure.aqua_left_12)


def add_empty_hyperlink_cell(base_position: str, pipe_hyperlink: str, current_setup: dict):
    worksheet: xlsxwriter = current_setup.get('worksheet')
    structure: xlsxwriter = current_setup.get('structure')
    worksheet.write_url(base_position, pipe_hyperlink, structure.missing_cell, string='')


def add_pipe_name_column(clean_console_server: dict, current_setup: dict, console_server_data: dict):
    """
    Writes Pipe Name column in excel output
    """
    letter: str = 'C'
    worksheet: xlsxwriter = current_setup.get('worksheet')
    structure: xlsxwriter = current_setup.get('structure')
    current_position: int = current_setup.get('body_position')

    for index, pipe_name in enumerate(sorted(clean_console_server), start=0):
        all_machine_size: int = get_all_machine_size(console_server_data, pipe_name)

        if all_machine_size != 0:
            current_color: xlsxwriter = get_current_color(index, structure)
            pipe_hyperlink: str = get_pipe_hyperlink(console_server_data, pipe_name)
            clean_pipe_name: str = f' {process_pipe_name(pipe_name)}'
            pipe_size = clean_console_server[pipe_name]

            max_number: int = current_position + pipe_size - 1

            base_position: str = get_base_position(letter, current_position)
            merge_position: str = get_merge_position(base_position, letter, max_number)

            worksheet.merge_range(merge_position, clean_pipe_name, current_color)
            worksheet.write_url(base_position, pipe_hyperlink, current_color, string=clean_pipe_name)

            current_position: int = max_number + 2


def set_single_ticket_row(current_position, worksheet):
    worksheet.set_row(current_position - 1, 28.5)


def get_pipe_size(pipe_name, user_sorted_pipes):
    current_pipe = user_sorted_pipes.get(pipe_name)
    return get_current_pipe_size(current_pipe)


def get_current_pipe_size(current_pipe):
    tally: int = 0
    for item in current_pipe:
        current_machine: dict = current_pipe[item]
        if len(current_machine) == 0:
            tally += 1
        else:
            tally += len(current_machine)
    return tally


def get_pipe_hyperlink(console_server_data, pipe_name):
    host_group_id: str = console_server_data.get(pipe_name, {}).get('host_id', 'None')
    host_group_url: str = f'http://172.30.1.100/console/host_group_host_list.php?host_group_id={host_group_id}'
    return host_group_url


def count_issues_in_user_pipe(pipe_name, user_sorted_pipes):
    systems_issues: dict = user_sorted_pipes[pipe_name]

    count: int = 0
    for system_issue in systems_issues:
        bar = len(user_sorted_pipes[pipe_name][system_issue])
        count += bar
    return count


def count_systems_in_user_pipe(pipe_name, user_sorted_pipes):
    return len(user_sorted_pipes[pipe_name])


def write_hyperlink_cells(number_of_machines: int, letter: str, number_of_cell: str, max_number: int, hyperlink: str,
                          text: str, worksheet, structure, color):
    """
    Write uniform colors for excel output. Contains merge and hyperlinks
    :param structure:
    :param number_of_machines:
    :param letter:
    :param number_of_cell:
    :param max_number:
    :param hyperlink:
    :param color:
    :param text:
    :param worksheet:
    :return:
    """
    if check_missing(text) == 'None' and number_of_machines == 1:
        worksheet.write(f'{letter}{number_of_cell}', '', structure.missing_cell)

    elif check_missing(text) == 'None' and number_of_machines >= 2:
        worksheet.merge_range(f'{letter}{number_of_cell}:{letter}{max_number}',
                              '', structure.missing_cell)

    elif number_of_machines == 1:
        try:
            worksheet.write_url(f'{letter}{number_of_cell}', hyperlink, color,
                                string=text)
        except AttributeError:
            worksheet.write_url(f'{letter}{number_of_cell}', hyperlink, structure.missing_cell,
                                string='')

    elif number_of_machines >= 2:
        worksheet.merge_range(f'{letter}{number_of_cell}:{letter}{max_number}',
                              text, color)

        worksheet.write_url(f'{letter}{number_of_cell}', hyperlink, color,
                            string=text)


def add_machine_name_column(initial_point, user_sorted_pipes, worksheet, structure,
                            console_server_data, user_virtual_machines, virtual_machines_data, tally_storage):
    """
    Write host group column in excel output

    :param tally_storage:
    :param virtual_machines_data:
    :param user_virtual_machines:
    :param user_sorted_pipes:
    :param initial_point: starting point of writing column pipe names
    :param worksheet:
    :param structure:
    :param console_server_data: for host group IDs
    :return:
    """
    column: str = 'D'

    color_change: int = 0
    current_pipe_point: int = initial_point
    for index, pipe_name in enumerate(user_sorted_pipes):
        user_machines = sorted(list(user_sorted_pipes.get(pipe_name, {}).keys()))
        tally_storage['pipes'] += 1

        machine_order: int = 0
        pipe_max_size: int = current_pipe_point
        for start, machine_name in enumerate(user_machines):

            machine_issues: dict = user_sorted_pipes.get(pipe_name, {}).get(machine_name)
            machine_issues_total: int = get_machine_issues_size(machine_issues)

            max_number: int = pipe_max_size + machine_issues_total - 1
            number_of_cell = str(machine_order + current_pipe_point)

            machine_hyperlink: str = get_machine_hyperlink(console_server_data, machine_name, pipe_name)
            connection_status: str = get_machine_connection_status(console_server_data, machine_name, pipe_name)

            contrast_color = get_contrast_color(color_change, structure)

            if connection_status.upper() == 'DEAD':
                write_hyperlink_cells(machine_issues_total, column, number_of_cell, max_number, machine_hyperlink,
                                      machine_name, worksheet, structure, structure.dark_grey_middle)
                tally_storage['dead'] += 1
                tally_storage['total_machines'] += 1

            elif connection_status.upper() == 'MOSTLY_DEAD':
                write_hyperlink_cells(machine_issues_total, column, number_of_cell, max_number, machine_hyperlink,
                                      machine_name, worksheet, structure, structure.light_grey_middle)
                tally_storage['dead'] += 1
                tally_storage['total_machines'] += 1

            elif connection_status.upper() == 'ALIVE':
                write_hyperlink_cells(machine_issues_total, column, number_of_cell, max_number, machine_hyperlink,
                                      machine_name, worksheet, structure, contrast_color)
                tally_storage['alive'] += 1
                tally_storage['total_machines'] += 1

            else:
                write_hyperlink_cells(machine_issues_total, column, number_of_cell, max_number, machine_hyperlink,
                                      machine_name, worksheet, structure, structure.light_red_middle_11)

            if machine_issues_total == 1:
                actual_max_number = max_number - 1
                worksheet.set_row(actual_max_number, 19.5)

            machine_order += machine_issues_total
            pipe_max_size += machine_issues_total
            color_change += 1

        current_pipe_point: int = pipe_max_size + 1

    current_pipe_point += 1

    for index, item in enumerate(user_virtual_machines):
        contrast_color = get_contrast_color(color_change, structure)
        current_row = current_pipe_point + index
        virtual_machine_hyperlink: str = get_rdp_connection_hyperlink(item, virtual_machines_data)

        worksheet.write_url(f'{column}{current_row}', virtual_machine_hyperlink, contrast_color, string=item)
        color_change += 1
        tally_storage['virtual_machines'] += 1

    return tally_storage


def get_rdp_connection_hyperlink(item, virtual_machines):
    rdp_connection_string: str = virtual_machines.get(item, {}).get('rdp_connection_string')
    virtual_machine_hyperlink = f'http://172.30.1.100/guacamole/#/client/{rdp_connection_string}'
    return virtual_machine_hyperlink


def get_contrast_color(initial_point, structure):
    contrast_color = initial_point % 2
    if contrast_color == 0:
        return structure.blue_middle
    else:
        return structure.alt_blue_middle


def write_status_field_column(initial_point, user_sorted_pipes, worksheet, structure,
                              console_server_data, user_virtual_machines, virtual_machines_data):
    """
    Write host group column in excel output

    :param virtual_machines_data:
    :param user_virtual_machines:
    :param user_sorted_pipes:
    :param initial_point: starting point of writing column pipe names
    :param worksheet:
    :param structure:
    :param console_server_data: for host group IDs
    :return:
    """
    column: str = 'E'

    color_change: int = 0
    current_pipe_point: int = initial_point
    for index, pipe_name in enumerate(user_sorted_pipes):
        user_machines = sorted(list(user_sorted_pipes.get(pipe_name, {}).keys()))

        machine_order: int = 0
        pipe_max_size: int = current_pipe_point
        for start, machine_name in enumerate(user_machines):
            machine_issues: dict = user_sorted_pipes.get(pipe_name, {}).get(machine_name)
            machine_issues_total: int = get_machine_issues_size(machine_issues)

            machine_comment: str = get_machine_comment(console_server_data, machine_name, pipe_name)

            max_number: int = pipe_max_size + machine_issues_total - 1
            number_of_cell = str(machine_order + current_pipe_point)

            contrast_color = get_contrast_color(color_change, structure)

            write_hyperlink_cells(machine_issues_total, column, number_of_cell, max_number, '',
                                  machine_comment, worksheet, structure, contrast_color)

            machine_order += machine_issues_total
            pipe_max_size += machine_issues_total
            color_change += 1

        current_pipe_point: int = pipe_max_size + 1

    current_pipe_point += 1
    for index, item in enumerate(user_virtual_machines):
        contrast_color = get_contrast_color(color_change, structure)
        current_row = current_pipe_point + index

        virtual_machine_comment: str = virtual_machines_data.get(item, {}).get('comment')

        if not virtual_machine_comment:
            worksheet.write(f'{column}{current_row}', virtual_machine_comment, structure.missing_cell)
        else:
            worksheet.write(f'{column}{current_row}', virtual_machine_comment, contrast_color)
        color_change += 1


def get_machine_comment(console_server_data, machine_name, pipe_name):
    machine_comment: str = console_server_data.get(pipe_name, {}). \
        get('pipe_data', {}).get(machine_name, {}).get('comment', 'None')
    return machine_comment


def get_machine_issues_size(machine_issues):
    machine_issues_total: int = len(machine_issues)

    if machine_issues_total == 0:
        return 1
    else:
        return machine_issues_total


def get_machine_connection_status(console_server_data, machine_name, pipe_name):
    connection_status: str = console_server_data.get(pipe_name, {}). \
        get('pipe_data', {}).get(machine_name, {}).get('connection_status', 'None')
    return connection_status


def get_machine_hyperlink(console_server_data, machine_name, pipe_name):
    host_id: str = console_server_data.get(pipe_name, {}). \
        get('pipe_data', {}).get(machine_name, {}).get('id', 'None')
    host_url: str = f'http://172.30.1.100/console/host_details.php?host_id={host_id}'
    return host_url


def add_trr_column(initial_point, user_sorted_pipes, worksheet, structure, console_server_data):
    """
    Write Pipe Name column in excel output
    :param user_sorted_pipes:
    :param initial_point:
    :param console_server_data: for host group IDs
    :param worksheet:
    :param structure:
    :return:
    """
    column: str = 'L'

    color_change: int = 0
    current_pipe_point: int = initial_point
    for index, pipe_name in enumerate(user_sorted_pipes):
        user_machines = sorted(list(user_sorted_pipes.get(pipe_name, {}).keys()))

        machine_order: int = 0
        pipe_max_size: int = current_pipe_point
        for start, machine_name in enumerate(user_machines):
            machine_issues: dict = user_sorted_pipes.get(pipe_name, {}).get(machine_name)
            machine_issues_total: int = get_machine_issues_size(machine_issues)

            ticket_id: str = get_machine_ticket_id(console_server_data, machine_name, pipe_name)
            ticket_hyperlink: str = get_machine_ticket_hyperlink(console_server_data, machine_name, pipe_name)

            max_number: int = pipe_max_size + machine_issues_total - 1
            number_of_cell = str(machine_order + current_pipe_point)

            contrast_color = get_contrast_color(color_change, structure)

            write_hyperlink_cells(machine_issues_total, column, number_of_cell, max_number, ticket_hyperlink,
                                  ticket_id, worksheet, structure, contrast_color)

            machine_order += machine_issues_total
            pipe_max_size += machine_issues_total
            color_change += 1

        current_pipe_point: int = pipe_max_size + 1


def get_machine_ticket_id(console_server_data, machine_name, pipe_name):
    return console_server_data.get(pipe_name, {}). \
        get('pipe_data', {}).get(machine_name, {}).get('ticket', 'None')


def get_machine_ticket_hyperlink(console_server_data, machine_name, pipe_name):
    ticket_id: str = console_server_data.get(pipe_name, {}). \
        get('pipe_data', {}).get(machine_name, {}).get('ticket', 'None')
    return f'https://azurecsi.visualstudio.com/CSI%20Commodity%20Qualification/_workitems/edit/{ticket_id}'


def add_type_column(initial_point, user_sorted_pipes, worksheet, structure, console_server_data, azure_devops_data):
    """
    Write Pipe Name column in excel output
    :param azure_devops_data:
    :param user_sorted_pipes:
    :param initial_point:
    :param console_server_data: for host group IDs
    :param worksheet:
    :param structure:
    :return:
    """
    column: str = 'M'

    color_change: int = 0
    current_pipe_point: int = initial_point
    for index, pipe_name in enumerate(user_sorted_pipes):
        user_machines = sorted(list(user_sorted_pipes.get(pipe_name, {}).keys()))

        machine_order: int = 0
        pipe_max_size: int = current_pipe_point
        for start, machine_name in enumerate(user_machines):

            machine_issues: dict = user_sorted_pipes.get(pipe_name, {}).get(machine_name)
            machine_issues_total: int = get_machine_issues_size(machine_issues)

            ticket_id: str = get_machine_ticket_id(console_server_data, machine_name, pipe_name)
            ticket_hyperlink: str = get_machine_ticket_hyperlink(console_server_data, machine_name, pipe_name)

            ticket_type: str = get_ticket_type(azure_devops_data, ticket_id)

            max_number: int = pipe_max_size + machine_issues_total - 1
            number_of_cell = str(machine_order + current_pipe_point)

            contrast_color = get_contrast_color(color_change, structure)

            try:
                ticket_type: str = ticket_type.upper().replace(' TEST', '').replace('TEST', '')

                write_hyperlink_cells(machine_issues_total, column, number_of_cell, max_number, ticket_hyperlink,
                                      ticket_type, worksheet, structure, contrast_color)

                machine_order += machine_issues_total
                pipe_max_size += machine_issues_total
                color_change += 1
            except AttributeError:
                write_hyperlink_cells(machine_issues_total, column, number_of_cell, max_number, ticket_hyperlink,
                                      'None', worksheet, structure, contrast_color)

                machine_order += machine_issues_total
                pipe_max_size += machine_issues_total
                color_change += 1

            except TypeError:
                write_hyperlink_cells(machine_issues_total, column, number_of_cell, max_number, ticket_hyperlink,
                                      'None', worksheet, structure, contrast_color)

                machine_order += machine_issues_total
                pipe_max_size += machine_issues_total
                color_change += 1

        current_pipe_point: int = pipe_max_size + 1


def get_ticket_type(azure_devops_data, ticket_id):
    return azure_devops_data.get(ticket_id, {}).get('table_data', {}).get('request_type')


def clean_ticket_state(ticket_state):
    if not ticket_state:
        return ticket_state
    else:
        return ticket_state.replace('InProgress', 'In Progress'). \
            replace('Test completed', 'Test Completed'). \
            replace('Ready To Review', 'Ready to Review'). \
            replace('Ready to start', 'Ready to Start')


def get_machine_issues_count(machine_issues):
    if machine_issues == {} or len(machine_issues) == 1:
        return 1


def get_machine_issue(current_issue, machine_issues):
    return machine_issues.get(current_issue, {}).get('system_component')


def add_machine_issue_item(column, contrast_color, issue_item, row, worksheet):
    if issue_item == 'BIOS' or issue_item == 'BMC' or issue_item == 'OS' or issue_item == 'System' or \
            issue_item == 'MATCH' or issue_item == 'CPLD':
        worksheet.write(f'{column}{row}', issue_item, contrast_color)

    else:
        worksheet.write(f'{column}{row}', 'Ticket', contrast_color)


def add_machine_issue_state(column, issue_item, system_component, row, worksheet, structure):
    if 'MISMATCH' in issue_item:
        worksheet.write(f'{column}{row}', 'MISMATCH', structure.neutral_cell)
        return 'mismatch'
    elif 'Title' in system_component:
        worksheet.write(f'{column}{row}', 'INCOMPLETE', structure.light_red_middle_11)
        return 'incomplete'
    else:
        worksheet.write(f'{column}{row}', 'INCOMPLETE', structure.light_red_middle_11)
        return 'incomplete'


def get_user_machines(pipe_name, user_sorted_pipes) -> list:
    return sorted(list(user_sorted_pipes.get(pipe_name, {}).keys()))


def process_machine_names(all_machines_in_pipes: list):
    """

    :param all_machines_in_pipes:
    :return:
    """
    all_unique_machines: dict = {}
    unique_machines = sorted(list(set(all_machines_in_pipes)))
    for machine_name in unique_machines:
        all_unique_machines[machine_name] = 0

    for unique_machine in unique_machines:
        for machine_name in all_machines_in_pipes:
            if unique_machine in machine_name:
                all_unique_machines[machine_name] += 1

    return all_unique_machines


def process_unique_types(all_machines_in_pipes: list):
    """

    :param all_machines_in_pipes:
    :return:
    """
    machine_types = sorted(list(set(all_machines_in_pipes)))
    unique_machine_types: dict = {}
    all_machine_types: list = []

    for machine_name in machine_types:
        machine_type = machine_name[8:11]
        unique_machine_types[machine_type] = 0
        all_machine_types.append(machine_type)

    for machine_type in all_machine_types:
        for machine in unique_machine_types:
            if machine_type in machine:
                unique_machine_types[machine_type] += 1

    return unique_machine_types


def get_machine_issues(all_machines_in_pipes: list, issues_dict: dict):
    """

    :param issues_dict:
    :param all_machines_in_pipes:
    :return:
    """
    all_unique_machines: dict = {}
    unique_machines = sorted(list(set(all_machines_in_pipes)))
    for machine_name in unique_machines:
        all_unique_machines[machine_name] = []

    for unique_machine in unique_machines:
        for machine_name in issues_dict:
            if unique_machine in machine_name:
                all_unique_machines.get(unique_machine, []).append(machine_name)
    return all_unique_machines


def clean_console_server_data(console_server_data: dict) -> dict:
    """

    """
    real_pipes: list = get_real_pipes(console_server_data)

    process_console_server: dict = {}
    for real_pipe in sorted(real_pipes):
        group_unique_tickets: dict = console_server_data[real_pipe]['group_unique_tickets']
        all_machine_size: int = get_all_machine_size(console_server_data, real_pipe)

        process_console_server[real_pipe] = all_machine_size + len(group_unique_tickets)

    return process_console_server


def get_pipe_inventory_size(console_server_data, current_pipe, real_pipe):
    real_machines: list = get_real_machines(current_pipe)
    pipe_size: int = 0
    for machine_name in real_machines:
        unique_dimms: list = console_server_data[real_pipe]['pipe_data'][machine_name]['unique_dimms']
        unique_disks: list = console_server_data[real_pipe]['pipe_data'][machine_name]['unique_disks']
        unique_nvmes: list = console_server_data[real_pipe]['pipe_data'][machine_name]['unique_nvmes']
        machine_size: int = len(unique_nvmes) + len(unique_disks) + len(unique_dimms)
        pipe_size += machine_size
    return pipe_size


def get_all_machine_size(console_server_data, real_pipe):
    unique_dimms: dict = console_server_data[real_pipe]['pipe_data']['pipe_inventory']['dimms']
    unique_disks: dict = console_server_data[real_pipe]['pipe_data']['pipe_inventory']['disks']
    unique_nvmes: dict = console_server_data[real_pipe]['pipe_data']['pipe_inventory']['nvmes']
    return len(unique_dimms) + len(unique_disks) + len(unique_nvmes)


def get_all_machine_dimms(console_server_data: dict, real_pipe: str) -> dict:
    return console_server_data[real_pipe]['pipe_data']['pipe_inventory']['dimms']


def get_all_machine_disks(console_server_data: dict, real_pipe: str) -> dict:
    return console_server_data[real_pipe]['pipe_data']['pipe_inventory']['disks']


def get_all_machine_nvmes(console_server_data: dict, real_pipe: str) -> dict:
    return console_server_data[real_pipe]['pipe_data']['pipe_inventory']['nvmes']


def get_real_machines(current_pipe) -> list:
    real_machines: list = []
    for potential_machine in current_pipe:
        if '-VM-' not in potential_machine.upper() and 'VSE' in potential_machine.upper():
            real_machines.append(potential_machine)
    return real_machines


def get_real_pipes(console_server_data) -> list:
    real_pipes: list = []
    for potential_pipe in console_server_data:
        if 'Pipe-' in potential_pipe and 'OFFLINE' not in potential_pipe \
                and '(' not in potential_pipe and ')' not in potential_pipe:
            try:
                host_group_status: str = console_server_data.get(potential_pipe, {}).get('host_group_status')
                if 'OFFLINE' not in host_group_status.upper():
                    real_pipes.append(potential_pipe)
            except TypeError:
                pass
            except AttributeError:
                pass
    return real_pipes


def clean_cell_data(cell_data: str) -> str:
    """
    Assure standard data cleansing
    """
    if not cell_data or cell_data == ' ' or cell_data == '':
        return 'None'

    elif ' (' in cell_data:
        dirty_data: str = cell_data.split(" (")[-1]
        return cell_data.replace(dirty_data, '').replace('(', '').strip()

    else:
        return cell_data.strip()


def clean_number_cell(number_cell: str) -> int:
    """

    :param number_cell:
    :return:
    """
    if not number_cell:
        return 0
    else:
        number_cell = str(number_cell)
        if ' ' in number_cell:
            return int(number_cell.replace(' ', ''))
        else:
            return int(number_cell)


def get_inventory_data() -> dict:
    excel_path: str = r'Z:\Kirkland_Lab\PipeCleaner\inventory.xlsx'

    part_numbers: list = []
    availability: list = []

    try:
        wb_obj = openpyxl.load_workbook(excel_path)

        ws4 = wb_obj['Commodity Inventory']
        m_row = ws4.max_row

        for i in range(2, m_row + 1):
            part_number: str = ws4.cell(row=i, column=4).value
            available: str = ws4.cell(row=i, column=9).value

            clean_part: str = clean_cell_data(part_number)
            clean_available: int = clean_number_cell(available)

            part_numbers.append(clean_part)
            availability.append(clean_available)

        part_to_available: dict = {}
        for index, current_part in enumerate(list(set(part_numbers)), start=0):
            part_to_available[current_part] = 0

            for count, item in enumerate(part_numbers, start=0):
                if item in current_part:
                    part_to_available[current_part] += int(availability[count])

        return part_to_available

    except FileNotFoundError:
        other_excel_path: str = r'172.30.1.100\pxe\Kirkland_Lab\PipeCleaner\inventory.xlsx'

        part_numbers: list = []
        availability: list = []

        wb_obj = openpyxl.load_workbook(other_excel_path)

        ws4 = wb_obj['Commodity Inventory']
        m_row = ws4.max_row

        for i in range(2, m_row + 1):
            part_number: str = ws4.cell(row=i, column=4).value
            available: str = ws4.cell(row=i, column=9).value

            clean_part: str = clean_cell_data(part_number)
            clean_available: int = clean_number_cell(available)

            part_numbers.append(clean_part)
            availability.append(clean_available)

        part_to_available: dict = {}
        for index, current_part in enumerate(list(set(part_numbers)), start=0):
            part_to_available[current_part] = 0

            for count, item in enumerate(part_numbers, start=0):
                if item in current_part:
                    part_to_available[current_part] += int(availability[count])

        return part_to_available


def add_column_data(azure_devops_data: dict, console_server_data: dict, all_issues: list, current_setup: dict) -> None:
    """

    """
    clean_console_server: dict = clean_console_server_data(console_server_data)
    # inventory_data: dict = get_inventory_data()

    add_pipe_name_column(clean_console_server, current_setup, console_server_data)

    add_part_number_column(console_server_data, current_setup)

    add_section_column(clean_console_server, console_server_data, current_setup)

    add_item_column(clean_console_server, console_server_data, current_setup, azure_devops_data)

    add_console_part_number_column(clean_console_server, console_server_data, current_setup, azure_devops_data)

    add_count_column(clean_console_server, console_server_data, current_setup, azure_devops_data)

    # add_available_column(clean_console_server, console_server_data, current_setup, azure_devops_data, inventory_data)

    # add_inventory_column(console_server_data, current_setup, inventory_data)


def is_inventory_data(inventory_data: dict, current_item: str) -> str:
    if not current_item:
        return ''

    else:
        for commodity in inventory_data:
            if commodity in current_item:
                return commodity
        else:
            return ''


def add_available_column(clean_console_server: dict, console_server_data: dict, current_setup: dict,
                         azure_devops_data: dict, inventory_data) -> None:
    """

    """
    letter: str = 'I'

    worksheet: xlsxwriter = current_setup.get('worksheet')
    structure: xlsxwriter = current_setup.get('structure')
    current_position: int = current_setup.get('body_position')

    for index, pipe_number in enumerate(sorted(clean_console_server), start=0):

        current_color: xlsxwriter = get_current_color_11(index, structure)

        pipe_disks: dict = get_all_machine_disks(console_server_data, pipe_number)
        pipe_nvmes: dict = get_all_machine_nvmes(console_server_data, pipe_number)
        pipe_dimms: dict = get_all_machine_dimms(console_server_data, pipe_number)

        for disk_commodity in sorted(pipe_disks):
            base_position: str = get_base_position(letter, current_position)

            if not is_inventory_data(inventory_data, disk_commodity):
                worksheet.write(base_position, '', structure.missing_cell)
                current_position += 1

            else:
                part = is_inventory_data(inventory_data, disk_commodity)
                available_number = inventory_data[part]
                worksheet.write(base_position, available_number, current_color)
                current_position += 1

        for nvme_commodity in sorted(pipe_nvmes):
            base_position: str = get_base_position(letter, current_position)

            if not is_inventory_data(inventory_data, nvme_commodity):
                worksheet.write(base_position, '', structure.missing_cell)
                current_position += 1

            else:
                part = is_inventory_data(inventory_data, nvme_commodity)
                available_number = inventory_data[part]
                worksheet.write(base_position, available_number, current_color)
                current_position += 1

        for dimm_commodity in sorted(pipe_dimms):
            base_position: str = get_base_position(letter, current_position)

            if not is_inventory_data(inventory_data, dimm_commodity):
                worksheet.write(base_position, '', structure.missing_cell)
                current_position += 1

            else:
                part = is_inventory_data(inventory_data, dimm_commodity)
                available_number = inventory_data[part]
                worksheet.write(base_position, available_number, current_color)
                current_position += 1

        group_unique_tickets: dict = console_server_data[pipe_number]['group_unique_tickets']
        for unique_ticket in sorted(group_unique_tickets):
            base_position: str = get_base_position(letter, current_position)
            part_number: str = azure_devops_data.get(unique_ticket, {}).get('table_data', {}).get('part_number', '')

            if not part_number:
                worksheet.write(base_position, '', structure.missing_cell)
                current_position += 1
            else:
                if not is_inventory_data(inventory_data, part_number):
                    worksheet.write(base_position, '', structure.missing_cell)
                    current_position += 1

                else:
                    part = is_inventory_data(inventory_data, part_number)
                    available_number = inventory_data[part]
                    worksheet.write(base_position, available_number, current_color)
                    current_position += 1

        current_position: int = current_position + 1


def add_count_column(clean_console_server: dict, console_server_data: dict, current_setup: dict,
                     azure_devops_data: dict) -> None:
    """

    """
    letter: str = 'G'

    worksheet: xlsxwriter = current_setup.get('worksheet')
    structure: xlsxwriter = current_setup.get('structure')
    current_position: int = current_setup.get('body_position')

    for index, pipe_number in enumerate(sorted(clean_console_server), start=0):
        all_machine_size: int = get_all_machine_size(console_server_data, pipe_number)

        if all_machine_size != 0:
            current_color: xlsxwriter = get_current_color_11(index, structure)

            pipe_disks: dict = get_all_machine_disks(console_server_data, pipe_number)
            pipe_nvmes: dict = get_all_machine_nvmes(console_server_data, pipe_number)
            pipe_dimms: dict = get_all_machine_dimms(console_server_data, pipe_number)

            for disk_commodity in sorted(pipe_disks):
                base_position: str = get_base_position(letter, current_position)
                current_count: str = pipe_disks[disk_commodity]
                worksheet.write(base_position, current_count, current_color)
                current_position += 1

            for nvme_commodity in sorted(pipe_nvmes):
                base_position: str = get_base_position(letter, current_position)
                current_count: str = pipe_nvmes[nvme_commodity]
                worksheet.write(base_position, current_count, current_color)
                current_position += 1

            for dimm_commodity in sorted(pipe_dimms):
                base_position: str = get_base_position(letter, current_position)
                current_count: str = pipe_dimms[dimm_commodity]
                worksheet.write(base_position, current_count, current_color)
                current_position += 1

            group_unique_tickets: dict = console_server_data[pipe_number]['group_unique_tickets']
            for unique_ticket in sorted(group_unique_tickets):
                base_position: str = get_base_position(letter, current_position)

                worksheet.write(base_position, '', structure.missing_cell)
                current_position += 1

            current_position: int = current_position + 1


def add_console_part_number_column(clean_console_server: dict, console_server_data: dict, current_setup: dict,
                                   azure_devops_data: dict) -> None:
    """

    """
    letter: str = 'F'

    worksheet: xlsxwriter = current_setup.get('worksheet')
    structure: xlsxwriter = current_setup.get('structure')
    current_position: int = current_setup.get('body_position')

    for index, pipe_number in enumerate(sorted(clean_console_server), start=0):
        all_machine_size: int = get_all_machine_size(console_server_data, pipe_number)

        if all_machine_size != 0:
            current_color: xlsxwriter = get_current_color_11(index, structure)

            pipe_disks: dict = get_all_machine_disks(console_server_data, pipe_number)
            pipe_nvmes: dict = get_all_machine_nvmes(console_server_data, pipe_number)
            pipe_dimms: dict = get_all_machine_dimms(console_server_data, pipe_number)

            for disk_commodity in sorted(pipe_disks):
                base_position: str = get_base_position(letter, current_position)
                worksheet.write(base_position, disk_commodity, current_color)
                current_position += 1

            for nvme_commodity in sorted(pipe_nvmes):
                base_position: str = get_base_position(letter, current_position)
                worksheet.write(base_position, nvme_commodity, current_color)
                current_position += 1

            for dimm_commodity in sorted(pipe_dimms):
                base_position: str = get_base_position(letter, current_position)
                worksheet.write(base_position, dimm_commodity, current_color)
                current_position += 1

            group_unique_tickets: dict = console_server_data[pipe_number]['group_unique_tickets']
            for unique_ticket in sorted(group_unique_tickets):
                base_position: str = get_base_position(letter, current_position)
                part_number: str = azure_devops_data.get(unique_ticket, {}).get('table_data', {}).get('part_number', '')

                if not part_number:
                    worksheet.write(base_position, '', structure.missing_cell)
                    current_position += 1
                else:
                    worksheet.write(base_position, part_number, current_color)
                    current_position += 1

            current_position: int = current_position + 1


def get_request_type(azure_devops_data, unique_ticket):
    request_type: str = azure_devops_data.get(unique_ticket, {}).get('table_data', {}).get('request_type',
                                                                                           '')
    return request_type.replace(' TEST', '').replace('TEST', '')


def add_item_column(clean_console_server, console_server_data, current_setup, azure_devops_data) -> None:
    """

    """
    letter: str = 'E'

    worksheet: xlsxwriter = current_setup.get('worksheet')
    structure: xlsxwriter = current_setup.get('structure')
    current_position: int = current_setup.get('body_position')

    for index, pipe_number in enumerate(sorted(clean_console_server), start=0):
        all_machine_size: int = get_all_machine_size(console_server_data, pipe_number)

        if all_machine_size != 0:
            current_color: xlsxwriter = get_current_color_11(index, structure)

            pipe_disks: dict = get_all_machine_disks(console_server_data, pipe_number)
            pipe_nvmes: dict = get_all_machine_nvmes(console_server_data, pipe_number)
            pipe_dimms: dict = get_all_machine_dimms(console_server_data, pipe_number)

            for disk_commodity in sorted(pipe_disks):
                base_position: str = get_base_position(letter, current_position)
                worksheet.write(base_position, 'Disk', current_color)
                current_position += 1

            for nvme_commodity in sorted(pipe_nvmes):
                base_position: str = get_base_position(letter, current_position)
                worksheet.write(base_position, 'NVMe', current_color)
                current_position += 1

            for dimm_commodity in sorted(pipe_dimms):
                base_position: str = get_base_position(letter, current_position)
                worksheet.write(base_position, 'DIMM', current_color)
                current_position += 1

            group_unique_tickets: dict = console_server_data[pipe_number]['group_unique_tickets']
            for unique_ticket in sorted(group_unique_tickets):
                base_position: str = get_base_position(letter, current_position)
                request_type: str = get_request_type(azure_devops_data, unique_ticket)

                if not request_type:
                    worksheet.write(base_position, '', structure.missing_cell)
                    current_position += 1
                else:
                    worksheet.write(base_position, request_type, current_color)
                    current_position += 1

            current_position: int = current_position + 1


def add_section_column(clean_console_server, console_server_data, current_setup) -> None:
    """

    """
    letter: str = 'D'

    worksheet: xlsxwriter = current_setup.get('worksheet')
    structure: xlsxwriter = current_setup.get('structure')
    current_position: int = current_setup.get('body_position')

    for index, pipe_number in enumerate(sorted(clean_console_server), start=0):
        all_machine_size: int = get_all_machine_size(console_server_data, pipe_number)

        if all_machine_size != 0:
            pipe_size: int = clean_console_server[pipe_number]
            current_color: xlsxwriter = get_current_color_11(index, structure)
            base_position: str = get_base_position(letter, current_position)

            max_position: int = current_position + all_machine_size - 1
            merge_position = get_merge_position(base_position, letter, max_position)

            worksheet.merge_range(merge_position, 'Pipe Commodities', current_color)

            group_unique_tickets: dict = console_server_data[pipe_number]['group_unique_tickets']

            for count, unique_ticket in enumerate(sorted(group_unique_tickets), start=0):
                maximum_position: int = current_position + count + all_machine_size
                hyperlink = f'https://azurecsi.visualstudio.com/CSI%20Commodity%20Qualification/_workitems/edit/' \
                            f'{unique_ticket}'

                worksheet.write_url(f'{letter}{maximum_position}', hyperlink, current_color,
                                    string=unique_ticket)

            current_position: int = current_position + pipe_size + 1
    else:
        current_position += 1


def add_inventory_column(console_server_data: dict, current_setup: dict, inventory_data: dict) -> None:
    """

    """
    part_number_position: str = 'O'
    sum_position: str = 'P'

    worksheet: xlsxwriter = current_setup.get('worksheet')
    structure: xlsxwriter = current_setup.get('structure')
    current_position: int = current_setup.get('body_position')

    for index, disk_part_number in enumerate(sorted(inventory_data), start=0):

        current_color: xlsxwriter = get_current_color_11(index, structure)
        part_number_location: str = get_base_position(part_number_position, current_position)
        sum_location: str = get_base_position(sum_position, current_position)

        worksheet.write(part_number_location, disk_part_number, current_color)
        worksheet.write(sum_location, inventory_data[disk_part_number], current_color)

        current_position: int = current_position + 1
    else:
        current_position += 1


def add_part_number_column(console_server_data, current_setup) -> None:
    """

    """
    type_position: str = 'K'
    part_number_position: str = 'L'
    sum_position: str = 'M'

    worksheet: xlsxwriter = current_setup.get('worksheet')
    structure: xlsxwriter = current_setup.get('structure')
    current_position: int = current_setup.get('body_position')

    disks: dict = console_server_data.get('inventory', {}).get('commodities', {}).get('disks')
    nvmes: dict = console_server_data.get('inventory', {}).get('commodities', {}).get('nvmes')
    dimms: dict = console_server_data.get('inventory', {}).get('commodities', {}).get('dimms')

    for index, disk_part_number in enumerate(sorted(disks), start=0):

        current_color: xlsxwriter = get_current_color_11(index, structure)
        type_location: str = get_base_position(type_position, current_position)
        part_number_location: str = get_base_position(part_number_position, current_position)
        sum_location: str = get_base_position(sum_position, current_position)

        disk_sum: int = disks[disk_part_number]

        worksheet.write(type_location, 'Disk', current_color)
        worksheet.write(part_number_location, disk_part_number, current_color)
        worksheet.write(sum_location, disk_sum, current_color)

        current_position: int = current_position + 1
    else:
        current_position += 1

    for index, nvme_part_number in enumerate(sorted(nvmes), start=0):

        current_color: xlsxwriter = get_current_color_11(index, structure)
        type_location: str = get_base_position(type_position, current_position)
        part_number_location: str = get_base_position(part_number_position, current_position)
        sum_location: str = get_base_position(sum_position, current_position)

        nvme_sum: int = nvmes[nvme_part_number]

        worksheet.write(type_location, 'NVMe', current_color)
        worksheet.write(part_number_location, nvme_part_number, current_color)
        worksheet.write(sum_location, nvme_sum, current_color)

        current_position: int = current_position + 1
    else:
        current_position += 1

    for index, dimm_part_number in enumerate(sorted(dimms), start=0):
        current_color: xlsxwriter = get_current_color_11(index, structure)
        type_location: str = get_base_position(type_position, current_position)
        part_number_location: str = get_base_position(part_number_position, current_position)
        sum_location: str = get_base_position(sum_position, current_position)

        dimm_sum: int = dimms[dimm_part_number]

        worksheet.write(type_location, 'DIMM', current_color)
        worksheet.write(part_number_location, dimm_part_number, current_color)
        worksheet.write(sum_location, dimm_sum, current_color)

        current_position: int = current_position + 1


def get_tally_storage():
    tally_storage: dict = {'dead': 0,
                           'alive': 0,
                           'match': 0,
                           'mismatch': 0,
                           'incomplete': 0,
                           'empty': 0,
                           'total_machines': 0,
                           'virtual_machines': 0,
                           'pipes': 0}
    return tally_storage


def clean_user_sorted_pipes(all_issues, user_sorted_pipes):
    count: int = 0
    for user_pipe in user_sorted_pipes:
        for user_machine in user_sorted_pipes[user_pipe]:
            for issue_system in all_issues:
                issue_machine: str = issue_system.get('machine_name')

                if issue_machine in user_machine:
                    user_sorted_pipes[user_pipe][issue_machine][count] = issue_system
                    count += 1
    return user_sorted_pipes


def get_sorted_pipes_and_systems(user_systems, user_unique_pipes):
    """

    :param user_systems:
    :param user_unique_pipes:
    :return:
    """
    user_sorted_pipes: dict = get_user_sorted_pipes(user_unique_pipes)
    for user_system in user_systems:
        pipe_name: str = user_systems.get(user_system, {}).get('pipe_name')
        user_sorted_pipes[pipe_name][user_system] = {}

    return user_sorted_pipes


def get_user_sorted_pipes(user_unique_pipes) -> dict:
    user_sorted_pipes: dict = {}
    for unique_pipe in user_unique_pipes:
        user_sorted_pipes[unique_pipe] = {}
    return user_sorted_pipes


def get_user_pipes_issues(processed_issues: dict, user_systems: dict, user_unique_pipes: list) -> dict:
    """
    Get user assigned machines that have issues sorted within pipes as dictionary
    :param processed_issues: All issues
    :param user_systems:
    :param user_unique_pipes:
    :return:
    """
    pipes_issues: dict = {}

    for user_pipe in user_unique_pipes:
        try:
            pipe_issues: dict = processed_issues[user_pipe]

            pipes_issues[user_pipe] = {}
            count: int = 0
            for user_system in user_systems:
                for pipe_issue in pipe_issues:

                    if user_system in pipe_issue:
                        current_pipe_issue: dict = processed_issues[user_pipe][pipe_issue]
                        pipes_issues[user_pipe][count] = current_pipe_issue
                        count += 1
        except KeyError:
            pipes_issues[user_pipe] = ''
    return pipes_issues


def create_breakdown_graph(console_server_data: dict, workbook: xlsxwriter, worksheet: xlsxwriter, sheet_name: str,
                           mismatch_tally: str, missing_tally: str):
    """
    Create Graph for Issues
    :param console_server_data:
    :param workbook:
    :param worksheet:
    :param sheet_name:
    :param mismatch_tally:
    :param missing_tally:
    :return:
    """
    vse_log: int = console_server_data.get('host_groups_data', {}).get('vse_log', 0)

    # Add the worksheet data that the charts will refer to.
    headings: list = ['Number', 'Tallies']
    data = [
        ['Azure DevOps (TRRs)', 'Comparison', 'Veritas Engineering & Services'],
        [int(missing_tally), int(mismatch_tally), int(vse_log)],
    ]

    # Write to excel output to hold data for graph, bolded Title
    bold = workbook.add_format({'bold': 1})
    worksheet.write_row('A1', headings, bold)
    worksheet.write_column('A2', data[0])
    worksheet.write_column('B2', data[1])

    # Type of Graph
    chart_structure = workbook.add_chart({'type': 'bar'})

    # Structure Graph
    chart_structure.add_series({
        'name': "='" + sheet_name + "'!$B$1",
        'categories': "='" + sheet_name + "'!$A$2:$A$4",
        'values': "='" + sheet_name + "'!$B$2:$B$4",
        'points': [
            {'fill': {'color': '#7030A0'}},
            # {'fill': {'color': '#FF0000'}},
            {'fill': {'color': '#DCAA1B'}},
            {'fill': {'color': '#31869B'}},  # Aqua Color
        ],
    })

    # Configure a second series. Note use of alternative syntax to define ranges.
    chart_structure.add_series({
        'name': [f"{sheet_name}", 0, 2],
        'categories': [f"{sheet_name}", 1, 0, 3, 0],
        'values': [f"{sheet_name}", 1, 2, 3, 2],
    })

    # Add a chart title and some axis labels.
    chart_structure.set_title({'name': 'Breakdown of Issues'})

    # Chart Style of Graph
    chart_structure.set_style(11)
    chart_structure.set_legend({'none': True})

    # Size of Chart
    worksheet.insert_chart('J1', chart_structure, {'x_scale': 2.350, 'y_scale': 0.84})


class ExcelStructure:
    def __init__(self, current_setup: dict, console_server_data: dict):
        self.current_setup: dict = current_setup
        self.console_server_data: dict = console_server_data

        self.worksheet: xlsxwriter = self.current_setup.get('worksheet')
        self.structure: xlsxwriter = self.current_setup.get('structure')

        self.sheet_title: str = self.current_setup.get('sheet_title')
        self.site_location: str = self.current_setup.get('site_location')
        self.default_user_name: str = self.current_setup.get('default_user_name')
        self.version: str = self.current_setup.get('version')
        self.header_height: str = self.current_setup.get('header_height')


def set_sheet_structure(current_setup: dict) -> None:
    """
    Create dashboard structure
    """
    set_excel_design(current_setup)
    add_header_data(current_setup)


def add_header_data(current_setup: dict) -> None:
    """
    Add header data on ex. username, date, version, etc.
    """
    add_header_user_name(current_setup)
    add_header_sheet_title(current_setup)
    add_header_site_location(current_setup)
    add_header_date_and_version(current_setup)
    add_header_items_under_testing(current_setup)


def add_header_site_location(current_setup: dict) -> None:
    """
    Adds the site location to the header area in the top left corner
    """
    site_location: str = current_setup.get('site_location')

    worksheet: xlsxwriter = current_setup.get('worksheet')
    structure: xlsxwriter = current_setup.get('structure')

    worksheet.write('B6', f'        {site_location}', structure.bold_italic_blue_font)


def add_header_sheet_title(current_setup: dict) -> None:
    """
    Adds the excel sheet name to the header area in the top left corner
    """
    sheet_title: str = current_setup.get('sheet_title')

    worksheet: xlsxwriter = current_setup.get('worksheet')
    structure: xlsxwriter = current_setup.get('structure')

    worksheet.write('B5', f'  Pipe Cleaner - {sheet_title}', structure.big_blue_font)


def set_excel_design(current_setup: dict) -> None:
    """
    Set up excel output design/parameters.
    """
    set_rows_and_columns_sizes(current_setup)

    add_column_titles(current_setup)
    add_freeze_panes(current_setup)
    add_vse_logo_top_right(current_setup)


def add_header_user_name(current_setup: dict):
    """
    Add clean user name to the top left corner.
    """
    clean_name: str = current_setup.get('clean_name')

    worksheet: xlsxwriter = current_setup.get('worksheet')
    structure: xlsxwriter = current_setup.get('structure')

    worksheet.write('B7', f'            {clean_name}', structure.bold_italic_blue_font)


def get_user_virtual_machines(user_info):
    return user_info['virtual_machines']


def get_user_pipe_total(user_info) -> int:
    user_systems: dict = get_user_systems(user_info)
    user_pipes: list = get_user_pipes(user_systems)
    user_unique_pipes: list = get_user_unique_pipes(user_pipes)
    return len(user_unique_pipes)


def get_user_unique_pipes(user_pipes):
    return sorted(list(set(user_pipes)))


def get_user_unique_pipes(user_pipes):
    return sorted(list(set(user_pipes)))


def add_header_date_and_version(current_setup: dict) -> None:
    """
    Adds the current date/time and Pipe Cleaner version to the header area in the top left corner
    """
    current_time: str = strftime('%I:%M %p')
    current_date: str = strftime('%m/%d/%Y')
    pipe_cleaner_version: str = current_setup.get('version')

    worksheet: xlsxwriter = current_setup.get('worksheet')
    structure: xlsxwriter = current_setup.get('structure')

    pipe_cleaner_version: str = clean_pipe_cleaner_version(pipe_cleaner_version)

    worksheet.write('B8', f'            {current_date} - {current_time} - {pipe_cleaner_version}',
                    structure.italic_blue_font)


def add_header_items_under_testing(current_setup: dict) -> None:
    """
    These items under testing are meant to be components still not 100% confident
    """
    worksheet: xlsxwriter = current_setup.get('worksheet')
    structure: xlsxwriter = current_setup.get('structure')
    header_height: xlsxwriter = current_setup.get('header_height')
    upper_header: str = header_height - 1

    worksheet.merge_range(f'C{upper_header}:G{upper_header}', f'VSE - Console Server',
                          structure.teal_middle_14)
    worksheet.write(f'I{upper_header}', f'Outdated', structure.light_red_middle_14)
    worksheet.merge_range(f'K{upper_header}:M{upper_header}', f'Console Server', structure.teal_middle_14)
    worksheet.merge_range(f'O{upper_header}:P{upper_header}', f'Outdated', structure.light_red_middle_14)


def get_user_pipes(user_systems):
    all_pipes: list = []
    for item in user_systems:
        if 'VSE' in item and '-' in item:
            all_pipes.append(user_systems[item]['pipe_name'])
    return all_pipes


def get_user_systems(user_info) -> dict:
    return user_info['systems']


def get_user_info_alt(console_server_data, default_name) -> dict:
    default_name_underscore: str = default_name_period_to_underscore(default_name)
    # import json
    # foo = json.dumps(console_server_data['user_base'], sort_keys=True, indent=4)
    # print(foo)
    # input()

    try:
        for user_name in console_server_data['user_base']:
            alt_name = str(console_server_data['user_base'][user_name]['alt_name']).lower()

            if default_name.lower() in alt_name or default_name_underscore in user_name:
                return console_server_data['user_base'][user_name]
        else:
            return {}

    except KeyError:
        return {}


def default_name_period_to_underscore(default_name):
    if 'steph' in default_name and '.ak' in default_name:
        return 'steph_ak'
    else:
        return default_name.replace('.', '_').replace('-EXT', '')


def add_vse_logo_top_right(current_setup: dict) -> None:
    """
    Creates VSE Logo on the top left corner
    :param current_setup:
    """
    worksheet: xlsxwriter = current_setup.get('worksheet')

    worksheet.insert_image('A1', 'pipe_cleaner/img/vsei_logo.png')


def clean_pipe_cleaner_version(pipe_cleaner_version) -> str:
    """
    Version for documentation
    :param pipe_cleaner_version:
    :return: cleaner version
    """
    return f"v{pipe_cleaner_version.split(' ')[0]}"


def add_freeze_panes(current_setup: dict) -> None:
    """
    Allows information to the left to stay
    """
    header_height: int = current_setup.get('header_height')
    worksheet: xlsxwriter = current_setup.get('worksheet')

    worksheet.freeze_panes(header_height, 7)


def add_column_titles(current_setup: dict) -> None:
    """
    Set up Column Names in the Excel table for categorizing into vertical data later
    """
    header_height: int = current_setup.get('header_height')
    left_padding: int = current_setup.get('left_padding')
    column_names: tuple = current_setup.get('column_names')

    worksheet: xlsxwriter = current_setup.get('worksheet')
    structure: xlsxwriter = current_setup.get('structure')

    # Hyperlink to Host Group page within Console Server, should be for column title "Pipe"
    host_group_column: str = get_letter_for_column_position(initial=0, left_padding=2)

    for index, column_title in enumerate(column_names, start=0):
        position: str = get_column_title_position(header_height, index, left_padding)

        if 'PIPE' in column_title.upper() and host_group_column in position[0]:
            worksheet.write_url(position, 'http://172.30.1.100/console/host_groups.php',
                                structure.teal_middle, column_title)

        elif not column_title:
            add_white_cell(position, current_setup)

        else:
            add_column_title(position, column_title, current_setup)


def add_column_title(position: str, column_title: str, current_setup: dict) -> None:
    """
    Add column title to the current excel sheet
    """
    worksheet: xlsxwriter = current_setup.get('worksheet')
    structure: xlsxwriter = current_setup.get('structure')

    worksheet.write(position, column_title, structure.teal_middle)


def get_column_title_position(header_height: int, index: int, left_padding: int) -> str:
    """
    Get position of the column title based on excel position from the letter and number ex. A1, B4, C3
    """
    letter: str = get_letter_for_column_position(index, left_padding)
    return f'{letter}{header_height}'


def add_white_cell(position: str, current_setup) -> None:
    """
    Account for empty cells that don't have column title.
    Meant for giving space between different groups of data.
    """
    worksheet: xlsxwriter = current_setup.get('worksheet')
    structure: xlsxwriter = current_setup.get('structure')

    worksheet.write(position, '', structure.white)


def get_letter_for_column_position(initial: int, left_padding: int) -> str:
    """
    For positioning the column title based on starting point of the left padding.
    :return: letter of excel column
    """
    return convert_index_to_letter(initial + left_padding)


def set_rows_and_columns_sizes(current_setup) -> None:
    """
    Beginning of the Excel Structure
    """
    rows_height: tuple = current_setup.get('rows_height')
    columns_width: tuple = current_setup.get('columns_width')

    worksheet: xlsxwriter = current_setup.get('worksheet')
    structure: xlsxwriter = current_setup.get('structure')

    set_header_rows_height(rows_height, worksheet, structure)
    set_excel_column_width(columns_width, worksheet, structure)


def set_header_rows_height(rows_height: tuple, worksheet: xlsxwriter, structure: xlsxwriter) -> None:
    """
    Establishes current worksheet row heights for the header.
    """
    for index, row_size in enumerate(rows_height, start=0):
        worksheet.set_row(index, row_size, structure.white)


def set_excel_column_width(columns_width: tuple, worksheet: xlsxwriter, structure: xlsxwriter) -> None:
    """
    Establishes current worksheet column widths.
    """
    for index in range(0, len(columns_width)):
        current_letter: str = convert_index_to_letter(index)

        worksheet.set_column(f'{current_letter}:{current_letter}',
                             columns_width[index],
                             structure.white)


def convert_index_to_letter(index: int) -> str:
    """

    :param index: Current index due to how many columns we care about in the excel output sheet
    """
    lower_character = chr(ord('a') + index)
    return str(lower_character).upper()


def remove_excel_green_corners(current_setup) -> None:
    """
    Excel sometimes have green corners within a cell. Removes to clear up look of excel output.
    :param current_setup: Current worksheet
    """
    worksheet = current_setup.get('worksheet')

    worksheet.ignore_errors({'number_stored_as_text': 'A1:XFD1048576'})


def create_personal_issues_sheet(excel_setup: dict) -> dict:
    """
    Current excel sheet design to setup the excel tab for data to fill in later.
    """
    workbook: xlsxwriter = excel_setup.get('workbook')

    excel_setup['host_group_hyperlink']: str = 'http://172.30.1.100/console/host_groups.php'
    excel_setup['sheet_title']: str = 'Inventory'

    excel_setup['worksheet']: xlsxwriter = workbook.add_worksheet(excel_setup.get('sheet_title'))

    excel_setup['rows_height']: tuple = (12.0, 19.5, 19.5, 18.0, 20.25, 20.25, 20.25, 20.25, 3.75, 3.75, 3.75, 18.75)

    excel_setup['columns_width']: tuple = (0.5, 0.5, 21.0, 19.0, 11.0, 40.0, 12.0, 1.86, 16.0, 10.0, 9.0, 35.0, 10.0,
                                           10.0, 35.0, 10.0, 10.0, 10.0, 10.0, 10.0, 10.0, 10.0)

    excel_setup['column_names']: tuple = ('Pipe',
                                          'Section',
                                          'Item',
                                          'Part Number',
                                          'Current',
                                          '',
                                          'Available',
                                          '',
                                          'Type',
                                          'Part Number',
                                          'Sum',
                                          '',
                                          'Part Number',
                                          'Sum')

    return excel_setup


def main_method(azure_devops_data: dict, console_server_data: dict, excel_setup: dict, all_issues: list) -> None:
    """
    Create Personal Issues
    """
    current_setup: dict = create_personal_issues_sheet(excel_setup)

    set_sheet_structure(current_setup)

    add_column_data(azure_devops_data, console_server_data, all_issues, current_setup)

    remove_excel_green_corners(current_setup)
