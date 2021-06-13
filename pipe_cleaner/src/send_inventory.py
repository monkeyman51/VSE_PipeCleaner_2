import shutil
import os
import sys

from time import strftime
from openpyxl import load_workbook
from openpyxl.styles import Alignment

from colorama import Fore, Style
import win32com.client as client


def get_file_path(user_name: str, root_path: str, file_number: str) -> str:
    """
    Get file name dedicated towards documenting / logging data entry for inventory.  Plan to have file names
    represented as name_date_time
    :return:
    """
    return fr'{root_path}\{file_number}_{user_name}.xlsx'


def get_date_file_name(default_user_name):
    clean_name: str = default_user_name. \
        replace('.', ''). \
        lower()

    year: str = strftime('%Y')[-2:]
    current_date: str = strftime('%m%d')
    current_time: str = strftime('%I%M%S')
    return f'{current_date}{year}_{current_time}_{clean_name}.xlsx'


def get_backup_path(default_user_name: str) -> str:
    """
    Have backup file path in case original file path is corrupt.
    """
    file_name: str = get_date_file_name(default_user_name)
    current_date: str = strftime('%m%d')
    current_time: str = strftime('%I%M%S')
    granular_file_name: str = f'{current_date}_{current_time}_{file_name}.xlsx'

    return fr'Z:\Kirkland_Lab\Users\joe_ton\project_inventory\back_transaction_logs\{granular_file_name}'


def get_clean_user_name(default_user_name):
    if '.' not in default_user_name:
        indexed_clean_name: list = []

        for index, character in enumerate(default_user_name, start=0):
            index_character: str = default_user_name[index]

            if index_character.isupper() and not index == 0:
                indexed_clean_name.append(' ')
                indexed_clean_name.append(index_character)
            else:
                indexed_clean_name.append(index_character)

        return ''.join(indexed_clean_name).replace('-Ext', '').replace(' ', '').replace('_', '').lower()
    else:

        return default_user_name.replace('.', ' ').title().replace('-Ext', '').replace(' ', '').replace('_', '').lower()


def get_file_number(root_file_path: str) -> str:
    """
    Need to provide new unique name for documenting transaction logs.
    """
    try:
        raw_file_names: list = os.listdir(root_file_path)

        if len(raw_file_names) == 1 and raw_file_names[0] == '_inventory_transactions.xlsx':
            return '00001'

        elif len(raw_file_names) == 0:
            return '00001'

        else:
            raw_numbers: list = []
            for raw_file_name in raw_file_names:
                raw_numbers.append(raw_file_name[0:5])

            non_zero_names: list = []
            for index, raw_number in enumerate(raw_numbers, start=0):
                non_zero_characters: list = []

                flip: int = 0
                for character in raw_number:
                    if character != 0 and flip == 0:
                        non_zero_characters.append(character)
                        flip += 1
                    elif character != 0 and flip >= 1:
                        non_zero_characters.append(character)

                complete_name: str = ''.join(non_zero_characters)
                try:
                    non_zero_names.append(int(complete_name))
                except ValueError:
                    pass

            clean_non_zero_names: list = sorted(list(set(non_zero_names)))
            latest_file_number = int(clean_non_zero_names[-1])
            latest_file_number += 1

        return get_file_name_number(latest_file_number)

    except FileNotFoundError:

        alt_path = r'\\172.30.1.100\pxe\Kirkland_Lab\PipeCleaner\transaction_logs'
        raw_file_names: list = os.listdir(alt_path)

        if len(raw_file_names) == 0:
            return '00001'
        else:
            raw_numbers: list = []
            for raw_file_name in raw_file_names:
                raw_numbers.append(raw_file_name[0:5])

            non_zero_names: list = []
            for index, raw_number in enumerate(raw_numbers, start=0):
                non_zero_characters: list = []

                flip: int = 0
                for character in raw_number:
                    if character != 0 and flip == 0:
                        non_zero_characters.append(character)
                        flip += 1
                    elif character != 0 and flip >= 1:
                        non_zero_characters.append(character)

                complete_name: str = ''.join(non_zero_characters)
                non_zero_names.append(int(complete_name))

            clean_non_zero_names = sorted(list(set(non_zero_names)))
            latest_file_number = int(clean_non_zero_names[-1])
            latest_file_number += 1

        return get_file_name_number(latest_file_number)


def get_file_name_number(latest_file_number):
    number_length: int = len(str(latest_file_number))

    if number_length == 1:
        return f'0000{latest_file_number}'

    elif number_length == 2:
        return f'000{latest_file_number}'

    elif number_length == 3:
        return f'00{latest_file_number}'

    elif number_length == 4:
        return f'0{latest_file_number}'

    elif number_length == 5:
        return latest_file_number


def read_inventory_file(inventory_file_name: str) -> list:
    """
    Read inventory excel file to print in the terminal output.  Used for confirming user the content before sending.
    :return:
    """
    try:
        excel_sheet = load_workbook(filename=inventory_file_name)['Inventory Template']

        inventory_transaction: dict = get_inventory_transaction_data(excel_sheet)
        inventory_rows: dict = get_inventory_rows(inventory_transaction)

        filled_rows: list = []
        for row in inventory_rows:
            row_data = inventory_rows[row]

            count: int = 0
            for cell_data in row_data:
                try:
                    if not cell_data or 'Fill Here' in cell_data:
                        count += 1
                except TypeError:
                    if not str(cell_data) or 'Fill Here' in str(cell_data):
                        count += 1

            if count < 10:
                filled_rows.append(row_data)

        return filled_rows

    except FileNotFoundError:
        print(f'\n\tDo {Fore.RED}Inventory Mode{Style.RESET_ALL} first...')
        input(f'\tPress enter to close Pipe Cleaner')
        sys.exit()


def get_inventory_transaction_data(excel_sheet):
    inventory_transaction: dict = {'from': get_column_data('E', excel_sheet),
                                   'to': get_column_data('F', excel_sheet),
                                   'type': get_column_data('G', excel_sheet),
                                   'part_number': get_column_data('H', excel_sheet),
                                   'qty': get_column_data('I', excel_sheet),
                                   'supplier': get_column_data('J', excel_sheet),
                                   'pipe_number': get_column_data('K', excel_sheet),
                                   'task_number': get_column_data('L', excel_sheet),
                                   'po_number': get_column_data('M', excel_sheet),
                                   'notes': get_column_data('N', excel_sheet)}
    return inventory_transaction


def get_inventory_rows(inventory_transaction) -> dict:
    inventory_rows: dict = {'row_01': [], 'row_02': [], 'row_03': [], 'row_04': [], 'row_05': [], 'row_06': [],
                            'row_07': [], 'row_08': [], 'row_09': [], 'row_10': [], 'row_11': [], 'row_12': []}

    for column_name in inventory_transaction:
        column_data: dict = inventory_transaction[column_name]

        inventory_rows['row_01'].append(column_data[0])
        inventory_rows['row_02'].append(column_data[1])
        inventory_rows['row_03'].append(column_data[2])
        inventory_rows['row_04'].append(column_data[3])
        inventory_rows['row_05'].append(column_data[4])
        inventory_rows['row_06'].append(column_data[5])
        inventory_rows['row_07'].append(column_data[6])
        inventory_rows['row_08'].append(column_data[7])
        inventory_rows['row_09'].append(column_data[8])
        inventory_rows['row_10'].append(column_data[9])
        inventory_rows['row_11'].append(column_data[10])
        inventory_rows['row_12'].append(column_data[11])

    return inventory_rows


def get_column_data(letter: str, sheet) -> list:
    column_data: list = []

    count: int = 10
    while count < 23:
        column_data.append(sheet[f'{letter}{count}'].value)
        count += 1

    return column_data


def print_file_data(inventory_file: list) -> list:
    """
    Print in terminal output in Pipe Cleaner to showcase the filled in data already before finalizing send
    data to Z: Drive
    """
    for index, row in enumerate(inventory_file, start=1):
        print(f'\n\tInventory Transaction #{index}:')

        if not row[0] or row[0] == 'Fill Here':
            print(f'\t\tFrom:      {Fore.RED}{row[0]}{Style.RESET_ALL}')
        else:
            print(f'\t\tFrom:      {row[0]}')

        if not row[1] or row[1] == 'Fill Here':
            print(f'\t\tTo:        {Fore.RED}{row[1]}{Style.RESET_ALL}')
        else:
            print(f'\t\tTo:        {row[1]}')

        if not row[2] or row[2] == 'Fill Here':
            print(f'\t\tType:      {Fore.RED}{row[2]}{Style.RESET_ALL}')
        else:
            print(f'\t\tType:      {row[2]}')

        if not row[3] or row[3] == 'Fill Here':
            print(f'\t\tPart #:    {Fore.RED}{row[3]}{Style.RESET_ALL}')
        else:
            print(f'\t\tPart #:    {row[3]}')

        if not row[4] or row[4] == 'Fill Here':
            print(f'\t\tQTY:       {Fore.RED}{row[4]}{Style.RESET_ALL}')
        else:
            print(f'\t\tQTY:       {row[4]}')

        if not row[5] or row[5] == 'Fill Here':
            print(f'\t\tSupplier:  {Fore.RED}{row[5]}{Style.RESET_ALL}')
        else:
            print(f'\t\tSupplier:  {row[5]}')

        if not row[6] or row[6] == 'Fill Here':
            print(f'\t\tPipe #:    {Fore.RED}{row[6]}{Style.RESET_ALL}')
        else:
            print(f'\t\tPipe #:    {row[6]}')

        if not row[7] or row[7] == 'Fill Here':
            print(f'\t\tTask # :   {Fore.RED}{row[7]}{Style.RESET_ALL}')
        else:
            print(f'\t\tTask #:    {row[7]}')

        if not row[8] or row[8] == 'Fill Here':
            print(f'\t\tPO #:      {Fore.RED}{row[8]}{Style.RESET_ALL}')
        else:
            print(f'\t\tPO #:      {row[8]}')

        if not row[9] or row[9] == 'Fill Here':
            print(f'\t\tNotes:     {Fore.RED}{row[9]}{Style.RESET_ALL}')
        else:
            print(f'\t\tNotes:     {row[9]}')

    return inventory_file


def get_user_send_response() -> str:
    """
    Get input from user regarding to send current inventory data from excel.
    """
    print(f'\n\tChoose between these options...')
    print(f'\ty -> Yes - Send data to Z: Drive')
    print(f'\tn -> No - Do not send data to Z: Drive')
    return input(f'\n\tChoose option: ')


def is_authorized_user(default_user_name: str, inventory_authorized: tuple) -> bool:
    """
    Checks if current user is apart of the inventory authorized users.
    :param default_user_name: current user
    :param inventory_authorized: authorized users
    :return: True or False
    """
    clean_name: str = default_user_name.casefold()

    for authorized_user in inventory_authorized:
        first_name: str = authorized_user[0].casefold()
        last_name: str = authorized_user[1].casefold()

        if first_name in clean_name and last_name in clean_name:
            return True

    else:
        return False


def get_last_entry_row(excel_sheet) -> int:
    index: int = 10
    for row in range(10_000):
        date_value = str(excel_sheet[f'D{index}'].value)
        name_value = str(excel_sheet[f'E{index}'].value)

        if date_value == 'None' and name_value == 'None':
            return index

        index += 1


def add_row_data(workbook, worksheet, file_path, file_data, last_row, user_name, task_name) -> None:
    for index, row in enumerate(file_data, start=0):
        number: int = last_row + index

        worksheet[f'C{number}'] = str('')
        worksheet[f'D{number}'] = str(strftime('%m/%d/%Y'))
        worksheet[f'E{number}'] = str(user_name)
        worksheet[f'F{number}'] = str(row[0])
        worksheet[f'G{number}'] = str(row[1])
        worksheet[f'H{number}'] = str(row[2])
        worksheet[f'I{number}'] = str(row[3])
        worksheet[f'J{number}'] = str(row[4])
        worksheet[f'K{number}'] = str(row[5])
        worksheet[f'L{number}'] = str(row[6])
        worksheet[f'M{number}'] = str(row[7])
        worksheet[f'N{number}'] = str(row[8])
        worksheet[f'O{number}'] = str(row[9])
        worksheet[f'P{number}'] = task_name

        cell_00 = worksheet[f'C{number}']
        cell_00.alignment = Alignment(horizontal='center')

        cell_01 = worksheet[f'D{number}']
        cell_01.alignment = Alignment(horizontal='center')

        cell_02 = worksheet[f'E{number}']
        cell_02.alignment = Alignment(horizontal='center')

        cell_03 = worksheet[f'F{number}']
        cell_03.alignment = Alignment(horizontal='center')

        cell_04 = worksheet[f'G{number}']
        cell_04.alignment = Alignment(horizontal='center')

        cell_05 = worksheet[f'H{number}']
        cell_05.alignment = Alignment(horizontal='center')

        cell_06 = worksheet[f'I{number}']
        cell_06.alignment = Alignment(horizontal='center')

        cell_07 = worksheet[f'J{number}']
        cell_07.alignment = Alignment(horizontal='center')

        cell_08 = worksheet[f'K{number}']
        cell_08.alignment = Alignment(horizontal='center')

        cell_09 = worksheet[f'L{number}']
        cell_09.alignment = Alignment(horizontal='center')

        cell_10 = worksheet[f'M{number}']
        cell_10.alignment = Alignment(horizontal='center')

        cell_11 = worksheet[f'N{number}']
        cell_11.alignment = Alignment(horizontal='center')

        cell_12 = worksheet[f'O{number}']
        cell_12.alignment = Alignment(horizontal='center')

        cell_13 = worksheet[f'P{number}']
        cell_13.alignment = Alignment(horizontal='center')

    try:
        workbook.save(file_path)
    except PermissionError:
        print(f'\n\tClose {Fore.RED}_inventory_transactions.xlsx{Style.RESET_ALL} to send.')


def get_drive_transaction_logs(file_data: list, user_name: str, task_name: str):
    """
    Get transaction logs found in the Z: Drive or 172.30.1.100 network.
    :return:
    """
    root_path: str = r'Z:\Kirkland_Lab\PipeCleaner\transaction_logs\_inventory_transactions.xlsx'
    alt_path: str = r'\\172.30.1.100\pxe\Kirkland_Lab\PipeCleaner\transaction_logs\_inventory_transactions.xlsx'

    try:
        add_transaction_log(file_data, root_path, user_name, task_name)

        second_place: str = r'Z:\Kirkland_Lab\Users\joe_ton\backup\_inventory_transactions.xlsx'
        add_transaction_log(file_data, second_place, user_name, task_name)

    except FileNotFoundError:
        add_transaction_log(file_data, alt_path, user_name, task_name)

        second_place: str = r'\\172.30.1.100\pxe\Kirkland_Lab\Users\joe_ton\backup\_inventory_transactions.xlsx'
        add_transaction_log(file_data, second_place, user_name, task_name)


def add_transaction_log(file_data, root_path, user_name, task_name):
    workbook = load_workbook(root_path)
    worksheet = workbook['All Inventory Transaction']
    last_row: int = get_last_entry_row(worksheet)
    add_row_data(workbook, worksheet, root_path, file_data, last_row, user_name, task_name)


def add_email_body(inventory_file: list, file_path: str, file_name: str) -> str:
    """
    Add the inventory body of the email.
    """
    body_message: str = ''

    for index, request in enumerate(inventory_file, start=1):

        body_message += f'\nInventory Request #{index}'
        body_message += f'\n\tFrom:         {request[0]}'
        body_message += f'\n\tTo:              {request[1]}'
        body_message += f'\n\tType:          {request[2]}'
        body_message += f'\n\tPart #:        {request[3]}'
        body_message += f'\n\tQTY:            {request[4]}'
        body_message += f'\n\tSupplier:   {request[5]}'
        body_message += f'\n\tPipe #:       {request[6]}'
        body_message += f'\n\tCage #:       {request[7]}'
        body_message += f'\n\tPO #:          {request[8]}'
        body_message += f'\n\tNotes #:     {request[9]}\n'

    body_message += f'\n\nFile Name: {file_name}'
    body_message += f'\nFile Path: {file_path}'

    return body_message


def email_inventory_request(file_path: str, file_name: str, inventory_file: list, task_name: str) -> None:
    """
    Sends email to people responsible dealing with
    """
    print(f'\tWriting and sending email to Inventory_Kirkland@veritasdcservices.com ....')
    real_location: str = 'Inventory_Kirkland@veritasdcservices.com'
    # person_location: str = 'joe.ton@VeritasDCservices.com'

    outlook = client.Dispatch('Outlook.Application')
    message = outlook.CreateItem(0)
    message.To = real_location
    message.Subject = task_name

    message.Body = add_email_body(inventory_file, file_path, file_name)
    message.Send()


def get_task_name(inventory_file_name: str) -> str:
    """
    Get task name / task id from MFST Planner.  This is in relationship with the inventory request.
    """
    worksheet = load_workbook(filename=inventory_file_name)['Inventory Template']
    return worksheet['G5'].value


def check_task_name(task_name: str) -> bool:
    """
    Checks if task name is entered in from the user.
    """
    if not task_name:
        return False

    elif 'Copy Task Title Here' in task_name:
        return False

    elif 'Copy' in task_name and\
            'Task' in task_name and \
            'Title' in task_name and \
            'Here' in task_name:
        return False

    else:
        return True


def validate_task_name(inventory_file_name: str) -> str:
    """
    Ensures task name is placed within the inventory_transaction.xlsx from the user.
    """
    task_name: str = get_task_name(inventory_file_name)
    is_task_name: bool = check_task_name(task_name)

    if not is_task_name:
        print(f'\tWARNING!!!')
        print(f'\tWARNING!!!')
        print(f'\tWARNING!!!\n')
        print(f'\tTask name from MFST Planner is not filled out in inventory_transaction.xlsx')
        print(f'\tPlease enter the assigned task title at the top of the inventory_transaction.xlsx')
        print(f'\tPress enter to exit.')
        input()
        sys.exit()

    elif is_task_name:
        return task_name


def main_method(default_user_name: str, inventory_authorized: tuple) -> None:
    """
    Send data to specific locations in Z: Drive to store information.
    :return:
    """
    user_name: str = get_clean_user_name(default_user_name)

    if is_authorized_user(default_user_name, inventory_authorized):

        print(f'\tAuthorized User: {user_name}')

        root_file_path: str = r'Z:\Kirkland_Lab\PipeCleaner\transaction_logs'
        inventory_file_name: str = 'inventory_transaction.xlsx'

        file_number: str = get_file_number(root_file_path)
        task_name: str = validate_task_name(inventory_file_name)

        inventory_file: list = read_inventory_file(inventory_file_name)
        file_data: list = print_file_data(inventory_file)
        file_path: str = get_file_path(user_name, root_file_path, file_number)
        user_response: str = get_user_send_response().upper()

        if 'YES' in user_response or user_response == 'Y':
            shutil.copyfile('inventory_transaction.xlsx', file_path)
            get_drive_transaction_logs(file_data, user_name, task_name)

            email_inventory_request(file_path, f'{file_number}_{user_name}', inventory_file, task_name)

            print(f'\n\t{Fore.GREEN}Transaction log sent to Z: Drive{Style.RESET_ALL}')
            print(f'\n\tFile Path: {root_file_path}')
            print(f'\tFile Name: {file_number}_{user_name}.xlsx')

            input(f'\n\tPress enter to exit Pipe Cleaner...')
            sys.exit()

        elif 'NO' in user_response or user_response == 'N':
            print(f'\tDid not sent transaction log to Z: Drive')

            input(f'\n\tPress enter to exit Pipe Cleaner...')
            sys.exit()

    else:
        print(f'\tAuthorized User [FAILED]: {user_name}')
        print(f'\n\tPlease notify developer or inventory admin to allow access.')
        print(F'\tPress enter to continue:')
        input()
