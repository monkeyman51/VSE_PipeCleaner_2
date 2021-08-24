"""
Inventory Total - Run when needed to consolidate data from Console Server blades to physical count.
"""
import datetime
import os
from csv import reader
from datetime import datetime, date
from json import loads, dumps

import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side, BORDER_THIN


def get_physical_count(csv_file_path: str) -> dict:
    """
    Iterate through each row of physical count provided in the CSV file from Inventory Team.
    :return: Provide part numbers associated with physical location and count.
    """
    physical_count: dict = {}

    with open(csv_file_path, "r") as file:
        file_data = reader(file, delimiter=",", quotechar='"')

        for index, row in enumerate(file_data, start=0):

            if index != 0:
                physical_count: dict = get_physical_data(physical_count, row)

    return physical_count


def get_console_server_all_hosts() -> list:
    """
    Get all hosts found in the All Hosts section of ZT Console Server.
    """
    generate_data: dict = {
        'action': 'get_host_status'
    }
    response = requests.post(url='http://172.30.1.100/console/console_js.php', data=dumps(generate_data))
    return loads(response.text)


def get_hosts_data() -> list:
    """
    Gather all host data from Console Server's All Host page. Purpose is to gather individual host ID and IP for later
    async fetching.  Duplicates... 1153(Unique) -> 1171(Count) = 18 Delta
    """
    all_hosts: list = get_console_server_all_hosts()

    serial_numbers: list = []

    hosts_data: list = []
    for host in all_hosts:

        host_data: dict = {}
        host_ip: str = host.get("host_ip", "None").upper()
        host_id: str = host.get("id", "None").upper()
        serial: str = host.get("serial", "None").upper()

        if serial not in serial_numbers:
            serial_numbers.append(serial)

            host_data["host_ip"]: str = host_ip
            host_data["host_id"]: str = host_id
            host_data["serial"]: str = serial

            hosts_data.append(host_data)

    return hosts_data


def get_physical_data(physical_count: dict, row) -> dict:
    """
    Iterate through each row in CSV file.  Provide data and some level of data validation to ensure clean data.
    :param physical_count: data containing all physical count
    :param row: row data
    :return: clean data for later merging into Console Server
    """
    location: str = row[0]
    part_number: str = row[1]
    count: str = row[2]

    if part_number == "":
        pass

    elif part_number == "EMPTY":
        pass

    elif part_number == "DECOMMISIONED PARTS":
        pass

    elif part_number in physical_count:
        physical_count[part_number]["location"]: str = location
        physical_count[part_number]["count"] += int(count)

    elif count == "":
        physical_count[part_number]: dict = {}
        physical_count[part_number]["location"]: str = location
        physical_count[part_number]["count"]: int = 0

    else:
        physical_count[part_number]: dict = {}
        physical_count[part_number]["location"]: str = location
        physical_count[part_number]["count"] = int(count)

    return physical_count


def add_excel_data(physical_count: dict, machines_data: dict, total_inventory: dict) -> str:
    """
    Add physical count to the excel output.
    :param physical_count: data that Inventory Team provided
    :param machines_data: Console Server data
    :param total_inventory: physical count + Console Server data
    """
    workbook = load_workbook(f'settings/total_inventory_template.xlsx')
    worksheet = workbook['Sheet1']

    total_parts: int = 0
    total_cage: int = 0
    total_rack: int = 0

    # for index, part_number in enumerate(physical_count, start=9):
    #     count: int = physical_count[part_number]["count"]
    #     location: int = physical_count[part_number]["location"]
    #
    #     total_parts += count
    #     total_cage += count
    #
    #     worksheet[f"A{index}"].value = part_number
    #     worksheet[f"D{index}"].value = count
    #     worksheet[f"E{index}"].value = count
    #
    #     if location == "0":
    #         worksheet[f"F{index}"].value = "None"
    #     else:
    #         worksheet[f"F{index}"].value = location
    #
    #     worksheet[f"D{index}"].alignment = Alignment(horizontal='center')
    #     worksheet[f"E{index}"].alignment = Alignment(horizontal='center')
    #     worksheet[f"F{index}"].alignment = Alignment(horizontal='center')
    #
    #     set_left_border(index, worksheet)

    for index, part_number in enumerate(total_inventory, start=9):
        count: int = total_inventory.get(part_number, {}).get("count", 0)
        location: str = total_inventory.get(part_number, {}).get("location", "None")
        part_type: str = total_inventory.get(part_number, {}).get("type", "None")
        supplier: str = total_inventory.get(part_number, {}).get("supplier", "None")

        worksheet[f"A{index}"].value = part_number
        worksheet[f"B{index}"].value = part_type
        worksheet[f"C{index}"].value = supplier
        worksheet[f"D{index}"].value = count

        total_parts += count
        total_cage: int = add_cage_data(index, part_number, physical_count, total_cage, worksheet)
        total_rack: int = add_rack_data(index, machines_data, part_number, total_rack, worksheet)

        if location == "0":
            worksheet[f"F{index}"].value = "None"
        else:
            worksheet[f"F{index}"].value = location

        worksheet[f"D{index}"].alignment = Alignment(horizontal='center')
        worksheet[f"E{index}"].alignment = Alignment(horizontal='center')
        worksheet[f"F{index}"].alignment = Alignment(horizontal='center')

        set_left_border(index, worksheet)

    worksheet["E7"].value = f"Cage - {total_cage}"
    worksheet["G7"].value = f"Rack - {total_rack}"

    add_date(worksheet)
    add_inventory_total(total_parts, worksheet)
    add_time(worksheet)

    total_inventory: str = "total_inventory.xlsx"

    workbook.save(total_inventory)

    return total_inventory


def add_rack_data(index, machines_data, part_number, total_rack, worksheet):
    if part_number in machines_data:
        rack_count: int = machines_data.get(part_number, {}).get("count", 0)
        rack_location: str = machines_data.get(part_number, {}).get("location", "None")

        worksheet[f"E{index}"].value = rack_count
        worksheet[f"F{index}"].value = rack_location

        total_rack += rack_count
    return total_rack


def add_cage_data(index, part_number, physical_count, total_cage, worksheet):
    if part_number in physical_count:
        cage_count: int = physical_count.get(part_number, {}).get("count", 0)
        cage_location: str = physical_count.get(part_number, {}).get("location", "None")

        worksheet[f"E{index}"].value = cage_count
        worksheet[f"F{index}"].value = cage_location

        total_cage += cage_count
    return total_cage


def set_left_border(index, worksheet):
    thin_border = Border(
        left=Side(border_style=BORDER_THIN, color='00000000'),
    )
    worksheet.cell(row=index, column=5).border = thin_border


def add_time(worksheet):
    """

    :param worksheet:
    :return:
    """
    current_time: str = datetime.now().strftime('%H:%M')
    worksheet["B5"].value = f"{current_time} PST"


def generate_json_data(host_id: str) -> None:
    """
    Generate latest Console Server host data.
    """
    generate_data = {
        'action': 'get_json_data',
        'host_id': f'{host_id}'
    }
    requests.post(url='http://172.30.1.100/console/console_js.php', json=generate_data)


def get_console_server_json(product_serial: str) -> dict:
    """
    Gets the Generated data using the product_serial string and creates JSON file.
    Returns JSON of Host within Console Server.

    :param product_serial: string from generate_json_data method
    :return: JSON data of Host
    """
    data = {
        'action': 'get_json_data',
        'host_id': f'{product_serial}'
    }
    response = requests.post(url=f'http://172.30.1.100/results/{product_serial}.json', json=data)

    return loads(response.text)


def generate_console_server_json(host_id: str) -> str:
    """
    Generates the JSON data from the Host Details page using the Host ID.
    Returns product-serial if JSON data is generated properly.

    :param host_id: found on Host Details page in URL ?host_id=<some_id>get_console_server_json
    :return: returns product-serial string for getting JSON data
    """
    generate_data = {
        'action': 'get_json_data',
        'host_id': f'{host_id}'
    }
    host_name_data = {
        'action': 'get_host_name_data',
        'host_id': f'{host_id}'
    }

    requests.post(url='http://172.30.1.100/console/console_js.php', json=generate_data)
    host_response = requests.post(url='http://172.30.1.100/console/console_js.php', json=host_name_data)

    product = loads(str(host_response.text))['host_name_data']['product']
    serial = loads(str(host_response.text))['host_name_data']['serial']

    return f'{product}-{serial}'


def get_all_machines_data(hosts_data: list) -> dict:
    """
    Get all machine data from Console Server
    """
    console_server_inventory: dict = {}

    for index, host_data in enumerate(hosts_data, start=1):
        print(f'Rack Host - {index}')

        host_id: str = host_data["host_id"]

        generate_json_data(host_id)
        product_serial: str = generate_console_server_json(host_id)
        machine_data: dict = get_console_server_json(product_serial)

        machine_name: str = machine_data.get("machine_name", "None")

        if "-VM-" in machine_name:
            pass

        else:
            location: str = machine_data.get("location", "None")

            unique_nvmes: list = machine_data["nvme"]["nvmes"]
            for unique_nvme in unique_nvmes:
                part_number: str = unique_nvme.get("model")
                commodity_type: str = "NVMe"

                console_server_inventory: dict = add_console_server_data(commodity_type,
                                                                         console_server_inventory,
                                                                         part_number)

            unique_disks: list = machine_data["disk"]["disks"]
            for unique_disk in unique_disks:
                part_number: str = unique_disk.get("model", "None")
                commodity_type: str = "Disk"

                console_server_inventory: dict = add_console_server_data(commodity_type,
                                                                         console_server_inventory,
                                                                         part_number)

            unique_dimms: list = machine_data["dmi"]["dimms"]
            for unique_dimm in unique_dimms:
                part_number: str = unique_dimm.get("part", "None")
                commodity_type: str = "DIMM"

                console_server_inventory: dict = add_console_server_data(commodity_type,
                                                                         console_server_inventory,
                                                                         part_number)

    return console_server_inventory


def add_console_server_data(commodity_type: str, console_server_inventory: dict, part_number: str) -> dict:
    """

    :param commodity_type:
    :param console_server_inventory:
    :param part_number:
    :return:
    """
    if part_number not in console_server_inventory:
        console_server_inventory[part_number]: dict = {}
        console_server_inventory[part_number]["type"]: str = commodity_type
        console_server_inventory[part_number]["count"]: int = 1

    elif part_number in console_server_inventory:
        console_server_inventory[part_number]["type"]: str = commodity_type
        console_server_inventory[part_number]["count"] += 1

    return console_server_inventory


def add_inventory_total(total_parts: int, worksheet) -> None:
    """
    Add total
    :param total_parts:
    :param worksheet:
    :return:
    """
    worksheet["B2"].value = total_parts


def add_date(worksheet):
    year: str = str(date.today().year)
    month: str = str(date.today().month)
    day: str = str(date.today().day)
    worksheet["B4"].value = f"{month}-{day}-{year}"


def main_method() -> None:
    """

    """
    csv_file_path: str = "settings/physical_baseline.csv"

    physical_count: dict = get_physical_count(csv_file_path)
    hosts_data: list = get_hosts_data()
    machines_data: dict = get_all_machines_data(hosts_data)

    total_inventory: dict = {**physical_count, **machines_data}

    excel_data: str = add_excel_data(physical_count, machines_data, total_inventory)

    os.system(fr'start EXCEL.EXE {excel_data}')
