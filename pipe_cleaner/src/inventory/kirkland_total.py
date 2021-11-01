"""
Total Kirkland based on Rich's data plus inventory tool database.
"""
import collections
import os
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Side, Border
from json import dumps

from pipe_cleaner.src.log_database import access_database_document


def audit_transaction_date() -> list:
    """
    Lists of transaction dates to skip when iterating through all transactions.
    :return: dates to avoid.
    """
    return ["06/23/2021", "07/07/2021", "07/08/2021", "07/09/2021", "07/12/2021", "07/13/2021", "07/14/2021",
            "07/15/2021", "07/16/2021", "07/19/2021", "07/20/2021", "07/21/2021", "07/22/2021", "07/23/2021",
            "07/26/2021", "07/27/2021", "07/28/2021", "07/29/2021", "07/30/2021", "08/02/2021", "08/03/2021",
            "08/04/2021", "08/05/2021", "08/06/2021", "08/09/2021", "08/10/2021", "08/11/2021", "8/12/2021"]


def is_date_after(current_date: str) -> bool:
    """
    Iterates through dates that should not be included in transaction logs.
    :param current_date:
    :return: True / False
    """
    audit_dates: list = audit_transaction_date()

    for audit_date in audit_dates:
        if current_date in audit_date:
            return False
    else:
        return True


def setup_transaction_data(inbound_name: str, outbound_name: str) -> dict:
    """
    Creates consistent naming convention for inbound and outbound names.
    :return: High-level structure of transactions of IN / OUT
    """
    return {inbound_name: {}, outbound_name: {}}


def rid_of_duplicate_serials(scanned_serials: list) -> list:
    """
    Get rid of duplicate serials for scanned serials.
    :param scanned_serials:
    :return:
    """
    return list(set(scanned_serials))


def clean_part_number_from_serial(serial_number: str, part_number: str) -> str:
    """
    Get rid of part number in serial number.
    :param serial_number:
    :param part_number:
    :return:
    """
    return serial_number. \
        replace(f"_{part_number}", ""). \
        replace(f" {part_number}", ""). \
        replace(f"-{part_number}", ""). \
        replace(f".{part_number}", ""). \
        replace(f"+{part_number}", ""). \
        replace(f"{part_number}_", ""). \
        replace(f"{part_number} ", ""). \
        replace(f"{part_number}-", ""). \
        replace(f"{part_number}.", ""). \
        replace(f"{part_number}+", ""). \
        replace(f"{part_number}", ""). \
        upper().replace(" ", "")


def clean_scanned_serials(scanned_serials: list, part_number: str) -> list:
    """
    Iterate through serials scanned.
    :param scanned_serials:
    :param part_number:
    :return:
    """
    clean_serials: list = rid_of_duplicate_serials(scanned_serials)

    pure_serials: list = []
    for serial_number in clean_serials:
        clean_serial: str = clean_part_number_from_serial(serial_number, part_number)
        pure_serials.append(clean_serial)

    return pure_serials


def clean_receipt_shipment(transaction_data: dict) -> dict:
    """
    Clean transaction database looking for duplicates and part number in serial number.
    :param transaction_data:
    :return:
    """
    transaction_data.pop("scanned_02", None)

    transaction_receipt: dict = transaction_data["receipt"]
    transaction_shipment: dict = transaction_data["shipment"]

    return {"receipts": log_transactions(transaction_receipt, "receipt"),
            "shipments": log_transactions(transaction_shipment, "shipment")}


def log_transactions(transaction_receipt, transaction_type):
    """

    :param transaction_receipt:
    :param transaction_type:
    :return:
    """
    clean_transactions: list = []
    for part_number in transaction_receipt:

        receipt_serials: list = transaction_receipt[part_number]["scanned"]
        previous_location: list = transaction_receipt[part_number]["previous_location"]
        current_location: list = transaction_receipt[part_number]["current_location"]

        for receipt_serial in receipt_serials:
            current_entry: dict = {"serial_number": clean_part_number_from_serial(receipt_serial, part_number),
                                   "previous_location": previous_location,
                                   "current_location": current_location,
                                   "part_number": part_number,
                                   "transaction_type": transaction_type}

            clean_transactions.append(current_entry)

    return clean_transactions


def get_transactions_from_database(part_numbers_library: dict) -> dict:
    """
    From inventory database
    :return:
    """
    transaction_data: dict = setup_transaction_data("receipt", "shipment")

    for current_entry in access_inventory_database():
        correct_transaction: dict = get_correct_transaction(current_entry, transaction_data)
        transaction_data.update(correct_transaction)

    return clean_receipt_shipment(transaction_data)


def access_inventory_database() -> list:
    """
    Access and return all transaction entries from Inventory.
    :return: List of transaction entries.
    """
    document = access_database_document('transactions', '021')
    return document.find({})


def access_serial_numbers_database() -> list:
    """
    Access and return serial numbers from Inventory database for testing and mocking.
    :return: List of transaction entries.
    """
    document = access_database_document('serial_numbers', 'base_line')
    return document.find({})


def get_correct_transaction(current_entry: dict, transaction_data: dict) -> dict:
    """
    Get audited transaction logs.
    :param transaction_data: all transaction data
    :param current_entry: transaction log
    :return: combines current transaction with all transactions
    """
    current_date: str = current_entry["time"]["date_logged"]

    if is_date_after(current_date):
        previous_location: str = current_entry["location"]["previous"]
        current_location: str = current_entry["location"]["current"]

        if "Receipt" in previous_location:
            return add_transaction_data("receipt", current_entry, transaction_data)

        elif "Shipment" in current_location or "Customer" in current_location:
            return add_transaction_data("shipment", current_entry, transaction_data)

        else:
            return transaction_data

    else:
        return transaction_data


def is_duplicates(scanned_serials: list) -> bool:
    """
    Checks if duplicates in scanned serial numbers.
    :param scanned_serials:
    :return:
    """
    total_count: int = len(scanned_serials)
    non_duplicates: int = len(set(scanned_serials))

    if total_count != non_duplicates:
        return True

    elif total_count == non_duplicates:
        return False


def get_duplicates(check_duplicates: bool, scanned_serials: list) -> list:
    """
    If duplicates exist, return duplicate serial numbers.
    :param check_duplicates:
    :param scanned_serials:
    :return:
    """
    if check_duplicates is True:
        return [item for item, count in collections.Counter(scanned_serials).items() if count > 1]
    else:
        return []


def add_transaction_data(transaction_type: str, current_entry: dict, transaction_data: dict) -> dict:
    """
    Add to overall total based on transaction data.
    :param transaction_type: receipt or shipment
    :param current_entry: total data
    :param transaction_data: total data
    :return: all transaction data
    """
    scanned_serials: list = current_entry["scanned"]
    part_number: str = current_entry["part_number"]
    previous_location: str = current_entry["location"]["previous"]
    current_location: str = current_entry["location"]["current"]

    current_transaction: dict = transaction_data[transaction_type]

    if part_number in current_transaction:
        check_duplicates: bool = is_duplicates(scanned_serials)
        part_data: dict = current_transaction[part_number]
        duplicates: list = get_duplicates(check_duplicates, scanned_serials)

        if isinstance(scanned_serials, list):
            part_data["scanned"].append(scanned_serials[0])
        else:
            part_data["scanned"].append(scanned_serials)

        part_data["duplicates"].append(duplicates)
        part_data["is_duplicate"]: bool = check_duplicates
        part_data["previous_location"]: bool = previous_location
        part_data["current_location"]: bool = current_location
        part_data["count"] += len(scanned_serials)

    elif part_number not in current_transaction:
        check_duplicates: bool = is_duplicates(scanned_serials)
        current_transaction[part_number]: dict = {}
        part_data: dict = current_transaction[part_number]

        part_data["scanned"]: list = scanned_serials
        part_data["duplicates"]: list = get_duplicates(check_duplicates, scanned_serials)
        part_data["is_duplicate"]: bool = check_duplicates
        part_data["previous_location"]: bool = previous_location
        part_data["current_location"]: bool = current_location
        part_data["count"]: int = len(scanned_serials)

    transaction_data["scanned_02"]: list = []
    for raw_serial in scanned_serials:
        alternative_part: str = f'{part_number}_'
        if alternative_part in raw_serial:
            clean_version: str = raw_serial.replace(alternative_part, "")
            transaction_data["scanned_02"].append(clean_version)

    return transaction_data


def get_month_name(raw_month: str) -> str:
    """

    :param raw_month:
    :return:
    """
    if raw_month == "01":
        return "January"

    elif raw_month == "02":
        return "February"

    elif raw_month == "03":
        return "March"

    elif raw_month == "04":
        return "April"

    elif raw_month == "05":
        return "May"

    elif raw_month == "06":
        return "June"

    elif raw_month == "07":
        return "July"

    elif raw_month == "08":
        return "August"

    elif raw_month == "09":
        return "September"

    elif raw_month == "10":
        return "October"

    elif raw_month == "11":
        return "November"

    elif raw_month == "12":
        return "December"

    else:
        return "None"


def get_clean_day(raw_day: str) -> str:
    """

    :param raw_day:
    :return:
    """
    if raw_day[0] == "0" and len(raw_day) == 2:
        return raw_day[1]

    else:
        return raw_day


def get_current_date() -> str:
    """
    Get current date for logging later.
    :return:
    """
    raw_date: str = get_raw_date()
    raw_year: str = get_raw_year(raw_date)
    raw_month: str = get_raw_month(raw_date)
    raw_day = get_raw_day(raw_date)

    clean_month: str = get_month_name(raw_month)
    clean_day: str = get_clean_day(raw_day)

    return f"{clean_month} {clean_day}, {raw_year}"


def get_raw_day(raw_date: str) -> str:
    """
    Get raw day in terms of date
    :param raw_date:
    :return:
    """
    return raw_date[8:10]


def get_raw_month(raw_date: str) -> str:
    """
    Get month from raw date.
    :param raw_date: month, year, day
    :return: month
    """
    return raw_date[5:7]


def get_raw_year(raw_date: str) -> str:
    """
    Get current year for later string parsing. ex. 2021 / 2022
    :param raw_date: month, day, year
    :return: year
    """
    return raw_date[0:4]


def get_raw_date() -> str:
    """
    Get raw dates later later parsing and string manipulation.
    :return:
    """
    return datetime.now().strftime('%Y-%m-%d')


def get_current_time() -> str:
    """

    :return:
    """
    current_time: str = datetime.today().strftime("%I:%M %p")
    return current_time


def subtract_transaction_data(part_number: str, scanned_serials: list, transaction_data: dict) -> dict:
    """
    Minus to overall total based on transaction data.
    :param part_number: name of commodity
    :param scanned_serials: quantity being moved
    :param transaction_data: total data
    :return: all transaction data
    """
    shipment: dict = transaction_data["shipment"]
    check_duplicates: bool = is_duplicates(scanned_serials)

    if part_number in transaction_data:
        part_data: dict = shipment[part_number]
        duplicates: list = get_duplicates(check_duplicates, scanned_serials)

        part_data["scanned"].append(scanned_serials)
        part_data["duplicates"].append(duplicates)
        part_data["is_duplicate"]: bool = check_duplicates
        part_data["count"] += len(scanned_serials)

    elif part_number not in transaction_data:
        transaction_data[part_number]: dict = {}
        part_data: dict = shipment[part_number]

        part_data["scanned"]: list = scanned_serials
        part_data["duplicates"]: list = get_duplicates(check_duplicates, scanned_serials)
        part_data["is_duplicate"]: bool = check_duplicates
        part_data["count"]: int = len(scanned_serials)

    return transaction_data


def get_datetime_output() -> str:
    """
    Create default date and time format for excel report.
    :return: excel output date and time
    """
    current_date: str = get_current_date()
    current_time: str = get_current_time()
    return f"{current_date}  -  {current_time} PST"


def get_base_mock() -> dict:
    """

    :return:
    """
    return {"MTA18ASF2G72PZ-3G2R1": 100,
            "HMA84GR7CJR4N-VK": 5,
            "MTFDDAK960TDD": 8,
            "36ASF8G72PZ-3G2B2": 34,
            "MZ1LB960HAJQ-00AMV": 254,
            "HFS960GD0FEI-A430A": 65,
            "SSDPELKX960G8D": 22,
            "2LQ202-403": 3,
            "MTA36ASF8G72PZ-3G2B2": 54,
            "M393A8G40AB2-CWE": 20,
            "MTA36ASF4G72PZ-3G2R1": 10,
            "HFS3T8GD0FEI-A430B": 22
            }


def get_merge_data(transactions: dict, base_mock: dict) -> dict:
    """

    :param transactions:
    :param base_mock:
    :return:
    """
    receipts: dict = transactions["receipt"]
    shipments: dict = transactions["shipment"]

    for receipt_part_number in receipts:
        receipt_count = int(receipts[receipt_part_number])

        if receipt_part_number in base_mock:
            base_mock[receipt_part_number] += receipt_count

        elif receipt_part_number not in base_mock:
            base_mock[receipt_part_number] = receipt_count

    for shipment_part_number in shipments:
        shipment_count = int(shipments[shipment_part_number])

        if shipment_part_number in base_mock:
            current_count: int = base_mock[shipment_part_number]
            total_count: int = current_count - shipment_count
            base_mock[shipment_part_number] = total_count

        elif shipment_part_number not in base_mock:
            total_measure: int = 0 - int(shipment_count)
            base_mock[shipment_part_number] = total_measure

    return base_mock


def get_mock_serial_numbers() -> dict:
    """

    :return:
    """
    base_serials: dict = {}

    for entry in access_serial_numbers_database():
        part_number: str = entry["Commodity Part Number Generic"]
        serial_number: str = entry["Commodity Serial Number  "]
        supplier: str = entry["Part Supplier"]

        if part_number not in base_serials:
            base_serials[part_number]: dict = {}
            base_serials[part_number]["scanned"]: list = [serial_number]
            base_serials[part_number]["supplier"]: str = supplier

        elif part_number in base_serials:
            base_serials[part_number]["scanned"].append(serial_number)
            base_serials[part_number]["supplier"]: str = supplier

    return base_serials


def get_total(transactions: dict, base_serials: dict) -> dict:
    """

    :param transactions:
    :param base_serials:
    :return:
    """
    for receipt_part_number in transactions["receipt"]:
        part_data: dict = transactions["receipt"][receipt_part_number]
        scanned: dict = part_data["scanned"]

        if receipt_part_number not in base_serials:
            base_serials[receipt_part_number]: dict = {}
            base_serials[receipt_part_number]["scanned"]: list = scanned
            base_serials[receipt_part_number]["supplier"]: list = "None"

        elif receipt_part_number in base_serials:
            base_serials[receipt_part_number]["scanned"].append(scanned)

    for shipment_part_number in transactions["shipment"]:
        part_data: dict = transactions["shipment"][shipment_part_number]
        scanned: dict = part_data["scanned"]

        if shipment_part_number in base_serials:
            for shipment_serial in scanned:

                base_scanned: list = base_serials[shipment_part_number]["scanned"]
                if shipment_serial in base_scanned:
                    base_scanned.remove(shipment_serial)

    return base_serials


def get_total_count(total_data: dict) -> int:
    """
    Get total commodities
    :param total_data:
    :return:
    """
    total_count: int = 0

    for part_number in total_data:
        current_count: int = len(total_data[part_number]["scanned"])
        total_count += current_count

    return total_count


def setup_serial_number() -> dict:
    """
    Structure of serial number per transaction.
    :return: serial_number_id
    """
    return {"part_numbers": [],
            "from_locations": [],
            "to_locations": [],
            "scanners": [],
            "requesters": [],
            "times": [],
            "dates": [],
            "suppliers": [],
            "notes": [],
            "tool_versions": [],
            "in_house": True}


def pretty_print(current_dictionary: dict) -> None:
    """
    Temporary.  Prints out dictionary for readable.
    :param current_dictionary:
    :return:
    """
    print(dumps(current_dictionary, sort_keys=True, indent=4))
    input()


def fill_missing_supplier(part_number: str, supplier: str, part_numbers_library: dict) -> str:
    """
    Fill the missing supplier based on part number using the part numbers library
    :param part_number: current part number
    :param supplier: current supplier
    :param part_numbers_library: from part numbers
    :return: missing supplier if found
    """
    if check_none(supplier) is True:

        missing_supplier = part_numbers_library.get(part_number, "None")
        return missing_supplier
    else:
        return supplier


def get_serial_numbers_database(part_numbers_library: dict):
    """
    Get serial numbers for all records of Kirkland.
    :return:
    """
    potential_fake_serial: int = 0
    potential_fake_underscore: int = 0

    document = access_database_document("serial_numbers", 'version_02')
    serial_numbers_database: list = document.find({})

    database: dict = {}
    for entry in serial_numbers_database:
        serial_number: str = entry["_id"]
        in_house: str = entry["in_house"][-1]
        part_number: str = entry["part_numbers"][-1]
        location: str = entry["to_locations"][-1]
        supplier: str = entry["suppliers"][-1]

        potential_fake_serial += alert_potential_fake_serial(serial_number)
        potential_fake_underscore += alert_fake_underscore(serial_number)

        if serial_number not in database:
            database[serial_number]: dict = {}
            available_supplier: str = store_available_supplier(part_number, part_numbers_library, supplier)
            primary_part_number: str = store_primary_part_number(part_number, part_numbers_library)

            database[serial_number]["supplier"]: str = available_supplier
            database[serial_number]["part_number"]: str = primary_part_number
            database[serial_number]["in_house"]: str = in_house
            database[serial_number]["location"]: str = location

    print(f"potential_fake_serial: {potential_fake_serial}")
    print(f"potential_fake_underscore: {potential_fake_underscore}")

    return database


def alert_potential_fake_serial(serial_number: str) -> int:
    """
    Most instances, serial numbers should be a combination of just numbers and letters.  This is to alert of

    :param serial_number:
    :return:
    """
    for character in serial_number:
        if not character.isdigit() and not character.isalpha():
            return 1
    else:
        return 0


def alert_fake_underscore(serial_number: str) -> int:
    """
    Most instances, serial numbers should be a combination of just numbers and letters.  This is to alert of

    :param serial_number:
    :return:
    """
    for character in serial_number:
        if "_" in character:
            return 1
    else:
        return 0


def store_primary_part_number(part_number: str, part_numbers_library) -> str:
    """
    If missing part number from data set, search available part number's supplier in part number library
    :param part_number:
    :param part_numbers_library:
    :return:
    """
    part_number_alt: dict = part_numbers_library.get(part_number, None)

    if part_number_alt:
        return part_number_alt["primary"]

    else:
        return part_number


def store_available_supplier(part_number, part_numbers_library, supplier) -> str:
    """
    If missing supplier from data set, search available part number's supplier in part number library
    :param part_number:
    :param part_numbers_library:
    :param supplier:
    :return:
    """
    if supplier == "None":
        part_number_alt: dict = part_numbers_library.get(part_number, None)

        if part_number_alt:
            return part_number_alt["supplier"]

        else:
            return "None"
    else:
        return supplier


def clean_serial_number(serial_numbers: list) -> list:
    """
    Sometimes serial numbers will able as list type while it should be string type.
    :param serial_numbers:
    :return:
    """
    clean_serials: list = []
    for serial_number in serial_numbers:
        if isinstance(serial_number, list):
            serial_non_list: str = serial_number[0]
            clean_serials.append(serial_non_list)
        else:
            clean_serials.append(serial_number)

    return clean_serials


def setup_master_data() -> dict:
    """
    Structure master data for later excel output.
    :return:
    """
    return {"part_numbers": {},
            "cage_total": 0,
            "rack_total": 0,
            "quarantine_total": 0,
            "not_in_house": 0,
            "in_house": 0}


def get_clean_cage_location(master_part_numbers: dict, part_number: str, location_type: str):
    """
    Convert list into readable string inside excel cell.
    :param master_part_numbers:
    :param part_number:
    :param location_type:
    :return:
    """
    if location_type == "Cage":
        cage_location: str = master_part_numbers[part_number]["cage_location"]
        if cage_location == "":
            return "None"
        else:
            return ', '.join(cage_location)

    elif location_type == "Rack":
        rack_location: str = master_part_numbers[part_number]["rack_location"]
        if rack_location == "":
            return "None"
        else:
            return ', '.join(rack_location)

    elif location_type == "Quarantine":
        quarantine_location: str = master_part_numbers[part_number]["quarantine_location"]
        if quarantine_location == "":
            return "None"
        else:
            return ', '.join(quarantine_location)


def set_cell_alignment(index: int, worksheet, slots: dict):
    """
    Align cell in desired way.
    :param index:
    :param worksheet:
    :param slots:
    :return:
    """
    worksheet[f"{slots['part']}{index}"].alignment = Alignment(horizontal='left')
    worksheet[f"{slots['type']}{index}"].alignment = Alignment(horizontal='left')
    worksheet[f"{slots['supplier']}{index}"].alignment = Alignment(horizontal='left')
    worksheet[f"{slots['total_count']}{index}"].alignment = Alignment(horizontal='center')
    worksheet[f"{slots['cage_count']}{index}"].alignment = Alignment(horizontal='center')
    worksheet[f"{slots['rack_count']}{index}"].alignment = Alignment(horizontal='center')
    worksheet[f"{slots['quarantine_count']}{index}"].alignment = Alignment(horizontal='center')

    # For whatever reason, this cell is not taking left border manually.  Did this via code.
    worksheet[f"{slots['type']}{index}"].border = Border(left=Side(border_style='thin', color='A5A5A5'))
    worksheet[f"{slots['quarantine_count']}{index}"].border = Border(left=Side(border_style='thin', color='A5A5A5'))


def set_supplier_color(index: int, worksheet, supplier: str):
    """
    Font cell in desired way.
    """
    if supplier == "Unknown" or supplier == "None" or supplier == "":
        worksheet[f"C{index}"].font = Font(color="7B7B7B")


def create_workbook_for_report() -> load_workbook:
    """
    Create workbook for report.  Used for saving excel report later.
    :return:
    """
    base_names: dict = get_relevant_names()

    template_file_path: str = base_names["template_file_path"]
    return load_workbook(template_file_path)


def get_relevant_names() -> dict:
    """
    Get names for relevant application later. file name, sheet names, etc.

    file_name: report name for excel
    template_file: master inventory template path
    master_sheet_name:
    transactions_sheet_name:

    :return: dictionary of file names
    """
    raw_date: str = get_raw_date()
    clean_year: str = get_raw_year(raw_date)[2:4]
    raw_month: str = get_raw_month(raw_date)
    raw_day: str = get_raw_day(raw_date)

    return {"file_name": f"inventory_{raw_month}-{raw_day}-{clean_year}.xlsx",
            "template_file_path": "settings/master_inventory_template.xlsx",
            "master_sheet_name": "Master - Internal",
            "transactions_sheet_name": "Transactions"}


def save_and_launch(workbook) -> None:
    """
    Save output based off of the template.
    :param workbook: inventory template
    :return: None
    """
    base_names: dict = get_relevant_names()

    file_name: str = base_names["file_name"]
    workbook.save(file_name)
    launch_output_after_save()


def set_number_red(base_part_count: int, difference: int, index: int, part_count: int, worksheet_master):
    """
    If total count seems off based on numbers then will output in the excel as red font.
    """
    if difference != "":
        check_total: int = base_part_count + difference
        if part_count != check_total:
            worksheet_master[f"D{index}"].font = Font(color="FF0000")


def launch_output_after_save() -> None:
    """
    After creating excel report for inventory, launch the excel file for user to automatically see.
    :return: None
    """
    base_names: dict = get_relevant_names()

    file_name: str = base_names["file_name"]
    os.system(fr'start EXCEL.EXE {file_name}')


def create_worksheet_master(workbook: load_workbook) -> load_workbook:
    """
    Create worksheet for excel output.  Used for creating individual sheet.
    :param workbook: interaction with main workbook excel file
    :return: master_worksheet
    """
    base_names: dict = get_relevant_names()

    master_sheet_name: str = base_names["master_sheet_name"]
    return workbook[master_sheet_name]


def create_transactions_worksheet(workbook: load_workbook) -> load_workbook:
    """
    Create worksheet for later use and coding interface.
    :param workbook:
    :return:
    """
    base_names: dict = get_relevant_names()

    transactions_sheet_name: str = base_names["transactions_sheet_name"]
    return workbook[transactions_sheet_name]


def add_transaction_top_left(receipt_count, shipment_count, worksheet_transactions):
    """
    Add transaction to the top left corner.
    :param receipt_count:
    :param shipment_count:
    :param worksheet_transactions:
    :return:
    """
    worksheet_transactions["C3"].value = receipt_count
    worksheet_transactions["C4"].value = shipment_count
    worksheet_transactions["C5"].value = receipt_count - shipment_count


def align_transactions_cells(index, worksheet_transactions) -> None:
    """
    Align for transaction cells
    :param index:
    :param worksheet_transactions:
    :return:
    """
    worksheet_transactions[f"B{index}"].alignment = Alignment(horizontal='left')
    worksheet_transactions[f"C{index}"].alignment = Alignment(horizontal='left')
    worksheet_transactions[f"D{index}"].alignment = Alignment(horizontal='center')
    worksheet_transactions[f"E{index}"].alignment = Alignment(horizontal='center')


def add_transaction_shipment(shipment_count, transaction_entry):
    scanned_serials: list = transaction_entry["scanned"]
    transaction_type: str = transaction_entry["transaction_type"]

    if transaction_type.upper() == "SHIPMENT":
        shipment_count += len(scanned_serials)
    return shipment_count


def add_transaction_receipt(receipt_count, transaction_entry) -> int:
    transaction_type: str = transaction_entry["transaction_type"]
    scanned_serials: list = transaction_entry["scanned"]

    if transaction_type.upper() == "RECEIPT":
        receipt_count += len(scanned_serials)
    return receipt_count


def add_transaction_type(index, transaction_entry, worksheet_transactions) -> None:
    transaction_type: str = transaction_entry["transaction_type"]

    worksheet_transactions[f"E{index}"].value = transaction_type.title()


def add_transaction_total(index, scanned_serials, worksheet_transactions) -> None:
    """

    :param index:
    :param scanned_serials:
    :param worksheet_transactions:
    :return:
    """
    worksheet_transactions[f"D{index}"].value = len(scanned_serials)


def add_transaction_supplier(index, worksheet_transactions) -> None:
    worksheet_transactions[f"C{index}"].value = "Unknown"


def add_transaction_part_number(index, transaction_entry, worksheet_transactions) -> None:
    part_number: str = transaction_entry["part_number"]

    worksheet_transactions[f"B{index}"].value = str(part_number)


def get_part_numbers_library() -> dict:
    """
    Get part_numbers library to help fill in gaps later.
    :return:
    """
    collection = access_database_document("part_numbers", "library")
    documents: list = collection.find({})

    library: dict = {}
    for document in documents:
        serial_number: str = document["_id"]
        supplier: str = document["supplier"]
        alternatives: list = document["alternatives"]

        if serial_number not in library:
            library[serial_number]: dict = {}
            library[serial_number]["primary"]: str = serial_number
            library[serial_number]["supplier"]: str = supplier
            library[serial_number]["alternatives"]: list = alternatives

        for alternative in alternatives:
            if alternative not in library:
                clean_alternatives: list = get_clean_alternatives(alternative, alternatives, serial_number)

                library[serial_number]: dict = {}
                library[serial_number]["primary"]: str = serial_number
                library[serial_number]["supplier"]: str = supplier
                library[serial_number]["alternatives"]: list = clean_alternatives

    return library


def get_clean_alternatives(alternative, alternatives, serial_number) -> list:
    """

    :param alternative:
    :param alternatives:
    :param serial_number:
    :return:
    """
    clean_alternatives: list = [serial_number]
    for alt_name in alternatives:
        if alt_name != alternative:
            clean_alternatives.append(alt_name)

    return clean_alternatives


def combine_serial_with_transaction(serial_numbers_database: dict, transactions_database: dict) -> dict:
    """
    Combine serial numbers database with transaction database.
    :param serial_numbers_database: Contains the base data sent from Rich and Inventory Team
    :param transactions_database: Logged through Pipe Cleaner
    :return: Combined data
    """
    combined: dict = get_combined_structure()
    combined: dict = count_serial_database(combined, serial_numbers_database)
    combined: dict = count_transaction_receipt(combined, transactions_database)
    combined: dict = count_transaction_shipment(combined, transactions_database)

    return combined


def count_transaction_receipt(combined, transactions_database):
    receipts: dict = transactions_database["receipts"]

    for entry in receipts:
        part_number: str = entry["part_number"]
        serial_number: str = entry["serial_number"]

        combined: dict = add_part_number(combined, part_number)

        count_serials: list = combined["parts"][part_number]["count_serials"]

        if serial_number not in count_serials:
            combined["stats"]["in_house"] += 1
            combined["stats"]["cage"] += 1

            combined["parts"][part_number]["count"] += 1
            combined["parts"][part_number]["cage"] += 1

            combined["parts"][part_number]["count_serials"].append(serial_number)
            combined["parts"][part_number]["cage_serials"].append(serial_number)

        elif serial_number in count_serials:
            combined["stats"]["already_counts"] += 1
            combined["stats"]["already_serials"].append({"serial_number": serial_number,
                                                         "part_number": part_number})

            combined["parts"][part_number]["already_serials"].append(serial_number)

        combined["stats"]["shipped_in"] += 1

    return combined


def count_transaction_shipment(combined, transactions_database):
    shipments: dict = transactions_database["shipments"]

    for entry in shipments:
        part_number: str = entry["part_number"]
        serial_number: str = entry["serial_number"]
        previous_location: str = entry["previous_location"]

        combined: dict = add_part_number(combined, part_number)
        count_serials: list = combined["parts"][part_number]["count_serials"]

        if serial_number in count_serials:
            combined["stats"]["in_house"] -= 1
            combined["stats"]["cage"] -= 1

            combined["parts"][part_number]["count"] -= 1

            combined: dict = remove_combined_location(combined, part_number, previous_location, serial_number)

        elif serial_number not in count_serials:
            combined["stats"]["rejected_counts"] += 1
            # combined["stats"]["rejected_serials"].append((part_number, serial_number))

        combined["stats"]["shipped_out"] += 1

    return combined


def remove_combined_location(combined, part_number, previous_location, serial_number):
    """

    :param combined:
    :param part_number:
    :param previous_location:
    :param serial_number:
    :return:
    """
    if previous_location == "Cage":
        combined["stats"]["cage"] -= 1

        combined["parts"][part_number]["cage"] -= 1
        # try:
        #     combined["parts"][part_number]["cage_serials"].remove(serial_number)
        # except ValueError:
        #     pass

    elif previous_location == "Rack":
        combined["stats"]["rack"] -= 1

        combined["parts"][part_number]["rack"] -= 1
        # try:
        #     combined["parts"][part_number]["rack_serials"].remove(serial_number)
        # except ValueError:
        #     pass

    elif previous_location == "Quarantine":
        combined["stats"]["quarantine"] -= 1

        combined["parts"][part_number]["quarantine"] -= 1

        # try:
        #     combined["parts"][part_number]["quarantine_serials"].remove(serial_number)
        # except ValueError:
        #     pass

    return combined


def clean_supplier_name(supplier: str) -> str:
    """
    Consistent supplier name.
    :param supplier:
    :return:
    """
    western_digital: str = "WD / HGST / Aspen"

    if check_none(supplier) is True:
        return "None"

    elif supplier == "Aspen / Western Digital":
        return western_digital

    elif supplier == "Western Digital":
        return western_digital

    elif supplier == "HGST / Western Digital":
        return western_digital

    elif supplier == "Toshiba":
        return "Toshiba / Kioxia"

    elif supplier == "wiwynn":
        return "Wiwynn"

    else:
        return supplier


def check_none(string_field: str) -> bool:
    """
    Check for none and invalid
    :param string_field:
    :return: True or False
    """
    if string_field == "Unknown" or string_field == "None" or \
            string_field is None or string_field == "NONE" or \
            string_field == "":
        return True
    else:
        return False


def add_unique_supplier(combined: dict, supplier: str) -> None:
    """
    Add to supplier total for later count in the excel output report.
    :param combined: SN and Transaction DB
    :param supplier: current supplier
    :return: plus one if valid unique supplier
    """
    if check_none(supplier):
        pass

    elif supplier not in combined["stats"]["unique_suppliers"]:
        clean_supplier: str = clean_supplier_name(supplier)
        combined["stats"]["unique_suppliers"].append(clean_supplier)


def count_serial_database(combined: dict, serial_numbers_database: dict) -> dict:
    """
    Keep track of
    :param combined: 
    :param serial_numbers_database: 
    :return: 
    """
    for serial_number in serial_numbers_database:
        part_number: str = serial_numbers_database[serial_number]["part_number"]
        in_house: str = serial_numbers_database[serial_number]["in_house"]
        location: str = serial_numbers_database[serial_number]["location"]
        supplier: str = serial_numbers_database[serial_number]["supplier"]
        clean_supplier: str = clean_supplier_name(supplier)

        combined: dict = add_part_number(combined, part_number)
        combined: dict = account_supplier_data(combined, part_number, clean_supplier)

        if location != "Out":
            combined: dict = add_cage_count(combined, part_number, location)
            combined: dict = add_rack_count(combined, part_number, location)
            combined: dict = add_quarantine_count(combined, part_number, location)

            combined["stats"]["base_serials"] += 1
            combined["stats"]["in_house"] += count_in_house(in_house)

            combined["parts"][part_number]["count"] += 1

            combined["parts"][part_number]["count_serials"].append(serial_number)
            combined: dict = append_cage_serial(combined, part_number, serial_number, location)
            combined: dict = append_rack_serial(combined, part_number, serial_number, location)
            combined: dict = append_quarantine_serial(combined, part_number, serial_number, location)

    return combined


def add_quarantine_count(combined: dict, part_number: str, location: str) -> dict:
    """
    Account for quarantine count in both stats and parts level for later Excel output for Inventory Total.
    """
    quarantine_count: int = count_location_quarantine(location)

    combined["stats"]["quarantine"] += quarantine_count
    combined["parts"][part_number]["quarantine"] += quarantine_count

    return combined


def add_rack_count(combined: dict, part_number: str, location: str) -> dict:
    """
    Account for rack count in both stats and parts level for later Excel output for Inventory Total.
    """
    rack_count: int = count_location_rack(location)

    combined["stats"]["rack"] += rack_count
    combined["parts"][part_number]["rack"] += rack_count

    return combined


def add_cage_count(combined: dict, part_number: str, location: str) -> dict:
    """
    Account for cage count in both stats level and parts level for later Excel output for Inventory Total.
    """
    cage_count: int = count_location_cage(location)

    combined["stats"]["cage"] += cage_count
    combined["parts"][part_number]["cage"] += cage_count

    return combined


def account_supplier_data(combined: dict, part_number: str, supplier: str) -> dict:
    """
    Add to combined data for supplier based information.

    :param combined: SN and Transaction DB
    :param part_number: current part number
    :param supplier: current supplier
    :return:
    """
    combined: dict = add_supplier_stat(combined, supplier)

    add_unique_supplier(combined, supplier)
    combined["parts"][part_number]["supplier"] = supplier
    combined["stats"]["suppliers"][supplier] += 1

    return combined


def get_combined_structure():
    combined: dict = {"stats": {}, "parts": {}}
    combined["stats"]["in_house"]: int = 0
    combined["stats"]["cage"]: int = 0
    combined["stats"]["rack"]: int = 0
    combined["stats"]["quarantine"]: int = 0
    combined["stats"]["base_serials"]: int = 0
    combined["stats"]["rejected_counts"]: int = 0
    combined["stats"]["already_counts"]: int = 0
    combined["stats"]["shipped_out"]: int = 0
    combined["stats"]["shipped_in"]: int = 0
    combined["stats"]["unique_suppliers"]: list = []

    combined["stats"]["rejected_serials"]: list = []
    combined["stats"]["already_serials"]: list = []
    combined["stats"]["suppliers"]: dict = {}

    return combined


def append_cage_serial(combined, part_number, serial_number, location) -> dict:
    if location == "Cage":
        combined["parts"][part_number]["cage_serials"].append(serial_number)
    return combined


def append_rack_serial(combined, part_number, serial_number, location) -> dict:
    if location == "Rack":
        combined["parts"][part_number]["rack_serials"].append(serial_number)
    return combined


def append_quarantine_serial(combined, part_number, serial_number, location) -> dict:
    if location == "Quarantine":
        combined["parts"][part_number]["quarantine_serials"].append(serial_number)
    return combined


def add_supplier_stat(combined, supplier):
    if supplier not in combined["stats"]["suppliers"]:
        combined["stats"]["suppliers"][supplier]: int = 0
    return combined


def add_part_number(combined, part_number):
    if part_number not in combined["parts"]:
        combined["parts"][part_number]: dict = {}
        combined["parts"][part_number]["count"]: int = 0
        combined["parts"][part_number]["cage"]: int = 0
        combined["parts"][part_number]["rack"]: int = 0
        combined["parts"][part_number]["quarantine"]: int = 0

        combined["parts"][part_number]["count_serials"]: list = []
        combined["parts"][part_number]["cage_serials"]: list = []
        combined["parts"][part_number]["rack_serials"]: list = []
        combined["parts"][part_number]["quarantine_serials"]: list = []
        combined["parts"][part_number]["rejected_serials"]: list = []
        combined["parts"][part_number]["already_serials"]: list = []

    return combined


def count_in_house(in_house: str) -> int:
    """

    :param in_house:
    :return:
    """
    if in_house == "True":
        return 1
    else:
        return 0


def count_location_quarantine(location: str) -> int:
    """

    :param location:
    :return:
    """
    return add_correct_location(location, "Quarantine")


def count_location_cage(location: str) -> int:
    """

    """
    return add_correct_location(location, "Cage")


def add_correct_location(location: str, check_location: str) -> int:
    """

    """
    possible_locations: list = get_possible_locations(location)

    if location == check_location:
        return 1

    elif location in possible_locations:
        return 0

    else:
        return 0


def get_possible_locations(location: str) -> list:
    """

    """
    possible_locations: list = ["Cage", "Quarantine", "Rack"]
    possible_locations.remove(location)

    return possible_locations


def count_location_rack(location: str) -> int:
    """
    Account for location based on rack.
    :param location: current location
    :return: count 1 if rack
    """
    if location == "Rack":
        return 1
    else:
        return 0


def get_unique_parts(combined: dict) -> list:
    """
    Get unique part numbers based from combined data gathered from SN and transaction database.
    :param combined: SN and transaction database
    :return: list of unique part numbers
    """
    return list(combined["parts"].keys())


def get_master_slots() -> dict:
    """
    Get master inventory columns for slotting the correct location.
    :return:
    """
    return {"start": 10,
            "part": "B",
            "type": "C",
            "supplier": "D",
            "total_count": "E",
            "cage_count": "F",
            "rack_count": "G",
            "quarantine_count": "H",
            "date_time": "C3",
            "unique_parts": "C4",
            "unique_suppliers": "C5",
            "location_percentages": "C6",
            "in_house_total": "E9",
            "cage_total": "F9",
            "rack_total": "G9",
            "quarantine_total": "H9"}


def make_sheet_master(combined: dict, worksheet_master) -> None:
    """
    Create sheet master for inventory output.
    :param combined:
    :param worksheet_master:
    :return:
    """
    parts: dict = combined["parts"]
    slots: dict = get_master_slots()

    add_master_main_info(combined, worksheet_master, slots)
    add_master_cells(parts, slots, worksheet_master)


def add_master_cells(parts: dict, slots: dict, worksheet_master):
    """
    Add inventory master cells for parts' descriptions and counts
    :param parts:
    :param slots:
    :param worksheet_master:
    :return:
    """
    start: int = slots["start"]
    slot_parts: str = slots["part"]
    slot_type: str = slots["type"]
    slot_supplier: str = slots["supplier"]
    slot_total_count: str = slots["total_count"]
    slot_cage_count: str = slots["cage_count"]
    slot_rack_count: str = slots["rack_count"]
    slot_quarantine_count: str = slots["quarantine_count"]

    for index, part_number in enumerate(parts, start=start):
        supplier: str = parts[part_number].get("supplier")
        count: int = parts[part_number]["count"]
        cage: int = parts[part_number]["cage"]
        rack: int = parts[part_number]["rack"]
        quarantine: int = parts[part_number]["quarantine"]

        worksheet_master[f"{slot_parts}{index}"].value = part_number
        worksheet_master[f"{slot_type}{index}"].value = ""
        worksheet_master[f"{slot_supplier}{index}"].value = supplier

        if count == 0:
            worksheet_master[f"{slot_total_count}{index}"].value = count
            worksheet_master[f"{slot_cage_count}{index}"].value = "-"
            worksheet_master[f"{slot_rack_count}{index}"].value = "-"
            worksheet_master[f"{slot_quarantine_count}{index}"].value = "-"

        else:
            worksheet_master[f"{slot_total_count}{index}"].value = count
            worksheet_master[f"{slot_cage_count}{index}"].value = cage
            worksheet_master[f"{slot_rack_count}{index}"].value = rack
            worksheet_master[f"{slot_quarantine_count}{index}"].value = quarantine

        set_cell_alignment(index, worksheet_master, slots)
        set_supplier_color(index, worksheet_master, supplier)


def add_master_main_info(combined: dict, worksheet_master: load_workbook, slots: dict) -> None:
    """
    Add to master sheet for inventory for basic information.  Ex. date / time, total, unique parts, etc.
    :param combined: SN + Transaction DB
    :param worksheet_master: inventory main sheet
    :param slots: predefined locations
    :return:
    """
    slot_date_time: str = slots["date_time"]
    slot_unique_parts: str = slots["unique_parts"]
    slot_unique_suppliers: str = slots["unique_suppliers"]
    slot_in_house: str = slots["in_house_total"]
    slot_cage: str = slots["cage_total"]
    slot_rack: str = slots["rack_total"]
    slot_quarantine: str = slots["quarantine_total"]

    unique_parts: list = get_unique_parts(combined)

    worksheet_master[slot_date_time].value = get_datetime_output()
    worksheet_master[slot_unique_parts].value = len(unique_parts)
    worksheet_master[slot_unique_suppliers].value = len(combined["stats"]["unique_suppliers"])
    worksheet_master[slot_in_house].value = combined["stats"]["in_house"]
    worksheet_master[slot_cage].value = combined["stats"]["cage"]
    worksheet_master[slot_rack].value = combined["stats"]["rack"]
    worksheet_master[slot_quarantine].value = combined["stats"]["quarantine"]


def main_method(basic_data: dict) -> None:
    """
    Extract data from Rich while adding / subtracting from database's transactions on plus or minus from Kirkland site.
    :return: None
    """
    workbook = create_workbook_for_report()
    worksheet_master = create_worksheet_master(workbook)
    # worksheet_transactions = create_transactions_worksheet(workbook)

    part_numbers_library: dict = get_part_numbers_library()
    serial_numbers_database: dict = get_serial_numbers_database(part_numbers_library)
    transactions_database: dict = get_transactions_from_database(part_numbers_library)

    pretty_print(transactions_database)

    combined: dict = combine_serial_with_transaction(serial_numbers_database, transactions_database)

    make_sheet_master(combined, worksheet_master)

    save_and_launch(workbook)
