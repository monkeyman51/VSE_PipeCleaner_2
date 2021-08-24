"""
Update form after request form is filled properly for inventory movement.
"""
import sys
from os import system
from time import strftime, sleep

import win32com.client as client
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from psutil import process_iter as task_manager, NoSuchProcess
from pipe_cleaner.src.log_database import access_database_document


def close_message(file_name: str) -> None:
    """
    Close message when excel closed.
    """
    print(f'\t\t- {file_name} closed')


def is_excel_file_running(file_name: str) -> bool:
    """
    Check if excel file is running through the task manager.
    """
    for application in task_manager():
        if 'EXCEL.EXE' in application.name().upper():
            try:
                for excel_file in application.as_dict()['cmdline']:
                    if file_name in excel_file:
                        return True

            except ProcessLookupError:
                close_message(file_name)
                return False
            except NoSuchProcess:
                close_message(file_name)
                return False
    else:
        close_message(file_name)
        return False


def open_excel_file_for_inventory_forms(file_name: str) -> None:
    """
    Automatically opens excel file.
    """
    system(f'start EXCEL.EXE {file_name}')


def is_excel_file_closed(file_name: str, form_type: str) -> bool:
    """
    Pipe Cleaner continues to check excel file runtime until it terminates by the user.
    """
    print(f'\t\t- {file_name} opened')
    print(f'\t\t- Fill out {form_type} form...\n')

    while True:
        if not is_excel_file_running(file_name):
            return True


def setup_update_form(request_form: dict) -> None:
    """
    Setup up the fields to have consistent look.
    """
    form_number: str = request_form['_id']

    date: str = request_form['basic_info']['date']
    end: str = request_form['basic_info']['end']
    location: str = request_form['basic_info']['location']
    quantity: str = request_form['basic_info']['quantity']
    seconds: str = request_form['basic_info']['seconds']
    start: str = request_form['basic_info']['start']
    time: str = request_form['basic_info']['time']
    user: str = request_form['basic_info']['user']
    version: str = request_form['basic_info']['version']

    excel_date: str = request_form['excel_data']['date']
    machine_name: str = request_form['excel_data']['machine_name']
    notes: str = request_form['excel_data']['notes']
    part_number: str = request_form['excel_data']['part_number']
    pipe_number: str = request_form['excel_data']['pipe_number']
    purchase_order: str = request_form['excel_data']['purchase_order']
    seconds: str = request_form['excel_data']['seconds']
    task_name: str = request_form['excel_data']['task_name']
    excel_time: str = request_form['excel_data']['time']
    trr_number: str = request_form['excel_data']['trr_number']

    update_workbook: load_workbook = load_workbook(f'settings/update_template.xlsx')
    update_worksheet = update_workbook['Update']

    update_worksheet['C13'] = 'Inventory Supervisor'
    update_worksheet['C16'] = 'Scan Material P/N'
    update_worksheet['C19'] = 'Scan Rack / Storage / Offsite'
    update_worksheet['C22'] = 'Scan Rack / Storage / Offsite'

    update_worksheet['D5'] = f'{start.upper()} TO {end.upper()} - UPDATE FORM'

    update_worksheet['C25'] = task_name
    update_worksheet['C28'] = notes
    update_worksheet['C29'] = pipe_number
    update_worksheet['C30'] = machine_name
    update_worksheet['C31'] = trr_number
    update_worksheet['C32'] = purchase_order

    update_worksheet['K2'] = excel_date
    update_worksheet['K3'] = excel_time
    update_worksheet['K4'] = user
    update_worksheet['K5'] = location
    update_worksheet['K6'] = form_number
    update_worksheet['K7'] = version
    update_worksheet['M10'] = quantity

    try:
        update_workbook.save(f'update_form.xlsx')

    except PermissionError:
        print(f'\tPermission Denied: update_template.xlsx is already open')
        print(f'\tPlease close file and restart Pipe Cleaner')
        input(f'\tPress enter to close:')
        sys.exit()


def start_update_form(form_number: str, basic_data: dict) -> None:
    """
    After sending email to inventory team. Have update form ready.
    """
    file_name: str = 'update_form.xlsx'
    site = str(basic_data['site']).replace(' Lab Site', '').replace('Kirkland', '0')

    current_year: str = strftime('%m/%d/%Y')[8:10]
    document = access_database_document('request_forms', f'{site}{current_year}')
    request_form: dict = document.find_one({'_id': form_number})

    setup_update_form(request_form)
    open_excel_file_for_inventory_forms(file_name)

    if is_excel_file_closed(file_name, 'update'):
        check_update_form(file_name, request_form, form_number)


def is_update_fields_filled(request_fields: dict) -> bool:
    """
    Checks to make sure the four essential fields are filled in for the request form.
    """
    is_quantity_match: bool = is_all_serial_numbers_scanned(request_fields)

    if is_field_correct(request_fields['approved_by']) and \
            is_field_correct(request_fields['part_number']) and \
            is_field_correct(request_fields['start_location']) and \
            is_field_correct(request_fields['end_location']) and \
            is_field_correct(request_fields['task_name']) and \
            is_quantity_match:
        return True

    else:
        return False


def print_line_divider(new_section_title: str) -> None:
    """
    Print terminal line divider.
    """
    print(f'\n\n\t{"-" * 60}\n\n')
    print(f'\t{new_section_title}')


def check_update_form(file_name: str, request_form: dict, form_number: str) -> None:
    """
    Check if inventory request form done correctly.
    """
    print(f'\t\t- Checking update form')

    fields_data: dict = get_update_fields(request_form, load_workbook(f'update_form.xlsx')['Update'])

    if is_update_fields_filled(fields_data):
        print_line_divider('Log Serial Numbers and Inventory Transaction:')
        print(f'\t\tY  ->  Yes')
        print(f'\t\tN  ->  No')
        response: str = input(f'\n\tResponse: ').upper()

        if 'Y' == response:
            update_serial_numbers(fields_data)
            update_transactions(fields_data, form_number)
            print(f'\n\tSuccess: Updates to Serial Numbers and Transactions database...')
            input(f'\tPress enter to exit: ')
            sys.exit()

        elif 'N' == response:
            print(f'\tNot updated to database.')
            input(f'\tPress enter to exit: ')
            sys.exit()

    else:
        print(f'\n-------------------------------------------------------------------')
        print(f'\n\t\t*** INVALID UPDATE FORM ***')
        respond_update_form_wrong(fields_data)
        print(f'-------------------------------------------------------------------')
        print(f'\n\t\t* Need to fix inventory update form...')
        input(f'\t\t- Press enter to open excel:')
        open_excel_file_for_inventory_forms(file_name)
        check_update_form(file_name, request_form, form_number)


def get_current_transactions_count(document) -> int:
    all_transactions: list = document.find({})

    count: int = 0
    for item in all_transactions:
        count += 1
    return count


def update_transactions(fields_data: dict, form_number: str) -> None:
    """
    Update serial numbers based off of update form's fields.
    """
    document = access_database_document('transactions', '021')
    transactions_count: int = get_current_transactions_count(document)
    new_count = str(transactions_count + 1)

    transaction_log: dict = get_transaction_log(fields_data, new_count, form_number)

    document.insert_one(transaction_log)


def get_transaction_log(fields_data: dict, _id: str, form_number: str) -> dict:
    """
    Data being submitted to database as new inventory transaction log.
    """
    transaction_log: dict = {"_id": str(_id),
                             "part_number": fields_data['part_number'],
                             "scanned": fields_data['current_quantity'],
                             "time": {},
                             "location": {},
                             "source": {}}

    transaction_log['time']['date_entry'] = fields_data['date']
    transaction_log['time']['time_entry'] = fields_data['time']
    transaction_log['time']['date_logged'] = strftime('%m/%d/%Y')
    transaction_log['time']['time_logged'] = strftime('%I:%M %p')

    transaction_log['location']['site'] = fields_data['site']
    transaction_log['location']['current'] = fields_data['end_location']
    transaction_log['location']['previous'] = fields_data['start_location']
    transaction_log['location']['rack'] = 'None'
    transaction_log['location']['machine'] = fields_data['machine_name']
    transaction_log['location']['pipe'] = fields_data['pipe_number']

    transaction_log['source']['approved_by'] = fields_data['approved_by']
    transaction_log['source']['verified_by'] = fields_data['user']
    transaction_log['source']['version'] = fields_data['version']
    transaction_log['source']['task'] = fields_data['task_name']
    transaction_log['source']['trr'] = fields_data['trr_number']
    transaction_log['source']['comment'] = fields_data['notes']
    transaction_log['source']['form_number'] = form_number

    return transaction_log


def get_update_form_data(fields_data: dict) -> dict:
    """
    Get data from update form to log in the cloud database.
    """
    serial_number_logs: dict = {}

    serial_numbers_entered: list = fields_data['current_quantity']

    for serial_number in serial_numbers_entered:
        serial_number_logs[serial_number]: dict = {}
        current: dict = serial_number_logs[serial_number]

        current['approved_by'] = fields_data['approved_by']
        current['date'] = fields_data['date']
        current['end_location'] = fields_data['end_location']
        current['form'] = fields_data['form']
        current['machine_name'] = fields_data['machine_name']
        current['notes'] = fields_data['notes']
        current['part_number'] = fields_data['part_number']
        current['pipe_number'] = fields_data['pipe_number']
        current['pipe_number'] = fields_data['pipe_number']
        current['purchase_order'] = fields_data['purchase_order']
        current['site'] = fields_data['site']
        current['start_location'] = fields_data['start_location']
        current['task_name'] = fields_data['task_name']
        current['time'] = fields_data['time']
        current['total_quantity'] = fields_data['total_quantity']
        current['trr_number'] = fields_data['trr_number']
        current['user'] = fields_data['user']
        current['version'] = fields_data['version']

    return serial_number_logs


def get_transaction_for_serial_number_log(fields_data: dict) -> dict:
    """
    Transaction for serial numbers.
    """
    return {"time": {"time_entry": fields_data['time'],
                     "date_entry": fields_data['date'],
                     "time_logged": strftime('%m/%d/%Y'),
                     "date_logged": strftime('%I:%M %p')},

            "location": {"site": fields_data['site'],
                         "current": fields_data[
                             'end_location'],
                         "previous": fields_data[
                             'start_location'],
                         "rack": "None",
                         "machine": fields_data[
                             'machine_name'],
                         "pipe": fields_data[
                             'pipe_number']},

            "source": {"approved_by": fields_data['approved_by'],
                       "verified_by": fields_data['user'],
                       "version": fields_data['version'],
                       "task": fields_data['task_name'],
                       "trr": fields_data['trr_number'],
                       "comment": fields_data['notes']}}


def update_serial_numbers(fields_data: dict) -> None:
    """
    Update serial numbers based off of update form's fields.
    """
    update_data: dict = get_update_form_data(fields_data)
    document = access_database_document('serial_numbers', 'all')

    for serial_number in update_data:

        db_serial_numbers: dict = document.find_one({'_id': serial_number})

        if not db_serial_numbers:
            entry: dict = get_new_serial_number_log(fields_data, serial_number)

            document.insert_one(entry)

        else:
            transaction: dict = get_transaction_for_serial_number_log(fields_data)
            document.update_one({"_id": serial_number},
                                {"$push": {"transactions": transaction}},
                                upsert=False)


def get_new_serial_number_log(fields_data: dict, _id: str) -> dict:
    """
    Data being submitted to database as new inventory transaction log.
    """
    transaction_log: dict = {"_id": str(_id),
                             "part_number": fields_data['part_number'],
                             "transactions": []}

    transaction: dict = get_transaction_for_serial_number_log(fields_data)
    transaction_log["transactions"].append(transaction)

    return transaction_log


def is_field_correct(field_input: str) -> bool:
    """
    Checks if valid input for a given field or is empty.
    """
    try:
        field_input: str = field_input.upper()

        if not field_input:
            return False

        elif field_input == '':
            return False

        elif field_input == 'COPY TASK TITLE':
            return False

        elif field_input == 'REQUIRED':
            return False

        elif field_input == 'COPY FROM TASK':
            return False

        elif field_input == 'RACK / STORAGE / OFFSITE':
            return False

        elif field_input == 'OPTIONAL':
            return False

        elif field_input == 'COPY FROM PACKAGE':
            return False

        elif field_input == 'INVENTORY SUPERVISOR':
            return False

        elif field_input == 'SCAN MATERIAL P/N':
            return False

        elif field_input == 'SCAN RACK / STORAGE / OFFSITE':
            return False

        else:
            return True

    except AttributeError:
        return False


def respond_update_form_wrong(update_fields: dict) -> None:
    """
    Message printed if update form is filled incorrectly.
    """
    bad_message: str = 'Invalid or Missing'

    if not is_field_correct(update_fields['approved_by']):
        print(f'\t\t\t- Approved By: {bad_message}')

    if not is_field_correct(update_fields['task_name']):
        print(f'\t\t\t- Task Name: {bad_message}')

    if not is_field_correct(update_fields['part_number']):
        print(f'\t\t\t- Part Number: {bad_message}')

    if not is_field_correct(update_fields['start_location']):
        print(f'\t\t\t- Start Location: {bad_message}')

    if not is_field_correct(update_fields['end_location']):
        print(f'\t\t\t- End Location: {bad_message}')

    if not is_all_serial_numbers_scanned(update_fields):
        print(f'\t\t\t- Serial Numbers Scanned: Not Enough Scanned')

    print(f'')


def is_all_serial_numbers_scanned(request_fields: dict) -> bool:
    """
    Are all the serial numbers scanned properly. Do they match the request with the total scanned?
    """
    total_quantity = int(request_fields['total_quantity'])
    current_quantity = int(len(request_fields['current_quantity']))

    if total_quantity == current_quantity:
        return True
    elif total_quantity != current_quantity:
        return False


def get_update_fields(request_form: dict, update_worksheet) -> dict:
    """
    Get update fields information.
    """
    quantity: str = request_form['basic_info']['quantity']
    machine_name: str = request_form['excel_data']['machine_name']
    notes: str = request_form['excel_data']['notes']
    pipe_number: str = request_form['excel_data']['pipe_number']
    purchase_order: str = request_form['excel_data']['purchase_order']
    task_name: str = request_form['excel_data']['task_name']
    trr_number: str = request_form['excel_data']['trr_number']

    return {'approved_by': update_worksheet['C13'].value,
            'part_number': update_worksheet['C16'].value,
            'start_location': update_worksheet['C19'].value,
            'end_location': update_worksheet['C22'].value,
            'task_name': task_name,
            'notes': clean_optional_field(notes),
            'pipe_number': clean_optional_field(pipe_number),
            'machine_name': clean_optional_field(machine_name),
            'trr_number': clean_optional_field(trr_number),
            'purchase_order': clean_optional_field(purchase_order),
            'current_quantity': get_serial_numbers(update_worksheet),
            'total_quantity': quantity,
            'date': update_worksheet['K2'].value,
            'time': update_worksheet['K3'].value,
            'user': update_worksheet['K4'].value,
            'site': update_worksheet['K5'].value,
            'form': update_worksheet['K6'].value,
            'version': update_worksheet['K7'].value}


def get_serial_numbers(worksheet: load_workbook) -> list:
    """
    Get serial numbers from Update Form sheet.
    """
    print(f'\t\t- Collecting Serial Numbers')
    serial_numbers: list = []
    for number in range(12, 1012):
        current_value: str = worksheet[f'H{number}'].value

        if not current_value or current_value == '' or current_value == 'Scan Here':
            pass
        else:
            serial_numbers.append(current_value)

    return serial_numbers


def clean_optional_field(field_input: str) -> str:
    """
    Get rid of optional or empty of optional fields.
    """
    upper_field_input = str(field_input).upper()

    if 'OPTIONAL' in upper_field_input:
        return 'None'
    elif upper_field_input == '':
        return 'None'
    else:
        return field_input