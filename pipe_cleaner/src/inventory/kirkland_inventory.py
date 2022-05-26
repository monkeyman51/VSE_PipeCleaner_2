"""
Kirkland inventory based on static, manually counted data plus transactions.
"""
from pipe_cleaner.src.log_database import access_database_document
from openpyxl import load_workbook
from os import system
from datetime import datetime
from openpyxl.styles import Alignment, Font, Side, Border


def get_clean_alternatives(alternative, alternatives, serial_number) -> list:
    """

    """
    clean_alternatives: list = [serial_number]
    for alt_name in alternatives:
        if alt_name != alternative:
            clean_alternatives.append(alt_name)

    return clean_alternatives


def access_serial_numbers_database() -> list:
    """
    Access and return serial numbers from Inventory database for testing and mocking.
    :return: List of transaction entries.
    """
    document = access_database_document('serial_numbers', 'base_line')
    return document.find({})


def is_same_serial_length(raw_serial_number: str, clean_serial: str) -> bool:
    """
    Temporary.  Checks if the original serial number is the same as the clean serial number.
    """
    if len(raw_serial_number) == len(clean_serial):
        return True
    elif len(raw_serial_number) != len(clean_serial):
        return False


def clean_serial_number(raw_serial_number: str, part_number: str) -> str:
    """

    """
    serial_number = str(raw_serial_number).upper().strip()

    if " " in serial_number:
        serial_1 = serial_number.split(" ")[0]
        serial_2 = serial_number.split(" ")[1]

        if part_number in serial_1:

            clean_serial: str = serial_number. \
                replace(f"{serial_1}_", ""). \
                replace(f"{serial_1}+", ""). \
                replace(f"{serial_1} ", "")

            return clean_serial.upper().strip()

        elif part_number in serial_2:
            clean_serial: str = serial_number. \
                replace(f"_{serial_2}", ""). \
                replace(f"+{serial_2}", ""). \
                replace(f" {serial_2}", "")

            return clean_serial.upper().strip()

    elif part_number in serial_number:
        clean_serial: str = serial_number. \
            replace(f"{part_number}_", ""). \
            replace(f"{part_number}+", ""). \
            replace(f"{part_number} ", "")

        return clean_serial.upper().strip()

    else:
        return str(serial_number).upper().strip()


def clean_part_name(part_name: str) -> str:
    """
    Erase any known characters that do not belong in the part name.

    :param part_name: raw part name
    :return: clean part name
    """
    clean_part: str = part_name.replace(":", "").upper().strip().split(" (")[0]

    return substitute_part_number(clean_part)


def get_serial_numbers_database():
    """
    Get serial numbers for all records of Kirkland.
    """
    document = access_database_document("serial_numbers", 'version_02')
    serial_numbers_database: list = document.find({})

    database: dict = {}
    for entry in serial_numbers_database:
        serial_number: str = entry["_id"].upper().strip()
        current_location: str = entry["to_locations"][-1].title().strip()
        previous_location: str = entry["from_locations"][-1].title().strip()
        date_logged: str = entry["dates"][-1].strip()
        part_number: str = clean_part_name(entry["part_numbers"][-1])

        clean_serial: str = clean_serial_number(serial_number, part_number)

        database[clean_serial]: dict = {}
        database[clean_serial]["current_location"]: str = current_location
        database[clean_serial]["previous_location"]: str = previous_location
        database[clean_serial]["date_logged"]: str = date_logged
        database[clean_serial]["part_number"]: str = part_number

    return database


def access_inventory_database() -> list:
    """
    Access and return all transaction entries from Inventory.
    :return: List of transaction entries.
    """
    document = access_database_document('transactions', '021')
    return document.find({})


def audit_transaction_date() -> list:
    """
    Lists of transaction dates to skip when iterating through all transactions.
    :return: dates to avoid.
    """
    return ["06/23/2021", "07/07/2021", "07/08/2021", "07/09/2021", "07/12/2021", "07/13/2021", "07/14/2021",
            "07/15/2021", "07/16/2021", "07/19/2021", "07/20/2021", "07/21/2021", "07/22/2021", "07/23/2021",
            "07/26/2021", "07/27/2021", "07/28/2021", "07/29/2021", "07/30/2021", "08/02/2021", "08/03/2021",
            "08/04/2021", "08/05/2021", "08/06/2021", "08/09/2021", "08/10/2021", "08/11/2021", "8/12/2021"]


def is_date_after(current_date: str) -> bool:
    """
    Iterates through dates that should not be included in transaction logs.
    :param current_date:
    :return: True / False
    """
    audit_dates: list = audit_transaction_date()

    for audit_date in audit_dates:
        if current_date in audit_date:
            return False
    else:
        return True


def get_transactions_from_database() -> list:
    """
    From inventory database
    :return:
    """
    transactions: list = []

    for current_entry in access_inventory_database():
        part_number: str = current_entry["part_number"]
        current_date: str = current_entry["time"]["date_logged"]

        if part_number.upper() != "TEST" and is_date_after(current_date):

            scanned: list = current_entry["scanned"]
            current_location: str = current_entry["location"]["current"]
            previous_location: str = current_entry["location"]["previous"]
            clean_part: str = clean_part_name(part_number)

            for serial_number in scanned:

                clean_serial: str = clean_serial_number(serial_number, part_number)

                current_serial: dict = {"part_number": clean_part,
                                        "serial_number": clean_serial,
                                        "current_location": current_location,
                                        "date_logged": current_date,
                                        "previous_location": previous_location}
                transactions.append(current_serial)

    return transactions


def set_supplier_color(index: int, worksheet, supplier: str):
    """
    Font cell in desired way.
    """
    if supplier == "Unknown" or supplier == "None" or supplier == "":
        worksheet[f"C{index}"].font = Font(color="7B7B7B")


def update_serial_history(base_serials: dict, transactions: list) -> dict:
    """

    """
    for serial_entry in transactions:

        serial_number: str = serial_entry["serial_number"]
        part_number: str = serial_entry["part_number"]
        current_location: str = serial_entry["current_location"]
        previous_location: str = serial_entry["previous_location"]
        date_logged: str = serial_entry["date_logged"]

        if serial_number not in base_serials:

            base_serials[serial_number]: dict = {}
            base_serials[serial_number]["date_logged"]: str = date_logged
            base_serials[serial_number]["part_number"]: str = part_number
            base_serials[serial_number]["current_location"]: str = current_location
            base_serials[serial_number]["previous_location"]: str = previous_location

        elif serial_number in base_serials:

            base_serials[serial_number]["date_logged"]: str = date_logged
            base_serials[serial_number]["part_number"]: str = part_number
            base_serials[serial_number]["current_location"]: str = current_location
            base_serials[serial_number]["previous_location"]: str = previous_location

    return base_serials


def output_excel(update_serials: dict) -> None:
    """

    :return:
    """
    workbook = load_workbook("settings/inventory/serials_history_template.xlsx")
    worksheet = workbook["Sheet1"]

    for index, serial_number in enumerate(update_serials, start=2):
        serial_entry: dict = update_serials[serial_number]
        part_number: str = serial_entry["part_number"]
        current_location: str = serial_entry["current_location"]
        previous_location: str = serial_entry["previous_location"]
        date_logged: str = serial_entry["date_logged"]

        worksheet[f"A{index}"].value = clean_serial_number(serial_number, part_number)
        worksheet[f"B{index}"].value = part_number
        worksheet[f"C{index}"].value = current_location
        worksheet[f"D{index}"].value = previous_location
        worksheet[f"E{index}"].value = date_logged

    workbook.save("serials_history.xlsx")
    system(fr'start EXCEL.EXE serials_history.xlsx')


def get_part_to_counts() -> dict:
    """
    Create part number counts for later data storage.
    """
    return {"parts": {},
            "overall": {"cage": 0,
                        "rack": 0,
                        "quarantine": 0,
                        "offsite": 0,
                        "exceptions": 0,
                        "onsite": 0}}


def sort_serials_into_categories(update_serials: dict) -> dict:
    """
    Organize serials into categories.  Pipe
    :param update_serials:
    :return:
    """
    part_to_counts: dict = get_part_to_counts()

    for unique_serial in update_serials:
        current_serial: dict = update_serials[unique_serial]
        part_number: str = current_serial["part_number"]
        current_location: str = current_serial["current_location"].upper()  # Upper for consistent comparison

        if part_number not in part_to_counts["parts"]:
            part_to_counts["parts"][part_number]: dict = {}
            serial_part_numbers: dict = part_to_counts["parts"][part_number]

            serial_part_numbers["cage"]: int = 0
            serial_part_numbers["rack"]: int = 0
            serial_part_numbers["quarantine"]: int = 0
            serial_part_numbers["offsite"]: int = 0
            serial_part_numbers["exceptions"]: int = 0
            serial_part_numbers["onsite"]: int = 0

            part_to_counts: dict = count_part_serials(current_location, part_number, part_to_counts)
            part_to_counts: dict = count_overall_serials(current_location, part_to_counts)

        elif part_number in part_to_counts["parts"]:
            part_to_counts: dict = count_part_serials(current_location, part_number, part_to_counts)
            part_to_counts: dict = count_overall_serials(current_location, part_to_counts)

    return part_to_counts


def count_overall_serials(current_location: str, part_to_counts: dict) -> dict:
    """
    Add count to overall part locations for overview of different categories.  These categories include Rack, Cage,
    Quarantine, Offsite, or Other (Exemptions)

    WARNING: count_part_serials should be aligned here.
    """
    overall_count: dict = part_to_counts["overall"]

    if "TURBO" in current_location and "CAT" in current_location:
        overall_count["rack"] += 1
        overall_count["onsite"] += 1

    elif "TURBOCATS" == current_location:
        overall_count["rack"] += 1
        overall_count["onsite"] += 1

    elif "QUARANTINE" == current_location or "QUAR" == current_location:
        overall_count["quarantine"] += 1
        overall_count["onsite"] += 1

    elif "CAGE" == current_location or "PICTURE" in current_location:
        overall_count["cage"] += 1
        overall_count["onsite"] += 1

    elif "RACK" == current_location or "PIPE" == current_location:
        overall_count["rack"] += 1
        overall_count["onsite"] += 1

    elif "CUSTOMER" == current_location or "OUT" == current_location or "SHIPMENT" == current_location:
        overall_count["offsite"] += 1

    else:
        overall_count["exceptions"] += 1

    return part_to_counts


def count_part_serials(current_location: str, part_number: str, part_to_counts: dict) -> dict:
    """
    Add count based on parts' state.  Will go to either Rack, Cage, Quarantine, Offsite, or Other (Exemptions)

    WARNING: count_overall_serials should be aligned here.
    """
    serials: dict = part_to_counts["parts"]

    if "TURBO" in current_location and "CAT" in current_location:
        serials[part_number]["rack"] += 1
        serials[part_number]["onsite"] += 1

    if "QUARANTINE" == current_location or "QUAR" == current_location:
        serials[part_number]["quarantine"] += 1
        serials[part_number]["onsite"] += 1

    elif "CAGE" == current_location or "PICTURE" in current_location:
        serials[part_number]["cage"] += 1
        serials[part_number]["onsite"] += 1

    elif "RACK" == current_location or "PIPE" == current_location:
        serials[part_number]["rack"] += 1
        serials[part_number]["onsite"] += 1

    elif "CUSTOMER" == current_location or "OUT" == current_location or "SHIPMENT" == current_location:
        serials[part_number]["offsite"] += 1

    else:
        serials[part_number]["exceptions"] += 1

    return part_to_counts


def get_unique_current_location(base_serial_numbers: dict) -> list:
    """
    Temp function to find unique list of current locations from static base serials database.
    """
    all_locations: list = []

    for unique_serial in base_serial_numbers:
        serial_data: dict = base_serial_numbers[unique_serial]
        current_location: str = serial_data["current_location"]

        all_locations.append(current_location)

    return list(set(all_locations))


def get_raw_date() -> str:
    """
    Get raw dates later later parsing and string manipulation.
    :return:
    """
    return datetime.now().strftime('%Y-%m-%d')


def get_relevant_names() -> dict:
    """
    Get names for relevant application later. file name, sheet names, etc.

    file_name: report name for excel
    template_file: master inventory template path
    master_sheet_name:
    transactions_sheet_name:

    :return: dictionary of file names
    """
    raw_date: str = get_raw_date()
    clean_year: str = get_raw_year(raw_date)[2:4]
    raw_month: str = get_raw_month(raw_date)
    raw_day: str = get_raw_day(raw_date)

    return {"file_name": f"inventory_{raw_month}-{raw_day}-{clean_year}.xlsx",
            "template_file_path": "settings/master_inventory_template.xlsx",
            "master_sheet_name": "Master - Internal",
            "transactions_sheet_name": "Transactions"}


def get_raw_year(raw_date: str) -> str:
    """
    Get current year for later string parsing. ex. 2021 / 2022
    """
    return raw_date[0:4]


def get_raw_day(raw_date: str) -> str:
    """
    Get raw day in terms of date
    """
    return raw_date[8:10]


def get_raw_month(raw_date: str) -> str:
    """
    Get month from raw date.
    :param raw_date: month, year, day
    :return: month
    """
    return raw_date[5:7]


def create_workbook_for_report() -> load_workbook:
    """
    Create workbook for report.  Used for saving excel report later.
    :return:
    """
    base_names: dict = get_relevant_names()

    template_file_path: str = base_names["template_file_path"]
    return load_workbook(template_file_path)


def create_worksheet_master(workbook: load_workbook) -> load_workbook:
    """
    Create worksheet for excel output.  Used for creating individual sheet.
    :param workbook: interaction with main workbook excel file
    :return: master_worksheet
    """
    base_names: dict = get_relevant_names()

    master_sheet_name: str = base_names["master_sheet_name"]
    return workbook[master_sheet_name]


def get_master_slots() -> dict:
    """
    Get master inventory columns for slotting the correct location.
    :return:
    """
    return {"start": 10,
            "part": "B",
            "type": "C",
            "supplier": "D",
            "total_count": "E",
            "cage_count": "F",
            "rack_count": "G",
            "quarantine_count": "H",
            "date_time": "C3",
            "unique_parts": "C4",
            "unique_suppliers": "C5",
            "location_percentages": "C6",
            "in_house_total": "E9",
            "cage_total": "F9",
            "rack_total": "G9",
            "quarantine_total": "H9"}


def access_database() -> list:
    """

    :return:
    """
    document_whole = access_database_document("part_numbers", "library")
    return document_whole.find({})


def access_database_part_number() -> dict:
    """
    Access database part number.
    :return:
    """
    part_numbers: list = access_database()

    library: dict = {}
    for document in part_numbers:
        main_part: str = document["_id"]
        supplier: str = document["supplier"]
        part_type: str = document["part_type"]
        alternatives: list = document["alternatives"]

        if main_part not in library:
            library[main_part]: dict = {}
            library[main_part]["supplier"]: str = supplier
            library[main_part]["part_type"]: str = part_type
            library[main_part]["alternatives"]: list = alternatives

    return library


def write_report(sorted_serials: dict, worksheet_master) -> None:
    """
    Write Serials History report.
    """
    part_number_library: dict = access_database_part_number()

    make_sheet_master(sorted_serials, worksheet_master, part_number_library)


def get_month_name(raw_month: str) -> str:
    """

    :param raw_month:
    :return:
    """
    if raw_month == "01":
        return "January"

    elif raw_month == "02":
        return "February"

    elif raw_month == "03":
        return "March"

    elif raw_month == "04":
        return "April"

    elif raw_month == "05":
        return "May"

    elif raw_month == "06":
        return "June"

    elif raw_month == "07":
        return "July"

    elif raw_month == "08":
        return "August"

    elif raw_month == "09":
        return "September"

    elif raw_month == "10":
        return "October"

    elif raw_month == "11":
        return "November"

    elif raw_month == "12":
        return "December"

    else:
        return "None"


def get_clean_day(raw_day: str) -> str:
    """

    :param raw_day:
    :return:
    """
    if raw_day[0] == "0" and len(raw_day) == 2:
        return raw_day[1]

    else:
        return raw_day


def get_current_date() -> str:
    """
    Get current date for logging later.
    :return:
    """
    raw_date: str = get_raw_date()
    raw_year: str = get_raw_year(raw_date)
    raw_month: str = get_raw_month(raw_date)
    raw_day = get_raw_day(raw_date)

    clean_month: str = get_month_name(raw_month)
    clean_day: str = get_clean_day(raw_day)

    return f"{clean_month} {clean_day}, {raw_year}"


def get_unique_parts(combined: dict) -> list:
    """
    Get unique part numbers based from combined data gathered from SN and transaction database.
    :param combined: SN and transaction database
    :return: list of unique part numbers
    """
    return list(combined["parts"].keys())


def get_current_time() -> str:
    """

    :return:
    """
    current_time: str = datetime.today().strftime("%I:%M %p")
    return current_time


def get_datetime_output() -> str:
    """
    Create default date and time format for excel report.
    :return: excel output date and time
    """
    current_date: str = get_current_date()
    current_time: str = get_current_time()
    return f"{current_date}  -  {current_time} PST"


def add_master_main_info(combined: dict, worksheet_master: load_workbook, slots: dict, all_data: dict) -> None:
    """
    Add to master sheet for inventory for basic information.  Ex. date / time, total, unique parts, etc.
    :param combined: SN + Transaction DB
    :param worksheet_master: inventory main sheet
    :param slots: predefined locations
    :param all_data:
    :return:
    """
    slot_date_time: str = slots["date_time"]
    slot_unique_parts: str = slots["unique_parts"]
    slot_unique_suppliers: str = slots["unique_suppliers"]
    slot_in_house: str = slots["in_house_total"]
    slot_cage: str = slots["cage_total"]
    slot_rack: str = slots["rack_total"]
    slot_quarantine: str = slots["quarantine_total"]

    # unique_parts: list = get_unique_parts(combined)

    worksheet_master[slot_date_time].value = get_datetime_output()
    # worksheet_master[slot_unique_parts].value = len(unique_parts)
    # worksheet_master[slot_unique_suppliers].value = len(combined["stats"]["unique_suppliers"])
    worksheet_master[slot_in_house].value = all_data["in_house"]
    worksheet_master[slot_cage].value = all_data["cage_total"]
    worksheet_master[slot_rack].value = all_data["rack_total"]
    worksheet_master[slot_quarantine].value = all_data["quarantine_total"]


def set_cell_alignment(index: int, worksheet, slots: dict):
    """
    Align cell in desired way.
    :param index:
    :param worksheet:
    :param slots:
    :return:
    """
    worksheet[f"{slots['part']}{index}"].alignment = Alignment(horizontal='left')
    worksheet[f"{slots['type']}{index}"].alignment = Alignment(horizontal='left')
    worksheet[f"{slots['supplier']}{index}"].alignment = Alignment(horizontal='left')
    worksheet[f"{slots['total_count']}{index}"].alignment = Alignment(horizontal='center')
    worksheet[f"{slots['cage_count']}{index}"].alignment = Alignment(horizontal='center')
    worksheet[f"{slots['rack_count']}{index}"].alignment = Alignment(horizontal='center')
    worksheet[f"{slots['quarantine_count']}{index}"].alignment = Alignment(horizontal='center')

    # For whatever reason, this cell is not taking left border manually.  Did this via code.
    worksheet[f"{slots['type']}{index}"].border = Border(left=Side(border_style='thin', color='A5A5A5'))
    worksheet[f"{slots['quarantine_count']}{index}"].border = Border(left=Side(border_style='thin', color='A5A5A5'))


def add_master_cells(parts: dict, slots: dict, worksheet_master, part_number_library: dict):
    """
    Add inventory master cells for parts' descriptions and counts
    :param parts:
    :param slots:
    :param worksheet_master:
    :param part_number_library:
    :return:
    """
    start: int = slots["start"]
    slot_parts: str = slots["part"]
    slot_type: str = slots["type"]
    slot_supplier: str = slots["supplier"]
    slot_total_count: str = slots["total_count"]
    slot_cage_count: str = slots["cage_count"]
    slot_rack_count: str = slots["rack_count"]
    slot_quarantine_count: str = slots["quarantine_count"]

    all_data: dict = {"in_house": 0,
                      "cage_total": 0,
                      "rack_total": 0,
                      "quarantine_total": 0}

    for index, part_number in enumerate(sorted(parts.keys()), start=start):
        # count: int = parts[part_number]["count"]
        cage_count: int = parts[part_number]["cage"]
        rack_count: int = parts[part_number]["rack"]
        quarantine_count: int = parts[part_number]["quarantine"]
        total_count: int = cage_count + rack_count + quarantine_count

        worksheet_master[f"{slot_parts}{index}"].value = part_number
        # worksheet_master[f"{slot_type}{index}"].value = ""
        # worksheet_master[f"{slot_supplier}{index}"].value = supplier

        part_attributes: dict = part_number_library.get(part_number, {})

        if part_attributes != {}:
            part_type: str = part_attributes["part_type"]
            supplier: str = part_attributes["supplier"]

            worksheet_master[f"{slot_type}{index}"].value = part_type
            worksheet_master[f"{slot_supplier}{index}"].value = supplier


        if total_count == 0:
            worksheet_master[f"{slot_total_count}{index}"].value = total_count
            worksheet_master[f"{slot_cage_count}{index}"].value = "-"
            worksheet_master[f"{slot_rack_count}{index}"].value = "-"
            worksheet_master[f"{slot_quarantine_count}{index}"].value = "-"

        else:
            worksheet_master[f"{slot_total_count}{index}"].value = total_count
            worksheet_master[f"{slot_cage_count}{index}"].value = cage_count
            worksheet_master[f"{slot_rack_count}{index}"].value = rack_count
            worksheet_master[f"{slot_quarantine_count}{index}"].value = quarantine_count

        set_cell_alignment(index, worksheet_master, slots)
        # set_supplier_color(index, worksheet_master, supplier)

        all_data["in_house"] += total_count
        all_data["cage_total"] += cage_count
        all_data["rack_total"] += rack_count
        all_data["quarantine_total"] += quarantine_count

    return all_data


def make_sheet_master(combined: dict, worksheet_master, part_number_library) -> None:
    """
    Create sheet master for inventory output.
    :param combined:
    :param worksheet_master:
    :param part_number_library:
    :return:
    """
    parts: dict = combined["parts"]
    slots: dict = get_master_slots()

    all_data: dict = add_master_cells(parts, slots, worksheet_master, part_number_library)
    add_master_main_info(combined, worksheet_master, slots, all_data)


def save_and_launch(workbook) -> None:
    """
    Save output based off of the template.
    :param workbook: inventory template
    :return: None
    """
    base_names: dict = get_relevant_names()

    file_name: str = base_names["file_name"]
    workbook.save(file_name)
    launch_output_after_save()


def launch_output_after_save() -> None:
    """
    After creating excel report for inventory, launch the excel file for user to automatically see.
    :return: None
    """
    base_names: dict = get_relevant_names()

    file_name: str = base_names["file_name"]
    system(fr'start EXCEL.EXE {file_name}')


def substitute_part_number(part_number: str) -> str:
    """
    Based off of Inventory Team's feedback, substitutes invalid P/N for real P/N

    :return: Data
    """
    if "KRM393A4K40DB2-CVF2030" == part_number:
        return "M393A4K40DB2-CVF"

    elif "HMA84GR7CJR4N-VKTNAD943" == part_number:
        return "HMA84GR7CJR4N-VK"

    elif "KRM393A4G40AB3-CWE2007" == part_number:
        return "M393A4G40AB3-CWE"

    elif "KRM393A4G40AB3-CWE2028" == part_number:
        return "M393A4G40AB3-CWE"

    elif "HMAA8GR7AJR4N-WMT4AD030" == part_number:
        return "HMAA8GR7AJR4N-WM"

    elif "HMA84GR7JJR4N-VKTFAC936" == part_number:
        return "HMA84GR7JJR4N-VK"

    elif "HMA84GR7CJR4N-VKTFAC904" == part_number:
        return "HMA84GR7CJR4N-VK"

    elif "HMA84GR7CJR4N-VKT3AD944" == part_number:
        return "HMA84GR7CJR4N-VK"

    elif "HMAA8GR7AJR4N-WMT4AC927" == part_number:
        return "HMAA8GR7AJR4N-WMT"

    elif "MTA36ASF4G72PZ-3G2R1112" == part_number:
        return "MTA36ASF4G72PZ-3G2R1"

    elif "HMA84GR7CJR4N-VKT3AD943" == part_number:
        return "HMA84GR7CJR4N-VK"

    elif "HMA82GR7CJR4N-VKT3AD930" == part_number:
        return "HMA82GR7CJR4N-VK"

    elif "HMA84GR7DJR4N-WMT4AC946" == part_number:
        return "HMA84GR7DJR4N-WM"

    elif "MTA18ASF4G72PZ-2G9E1945" == part_number:
        return "MTA18ASF4G72PZ-2G9E"

    elif "HMAA8GR7AJR4N-WMT8AC929" == part_number:
        return "HMAA8GR7AJR4N-WM"

    elif "KRM393A4K40DB2-CVF2001" == part_number:
        return "M393A4K40DB2-CVF"

    elif "KRM393A4K40CB2-CTD1902" == part_number:
        return "M393A4K40CB2-CTD"

    elif "KRM393A4K40CB2-CTD1902" == part_number:
        return "M393A4K40CB2-CTD"

    elif "HMA84GR7CJR4N-XNTGAA018" == part_number:
        return "HMA84GR7CJR4N-XN"

    elif "KRM393A4K40CB2-CTD1903" == part_number:
        return "M393A4K40CB2-CTD"

    elif "HMA84GR7CJR4N-VKT3AC905" == part_number:
        return "HMA84GR7CJR4N-VK"

    elif "HMA84GR7CJR4N-VKTNAC744" == part_number:
        return "HMA84GR7CJR4N-VK"

    elif "MTA36ASF4G72PZ-2G6E1819" == part_number:
        return "MTA36ASF4G72PZ-2G6E"

    elif "HMA84GR7CJR4N-VKT3AC743" == part_number:
        return "HMA84GR7CJR4N-VK"

    # Model Number -> Part Number
    elif "ST18000NM019J" == part_number:
        return "3AY212-401"

    # Model Number -> Part Number
    elif "WUH721816ALN6L6" == part_number:
        return "0F38420"

    elif "MG08SCA16TA" == part_number:
        return "HDEPN20SMA51F"

    elif "ST14000NM000G-2KG103" == part_number:
        return "2KG103-401"

    elif "ST12000NM007G-2RM102" == part_number:
        return "2RM102-402"

    elif "WUH721414ALN600" == part_number:
        return "0F31114"

    elif "HMAA4GR7AJHR4N-XN" == part_number:
        return "HMAA4GR7AJR4N-XN"

    elif "SSDPELKX019T8M2" == part_number:
        return "SSDPELKX019T8DM2"

    elif "SSDPELKX960G8D-203" == part_number:
        return "SSDPELKX960G8D"

    elif "ST16000NM000G-2KH103" == part_number:
        return "2KH103-402"

    elif "M393A2K43BB1CTD" == part_number:
        return "M393A2K43BB1-CTD"

    elif "MG07ACA12TA" == part_number:
        return "HDEPW21SMA51"

    elif "PE8110NVME1TB" == part_number:
        return "HFS960GDE0X098N"

    elif "ST14000NM021J-2TX112" == part_number:
        return "2TX112-401"

    elif "ST18000NM018J-3B1112" == part_number:
        return "3B1112-402"

    elif "ST6000NM0115-1YZ110" == part_number:
        return "1YZ110-003"

    elif "ST6000NM037A-2MQ101" == part_number:
        return "2MQ101-402"

    elif "WUH721414AL" == part_number:
        return "0F31114"

    elif "WUH721816AL4206" == part_number:
        return "0F38313"

    elif "WUH721818ALN6L6" == part_number:
        return "0F38423"

    elif "WUS4BB038D4MBE7" == part_number:
        return "0TS2381"

    elif "HFS960GDE0X089N" == part_number:
        return "HFS960GDE0X098N"

    elif "SSDPFXNV153TZEM" == part_number:
        return "SSDPFXNV153TZD"

    else:
        return part_number


def reverse_transactions(all_transactions: list) -> list:
    """

    """
    order: list = []

    for entry in all_transactions:
        order.append(entry)

    new_order: list = order[::-1]

    return new_order


def add_transaction_data(workbook) -> None:
    """
    Add inventory transaction data for report tab.
    """
    document = access_database_document('transactions', '021')
    worksheet = workbook["Transactions"]

    all_transactions: list = document.find({})
    transactions: list = reverse_transactions(all_transactions)

    for index, transaction_log in enumerate(transactions, start=9):

        date: str = transaction_log["time"]["date_logged"]
        previous: str = transaction_log["location"]["previous"]
        current: str = transaction_log["location"]["current"]
        part_number: str = transaction_log["part_number"]
        quantity = len(transaction_log["scanned"])
        task: str = transaction_log["source"]["task"]

        worksheet[f'B{index}'].value = date
        worksheet[f'C{index}'].value = previous
        worksheet[f'D{index}'].value = current
        worksheet[f'E{index}'].value = part_number
        worksheet[f'F{index}'].value = quantity
        worksheet[f'G{index}'].value = task


def main_method() -> None:
    """

    """
    base_serial_numbers: dict = get_serial_numbers_database()
    transactions: list = get_transactions_from_database()
    update_serials: dict = update_serial_history(base_serial_numbers, transactions)

    sorted_serials: dict = sort_serials_into_categories(update_serials)

    workbook = create_workbook_for_report()
    worksheet_master = create_worksheet_master(workbook)
    add_transaction_data(workbook)
    write_report(sorted_serials, worksheet_master)

    save_and_launch(workbook)

    # output_excel(update_serials)
