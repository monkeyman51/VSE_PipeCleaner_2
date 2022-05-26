"""
Store and edit part numbers that are alternative versions of each other from MongoDB.

Sometimes when comparing a serial number's part number, there could be variants of the same part number.  For instance,
ABC and ABC-1 hypothetically are the same part number but stored in the database differently.  This could arise from
scanning the wrong bar code to logging the wrong data manually.  Either way, there needs to be a thesaurus of
the same part number to allow adding / subtracting / updating correctly for a total inventory report.
"""
import psutil

from pipe_cleaner.src.log_database import access_database_document
from openpyxl import load_workbook
from openpyxl.styles import Side, Border
from datetime import datetime
import os
from psutil import process_iter as task_manager, NoSuchProcess


def access_database() -> list:
    """

    :return:
    """
    document_whole = access_database_document("part_numbers", "library")
    return document_whole.find({})


def access_database_part_number() -> dict:
    """
    Access database part number.
    :return:
    """
    part_numbers: list = access_database()

    library: dict = {}
    for document in part_numbers:
        main_part: str = document["_id"]
        supplier: str = document["supplier"]
        part_type: str = document["part_type"]
        model_number: str = document["model_number"]
        description: str = document["description"]
        alternatives: list = document["alternatives"]

        if main_part not in library:
            library[main_part]: dict = {}
            library[main_part]["supplier"]: str = supplier
            library[main_part]["part_type"]: str = part_type
            library[main_part]["model_number"]: str = model_number
            library[main_part]["description"]: str = description
            library[main_part]["alternatives"]: list = alternatives

    return library


def get_library_model(alternative: str, part_number: str, supplier_name: str) -> dict:
    """
    Structure of database entries for each part number.
    :return:
    """
    return {"_id": part_number,
            "main_part": supplier_name,
            "supplier": supplier_name,
            "alternatives": []}


def get_current_date() -> str:
    raw_date: str = datetime.now().strftime('%Y-%m-%d')
    raw_year: str = raw_date[0:4]
    raw_month: str = raw_date[5:7]
    raw_day: str = raw_date[8:10]

    return f"{raw_month}/{raw_day}/{raw_year}"


def get_excel_data(worksheet) -> dict:
    """
    Grab data from excel.
    :return:
    """
    library_data: dict = {}

    for row_number in range(9, 2_000):
        if worksheet[f"B{row_number}"].value is None:
            break

        main_part: str = worksheet[f"B{row_number}"].value
        part_type: str = worksheet[f"E{row_number}"].value
        supplier: str = worksheet[f"F{row_number}"].value
        model_number: str = worksheet[f"C{row_number}"].value
        description: str = worksheet[f"D{row_number}"].value
        alternatives: list = get_alternatives(row_number, worksheet)

        library_data[main_part]: dict = {}
        library_data[main_part]["part_type"]: str = part_type
        library_data[main_part]["supplier"]: str = supplier
        library_data[main_part]["model_number"]: str = model_number
        library_data[main_part]["description"]: str = description
        library_data[main_part]["alternatives"]: list = alternatives

    return library_data


def get_alternatives(row_number: int, worksheet) -> list:
    """
    Get alternatives shown in the excel output
    :param row_number:
    :param worksheet:
    :return:
    """
    alternatives: list = []
    alternative_04: str = worksheet[f"G{row_number}"].value
    alternative_05: str = worksheet[f"H{row_number}"].value
    alternative_06: str = worksheet[f"I{row_number}"].value
    alternative_07: str = worksheet[f"J{row_number}"].value
    alternative_08: str = worksheet[f"K{row_number}"].value
    alternative_09: str = worksheet[f"L{row_number}"].value
    alternative_10: str = worksheet[f"M{row_number}"].value

    alternatives.append(alternative_04)
    alternatives.append(alternative_05)
    alternatives.append(alternative_06)
    alternatives.append(alternative_07)
    alternatives.append(alternative_08)
    alternatives.append(alternative_09)
    alternatives.append(alternative_10)

    real_alternatives: list = []
    for alternative in alternatives:
        if alternative == "None" or alternative is None:
            pass
        else:
            real_alternatives.append(str(alternative))

    return real_alternatives


def fill_excel_output(access_database: list) -> None:
    """"""
    workbook = load_workbook("settings/inventory/serials_database_output.xlsx")
    worksheet = workbook["Part Numbers - Library"]

    pass


def launch_output_after_save() -> None:
    """
    After creating excel report for inventory, launch the excel file for user to automatically see.
    :return: None
    """
    os.system(fr'start EXCEL.EXE library/part_numbers_library.xlsx')


def save_excel_output(workbook) -> None:
    """
    Save output based off of the template.
    :param workbook: inventory template
    :return: None
    """
    try:
        workbook.save("library/part_numbers_library.xlsx")

    except PermissionError:
        print(f"\n\tpart_numbers_library.xlsx is already opened.  Close file then try again.")


def get_alphabet() -> list:
    """

    :return:
    """
    return ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S"]


def create_excel(worksheet, library_data: dict) -> None:
    """

    """
    for index, part_number in enumerate(sorted(library_data.keys()), start=9):
        supplier: str = library_data[part_number]["supplier"]
        part_type: str = library_data[part_number]["part_type"]
        model_number: str = library_data[part_number]["model_number"]
        description: str = library_data[part_number].get("description", "")
        alternatives: list = library_data[part_number]["alternatives"]

        worksheet[f"B{index}"].value = part_number
        worksheet[f"C{index}"].value = model_number
        worksheet[f"D{index}"].value = description
        worksheet[f"E{index}"].value = part_type
        worksheet[f"F{index}"].value = supplier

        worksheet[f"C{index}"].border = Border(left=Side(border_style='thin', color='A5A5A5'))
        worksheet[f"D{index}"].border = Border(left=Side(border_style='thin', color='A5A5A5'))
        worksheet[f"E{index}"].border = Border(left=Side(border_style='thin', color='A5A5A5'))
        worksheet[f"F{index}"].border = Border(left=Side(border_style='thin', color='A5A5A5'))

        # for number, alternative in enumerate(alternatives, start=4):
        #     alphabet: list = get_alphabet()
        #     worksheet[f"{alphabet[number]}{index}"].value = alternative


def replace_database_with_excel_data(database_library: dict, excel_data: dict) -> dict:
    """
    
    :param database_library: 
    :param excel_data: 
    :return: 
    """
    for part_number in excel_data:
        database_library[part_number]["supplier"]: str = excel_data[part_number]["supplier"]
        database_library[part_number]["alternatives"]: str = excel_data[part_number]["alternatives"]

    return database_library


def is_excel_file_opened(file_name: str) -> bool:
    """
    Assure that excel file opened.  Limits attempt to see if excel file opened after the 100th attempt.
    """
    count: int = 0

    while count <= 100:
        for application in task_manager():
            application_name: str = application.name().upper()

            if "EXCEL.EXE" == application_name:

                try:
                    current_excel_files: dict = application.as_dict()["cmdline"]
                    for excel_file in current_excel_files:

                        if file_name in excel_file:
                            return True

                except NoSuchProcess:
                    pass

        count += 1
    else:
        return False


def is_excel_file_done(file_name: str) -> bool:
    """
    Assure that excel file closed.  Limit on attempt is unbounded.
    """
    count: int = 0
    while count == 0:
        for application in task_manager():
            application_name: str = application.name().upper()

            if "EXCEL.EXE" == application_name:

                try:
                    current_excel_files: list = application.as_dict()["cmdline"]

                    clean_files: list = []
                    for excel_file in current_excel_files:
                        clean_file: str = excel_file.replace(r"library/", "")
                        clean_files.append(clean_file)

                    if file_name not in clean_files and r"/Embedding" not in clean_files:
                        count = 1

                except NoSuchProcess:
                    pass

            else:
                count = 1

    if count == 1:
        return True


def is_excel_file_running(file_name: str) -> bool:
    """
    Check if excel file is running through the task manager.
    """
    if is_excel_file_opened(file_name):
        if is_excel_file_done(file_name):
            input(f"\tPress Enter to Save: ")
            return False


def close_message(file_name: str) -> None:
    """
    Close message when excel closed.
    """
    print(f'\t\t- {file_name} closed')


def is_excel_file_closed(file_name: str, form_type: str) -> bool:
    """
    Pipe Cleaner continues to check excel file runtime until it terminates by the user.
    """
    print(f'\t\t- {file_name} opened')
    print(f'\t\t- Fill out {form_type} form...\n')

    if not is_excel_file_running(file_name):
        return True


def get_new_worksheet():
    workbook = load_workbook("library/part_numbers_library.xlsx")
    worksheet = workbook["Part Numbers - Library"]
    return worksheet


def start_request_stage() -> dict:
    """
    Start the request stage before notifying inventory team
    """
    file_closed: bool = is_excel_file_closed("part_numbers_library.xlsx", "part_numbers")
    print(f"\tFile Closed: {file_closed}")

    #  Checks file closed twice to ensure file closure.  Sometimes falsely closes.
    if file_closed is True:
        new_worksheet = get_new_worksheet()
        return get_excel_data(new_worksheet)


def update_database(replacement: dict) -> dict:
    """

    :param replacement:
    :return:
    """
    collection = access_database_document("part_numbers", "library")

    for part_number in replacement:
        supplier: dict = replacement[part_number]["supplier"]
        part_type: dict = replacement[part_number]["part_type"]
        model_number: dict = replacement[part_number]["model_number"]
        description: dict = replacement[part_number]["description"]
        alternatives: list = replacement[part_number]["alternatives"]

        collection.update_one({"_id": part_number},
                              {"$set": {"supplier": supplier,
                                        "part_type": part_type,
                                        "model_number": model_number,
                                        "description": description,
                                        "alternatives": alternatives}},
                              upsert=True)


def main_method(basic_data: dict) -> dict:
    """
    Main method starts here.
    :return:
    """
    workbook = load_workbook("settings/inventory/part_numbers_library_template.xlsx")
    worksheet = workbook["Part Numbers - Library"]

    database_library: dict = access_database_part_number()

    create_excel(worksheet, database_library)

    save_excel_output(workbook)
    launch_output_after_save()

    excel_data: dict = start_request_stage()

    update_database(excel_data)
