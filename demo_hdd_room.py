"""
Attempt to upload hard drive room logging.
"""
from openpyxl import load_workbook
from pipe_cleaner.src.log_database import access_database_document
from time import strftime


def clean_hdd_room_data() -> list:
    """
    Clean data as to upload to database successfully.
    """
    serial_numbers: list = []

    worksheet = load_workbook(fr'settings/hard_drive_room.xlsx')['Master']

    for index, row in enumerate(range(1, 5113), start=2):
        document: dict = {"_id": clean_input(worksheet[f"F{index}"].value),
                          "part_number": clean_input(worksheet[f"E{index}"].value),
                          "transactions": []}

        transaction: dict = set_transactions(index, worksheet)

        document['transactions'].append(transaction)
        serial_numbers.append(document)

    return serial_numbers


def set_transactions(index: int, worksheet: load_workbook) -> dict:
    """
    Set times, locations, and sources for information about transaction.
    """
    transaction: dict = {}
    transaction: dict = set_transaction_times(index, transaction, worksheet)
    transaction: dict = set_transaction_locations(index, transaction, worksheet)
    transaction: dict = set_transaction_sources(index, transaction, worksheet)

    return transaction


def set_transaction_times(index: int, transaction: dict, worksheet: load_workbook) -> dict:
    """
    Set times in transactions.
    """
    transaction["time"]: dict = {"time_entry": "None",
                                 "date_entry": clean_input(worksheet[f"A{index}"].value),
                                 "time_logged": strftime('%m/%d/%Y'),
                                 "date_logged": strftime('%I:%M %p')}

    return transaction


def set_transaction_sources(index: int, transaction: dict, worksheet: load_workbook) -> dict:
    """
    Set sources in transactions.
    """
    transaction["source"]: dict = {"approved_by": "Daniel Andersen",
                                   "verified_by": clean_input(worksheet[f"B{index}"].value),
                                   "version": "2.6.7",
                                   "task": "None",
                                   "trr": "None",
                                   "comment": "None"}

    return transaction


def set_transaction_locations(index: int, transaction: dict, worksheet: load_workbook) -> dict:
    """
    Set locations in transactions.
    """
    transaction["location"]: dict = {"site": "Kirkland, WA",
                                     "current": clean_input(worksheet[f"G{index}"].value),
                                     "previous": "None",
                                     "rack": "None",
                                     "machine": "None",
                                     "pipe": "None"}

    return transaction


def clean_input(data_input: str) -> str:
    """

    """
    return str(data_input).replace('null', 'None')


def main_method() -> None:
    """
    Access Database Document.
    """
    hdd_data: list = clean_hdd_room_data()
    document = access_database_document('serial_numbers', 'all')
    # find_all = document.find({})

    # count = 0
    # for item in find_all:
    #     count += 1
    # print(f'count: {count}')
    # input()

    for entry in hdd_data:
        serial_number: str = entry["_id"]

        db_serial_numbers: dict = document.find_one({'_id': serial_number})
        if not db_serial_numbers:
            document.insert_one(entry)

        else:
            transaction: dict = entry["transactions"][0]
            document.update_one({"_id": serial_number},
                                {"$push": {"transactions": transaction}},
                                upsert=False)


main_method()
