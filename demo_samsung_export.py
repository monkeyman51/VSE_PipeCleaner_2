"""
Gather data from Samsung export excel sheet to database stored in cloud.
"""
from pipe_cleaner.src.log_database import access_database_document
from openpyxl import load_workbook
from time import strftime


def get_part_numbers_from_excel(worksheet: load_workbook):
    """
    Fetch part numbers data from excel sheet.
    """
    part_numbers: list = []
    for index in range(2, 174):
        part_number_data: dict = {'part_number': worksheet[f'A{index}'].value,
                                  'ddr': worksheet[f'B{index}'].value,
                                  'type': worksheet[f'C{index}'].value,
                                  'size': worksheet[f'D{index}'].value,
                                  'rank': worksheet[f'E{index}'].value,
                                  'speed': worksheet[f'F{index}'].value,
                                  'voltage': worksheet[f'G{index}'].value}
        part_numbers.append(part_number_data)

    return part_numbers


def add_new_part_number_database(part_numbers_data: dict, part_number: str) -> dict:
    """
    Set standard schema for part numbers from excel sheet for cloud database.
    """
    new_entry: dict = {'_id': part_number,
                       'sources': []}
    return add_new_part_number_source(new_entry, part_numbers_data)


def add_new_part_number_source(new_entry: dict, part_numbers_data: dict) -> dict:
    """
    
    """
    new_source: dict = {'source': 'Samsung - Website Export',
                        'type': part_numbers_data['type'],
                        'size': part_numbers_data['size'],
                        'speed': part_numbers_data['speed'],
                        'ddr': part_numbers_data['ddr'],
                        'voltage': part_numbers_data['voltage'],
                        'friendly_name': 'None',
                        'model_number': 'None',
                        'supplier': 'Samsung',
                        'manufacturer': 'None',
                        'description': 'None',
                        'comment': 'None',
                        'time_logged': strftime('%I:%M %p'),
                        'date_logged': strftime('%m/%d/%Y')}

    new_entry['sources'].append(new_source)

    return new_entry


def main_method() -> None:
    """
    Access Database Document.
    """
    workbook = load_workbook(fr'settings/samsung_export.xlsx')
    worksheet = workbook['part_numbers']

    xlsx_part_numbers: list = get_part_numbers_from_excel(worksheet)

    document = access_database_document("part_numbers", "all")

    for part_number_data in xlsx_part_numbers:
        part_number: str = part_number_data['part_number']
        db_serial_numbers: dict = document.find_one({'_id': part_number})

        if not db_serial_numbers:
            new_entry: dict = add_new_part_number_database(part_number_data, part_number)
            document.insert_one(new_entry)
        else:
            new_source = add_new_part_number_database(part_number_data, part_number)

            document.update_one({"_id": part_number},
                                {"$push": {"sources": new_source}},
                                upsert=False)


main_method()
